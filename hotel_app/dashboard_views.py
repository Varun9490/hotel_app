import json
import datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.http import JsonResponse
from django.contrib.auth.models import User, Group
from django.db.models import Count, Avg, Q
from django.db.models.functions import TruncHour
from django.core.exceptions import PermissionDenied
from django.utils import timezone
from django.http import JsonResponse, HttpResponseBadRequest, HttpResponse
from django.views.decorators.http import require_http_methods
from django.db import connection, transaction
from django.conf import settings
from django.utils.text import slugify
from django.views.decorators.csrf import csrf_protect, csrf_exempt
from django.contrib.auth import get_user_model
import os

import os

import os
import logging

# Import all models from hotel_app
from hotel_app.models import (
    Department, Location, RequestType, Checklist,
    Complaint, BreakfastVoucher, Review, Guest,
    Voucher, VoucherScan, ServiceRequest, UserProfile, UserGroup, UserGroupMembership,
    Notification, GymMember, SLAConfiguration, DepartmentRequestSLA  # Add SLAConfiguration and DepartmentRequestSLA models
)

# Import all forms from the local forms.py
from .forms import (
    UserForm, DepartmentForm, GroupForm, LocationForm,
    RequestTypeForm, ChecklistForm, ComplaintForm,
    BreakfastVoucherForm, ReviewForm, VoucherForm, GymMemberForm
)

# Import local utils and services
from .utils import user_in_group, create_notification
from hotel_app.whatsapp_service import WhatsAppService
from .rbac_services import get_accessible_sections, can_access_section

# Import export/import utilities
from .export_import_utils import create_export_file, import_all_data, validate_import_data

# ---- Constants ----
ADMINS_GROUP = 'Admins'
STAFF_GROUP = 'Staff'
USERS_GROUP = 'Users'

User = get_user_model()

# Roles are not stored in DB. They are labels -> permission flags.
ROLES = ["Admins", "Staff", "Users"]

def _role_to_flags(role: str):
    r = (role or "").strip().lower()
    if r in ("admin", "admins", "administrator", "superuser"):
        return True, True
    if r in ("staff", "front desk", "front desk team"):
        return True, False
    # default user
    return False, False


# ---- Helper Functions ----
def is_admin(user):
    """Check if a user is an admin or superuser."""
    return user.is_superuser or user_in_group(user, ADMINS_GROUP)

def is_staff(user):
    """Check if a user is staff, admin, or superuser."""
    return (user.is_superuser or
            user_in_group(user, ADMINS_GROUP) or
            user_in_group(user, STAFF_GROUP))

@login_required
@user_passes_test(is_staff)
def dashboard(request):
    """Main dashboard view."""


@require_http_methods(['GET'])
def api_manage_users_filters(request):
    departments = list(Department.objects.order_by('name').values_list('name', flat=True))
    roles = ROLES  # not from database
    
    # Get all users with their profile information
    users = User.objects.select_related('userprofile').all()
    
    # Format users data
    formatted_users = []
    unassigned_users = []
    
    for user in users:
        # Get user profile if it exists
        profile = getattr(user, 'userprofile', None)
        
        # Format user data
        user_data = {
            'id': user.id,
            'username': user.username,
            'email': user.email,
            'full_name': getattr(profile, 'full_name', None) or f"{user.first_name} {user.last_name}".strip() or user.username,
            'avatar_url': getattr(profile, 'avatar_url', None),
            'role': getattr(profile, 'title', None) or 'Staff'
        }
        
        formatted_users.append(user_data)
        
        # Check if user is unassigned (no profile or no department in profile)
        if not profile or not getattr(profile, 'department', None):
            unassigned_users.append({
                **user_data,
                'department': 'Unassigned'
            })
    
    return JsonResponse({
        "departments": departments, 
        "roles": roles,
        "users": formatted_users,
        "unassigned_users": unassigned_users
    })
    # Get system status
    system_statuses = [
        {'name': 'Camera Health', 'value': '98%', 'color': 'green-500'},
        {'name': 'Import/Export Activity', 'value': 'Active', 'color': 'gray-900'},
        {'name': 'Billing Status', 'value': 'Current', 'color': 'green-500'},
    ]

    # Get checklist data
    checklists = [
        {'name': 'Housekeeping', 'completed': 18, 'total': 20, 'status_color': 'green', 'percentage': 90},
        {'name': 'Maintenance', 'completed': 12, 'total': 15, 'status_color': 'yellow', 'percentage': 80},
    ]

    # Get WhatsApp campaigns
    whatsapp_campaigns = [
        {'name': 'Welcome Message', 'time': '10:00 AM', 'status_color': 'green-500'},
        {'name': 'Checkout Reminder', 'time': '2:00 PM', 'status_color': 'yellow-400'},
        {'name': 'Feedback Request', 'time': '6:00 PM', 'status_color': 'sky-600'},
    ]

    # Get requests data for chart
    requests_data = {
        'labels': ['Housekeeping', 'Maintenance', 'Concierge', 'F&B', 'IT Support', 'Other'],
        'values': [95, 75, 45, 35, 25, 15]
    }

    # Get feedback data for chart
    feedback_data = {
        'labels': ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun'],
        'positive': [70, 80, 75, 90, 85, 95, 100],
        'neutral': [20, 25, 30, 25, 35, 30, 25],
        'negative': [15, 10, 20, 15, 12, 10, 8]
    }

    context = {
        'system_statuses': system_statuses,
        'checklists': checklists,
        'whatsapp_campaigns': whatsapp_campaigns,
        'requests_data': json.dumps(requests_data),
        'feedback_data': json.dumps(feedback_data),
    }
    
    return render(request, 'dashboard/dashboard.html', context)

def require_permission(group_names):
    """Decorator to require specific group permissions for a view."""
    if not isinstance(group_names, (list, tuple)):
        group_names = [group_names]
    
    def decorator(view_func):
        @login_required
        def wrapper(request, *args, **kwargs):
            if request.user.is_superuser or any(user_in_group(request.user, group) for group in group_names):
                return view_func(request, *args, **kwargs)
            raise PermissionDenied("You don't have permission to access this page.")
        return wrapper
    return decorator


def require_role(roles):
    """Decorator to require specific roles for a view."""
    if not isinstance(roles, (list, tuple)):
        roles = [roles]
    
    def decorator(view_func):
        @login_required
        def wrapper(request, *args, **kwargs):
            # Check if user has the required role
            if hasattr(request.user, 'userprofile'):
                user_role = request.user.userprofile.role
                if user_role in roles or request.user.is_superuser:
                    return view_func(request, *args, **kwargs)
            
            # Fallback to group-based permissions for backward compatibility
            if request.user.is_superuser or any(user_in_group(request.user, role) for role in roles):
                return view_func(request, *args, **kwargs)
                
            raise PermissionDenied("You don't have permission to access this page.")
        return wrapper
    return decorator


# ---- Notification Examples ----
# These are examples of how notifications would be created in real scenarios

def create_voucher_notification(voucher):
    """Create a notification when a voucher is issued"""
    # Create notification for the staff member who issued the voucher
    if voucher.issued_by:
        create_notification(
            recipient=voucher.issued_by,
            title="Voucher Issued",
            message=f"Voucher for {voucher.guest_name} has been issued successfully.",
            notification_type="voucher"
        )
    
    # Create notification for the guest (if we have a user account for them)
    # This would typically be implemented when guests have user accounts

def create_voucher_scan_notification(voucher, scanned_by):
    """Create a notification when a voucher is scanned"""
    # Create notification for the staff member who scanned the voucher
    create_notification(
        recipient=scanned_by,
        title="Voucher Scanned",
        message=f"Voucher for {voucher.guest_name} has been scanned successfully.",
        notification_type="voucher"
    )
    
    # Create notification for the staff member who issued the voucher
    if voucher.issued_by and voucher.issued_by != scanned_by:
        create_notification(
            recipient=voucher.issued_by,
            title="Voucher Redeemed",
            message=f"Voucher for {voucher.guest_name} has been redeemed.",
            notification_type="voucher"
        )

def create_service_request_notification(service_request):
    """Create a notification when a service request is created"""
    # Create notification for the department head
    if service_request.department and service_request.department.head:
        create_notification(
            recipient=service_request.department.head,
            title="New Service Request",
            message=f"A new service request has been submitted: {service_request.title}",
            notification_type="request"
        )
    
    # Create notification for the requester
    if service_request.requester:
        create_notification(
            recipient=service_request.requester,
            title="Service Request Submitted",
            message=f"Your service request '{service_request.title}' has been submitted successfully.",
            notification_type="request"
        )

# ---- Dashboard Home ----
@login_required
@require_role(['admin', 'staff', 'user'])
def dashboard_main(request):
    """Main dashboard view with key metrics."""
    total_users = User.objects.count()
    total_departments = Department.objects.count()
    total_locations = Location.objects.count()
    active_complaints = Complaint.objects.filter(status="pending").count()
    resolved_complaints = Complaint.objects.filter(status="resolved").count()
    vouchers_issued = Voucher.objects.count()
    vouchers_redeemed = Voucher.objects.filter(status="redeemed").count()
    average_review_rating = Review.objects.aggregate(Avg("rating"))["rating__avg"] or 0
    complaint_trends = Complaint.objects.values("status").annotate(count=Count("id"))

    context = {
        "total_users": total_users,
        "total_departments": total_departments,
        "total_locations": total_locations,
        "active_complaints": active_complaints,
        "resolved_complaints": resolved_complaints,
        "vouchers_issued": vouchers_issued,
        "vouchers_redeemed": vouchers_redeemed,
        "average_review_rating": f"{average_review_rating:.2f}",
        "complaint_trends": list(complaint_trends),
    }
    return render(request, "dashboard/main.html", context)


from django.contrib.auth.decorators import login_required


@login_required
def dashboard_view(request):
    """Render the dashboard with live metrics for users, departments and open complaints.

    - total_users: count of User objects
    - total_departments: count of Department objects
    - open_complaints: count of Complaint objects with pending/open status
    """
    today = timezone.localdate()

    # Live counts (defensive)
    try:
        total_users = User.objects.count()
    except Exception:
        total_users = 0

    try:
        total_departments = Department.objects.count()
    except Exception:
        total_departments = 0

    try:
        total_locations = Location.objects.count()
    except Exception:
        total_locations = 0

    try:
        open_complaints = Complaint.objects.filter(status__in=["pending", "in_progress"]).count()
    except Exception:
        try:
            open_complaints = Complaint.objects.count()
        except Exception:
            open_complaints = 0

    try:
        resolved_complaints = Complaint.objects.filter(status="resolved").count()
    except Exception:
        resolved_complaints = 0

    # Vouchers
    try:
        vouchers_issued = Voucher.objects.count()
        vouchers_redeemed = Voucher.objects.filter(status="redeemed").count()
        vouchers_expired = Voucher.objects.filter(status="expired").count()
    except Exception:
        vouchers_issued = vouchers_redeemed = vouchers_expired = 0

    # Reviews
    try:
        average_review_rating = Review.objects.aggregate(avg=Avg("rating"))["avg"] or 0
    except Exception:
        average_review_rating = 0

    # Complaint trends for charting
    try:
        complaint_trends = list(Complaint.objects.values("status").annotate(count=Count("id")))
    except Exception:
        complaint_trends = []

    # Requests chart data (try to derive from RequestType + ServiceRequest if available)
    try:
        request_types = list(RequestType.objects.all())
        requests_labels = [rt.name for rt in request_types]
        try:
            from hotel_app.models import ServiceRequest
            requests_values = [ServiceRequest.objects.filter(request_type=rt).count() for rt in request_types]
        except Exception:
            requests_values = [1 for _ in requests_labels]
    except Exception:
        requests_labels = ['Housekeeping', 'Maintenance', 'Concierge', 'F&B', 'IT Support', 'Other']
        requests_values = [95, 75, 45, 35, 25, 15]

    requests_data = {
        'labels': requests_labels,
        'values': requests_values,
    }

    # Feedback chart data (7-day buckets using Review if possible)
    try:
        labels = []
        positive = []
        neutral = []
        negative = []
        for i in range(6, -1, -1):
            day = today - datetime.timedelta(days=i)
            labels.append(day.strftime('%a'))
            reviews_on_day = Review.objects.filter(created_at__date=day)
            positive.append(reviews_on_day.filter(rating__gte=4).count())
            neutral.append(reviews_on_day.filter(rating=3).count())
            negative.append(reviews_on_day.filter(rating__lte=2).count())
        feedback_data = {
            'labels': labels,
            'positive': positive,
            'neutral': neutral,
            'negative': negative,
        }
    except Exception:
        feedback_data = {
            'labels': ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun'],
            'positive': [70, 80, 75, 90, 85, 95, 100],
            'neutral': [20, 25, 30, 25, 35, 30, 25],
            'negative': [15, 10, 20, 15, 12, 10, 8],
        }

    # Occupancy
    try:
        # Prefer datetime fields if set, otherwise use date fields
        occupancy_qs = Guest.objects.filter(
            Q(checkin_date__lte=today, checkout_date__gte=today) |
            Q(checkin_datetime__date__lte=today, checkout_datetime__date__gte=today)
        )
        occupancy_today = occupancy_qs.count()
        occupancy_rate = float(occupancy_today) / max(1, total_locations) * 100 if total_locations else 0
    except Exception:
        occupancy_today = 0
        occupancy_rate = 0

    occupancy_data = {'occupied': occupancy_today, 'rate': round(occupancy_rate, 1)}

    # DEBUG-only: seed minimal demo data if site is empty so dashboard looks functional locally
    try:
        if getattr(settings, 'DEBUG', False) and (total_users == 0 or Guest.objects.count() == 0):
            # Create demo department
            demo_dept, _ = Department.objects.get_or_create(name='Demo Department')

            # Create a demo user
            try:
                demo_user = User.objects.create_user(username='demo_user', password='password123', email='demo@example.com')
            except Exception:
                # If user exists or cannot be created, fetch any existing user
                demo_user = User.objects.first()

            # Create building/floor/location
            from hotel_app.models import Building, Floor, LocationType, LocationFamily, Booking
            building, _ = Building.objects.get_or_create(name='Main Building')
            floor, _ = Floor.objects.get_or_create(building=building, floor_number=1)
            ltype, _ = LocationType.objects.get_or_create(name='Guest Room')
            lfamily, _ = LocationFamily.objects.get_or_create(name='Rooms')
            location, _ = Location.objects.get_or_create(building=building, floor=floor, room_no='101', defaults={'name': 'Room 101', 'type': ltype, 'family': lfamily, 'capacity': 2})

            # Create demo guest
            guest, _ = Guest.objects.get_or_create(full_name='Demo Guest', defaults={
                'email': 'guest@example.com',
                'room_number': '101',
                'checkin_date': today - datetime.timedelta(days=1),
                'checkout_date': today + datetime.timedelta(days=1),
            })

            # Booking
            try:
                booking, _ = Booking.objects.get_or_create(guest=guest, room_number='101', defaults={'check_in': timezone.now() - datetime.timedelta(days=1), 'check_out': timezone.now() + datetime.timedelta(days=1)})
            except Exception:
                booking = None

            # Voucher
            try:
                if booking:
                    Voucher.objects.get_or_create(booking=booking, guest=guest, defaults={'guest_name': guest.full_name, 'room_number': '101', 'check_in_date': guest.checkin_date, 'check_out_date': guest.checkout_date, 'status': 'active', 'quantity': 1})
                else:
                    Voucher.objects.get_or_create(guest=guest, defaults={'guest_name': guest.full_name, 'room_number': '101', 'check_in_date': guest.checkin_date, 'check_out_date': guest.checkout_date, 'status': 'active', 'quantity': 1})
            except Exception:
                pass

            # Complaint
            try:
                Complaint.objects.get_or_create(subject='Demo complaint', defaults={'description': 'This is a demo complaint', 'status': 'pending'})
            except Exception:
                pass

            # Review
            try:
                Review.objects.get_or_create(guest=guest, defaults={'rating': 4, 'comment': 'Demo review'})
            except Exception:
                pass

            # Recompute counts
            total_users = User.objects.count()
            total_departments = Department.objects.count()
            total_locations = Location.objects.count()
            vouchers_issued = Voucher.objects.count()
            vouchers_redeemed = Voucher.objects.filter(status='redeemed').count()
            open_complaints = Complaint.objects.filter(status__in=['pending', 'in_progress']).count()
            average_review_rating = Review.objects.aggregate(avg=Avg('rating'))['avg'] or 0
    except Exception:
        # Do not let seeding errors break the dashboard
        pass

    context = {
        'total_users': total_users,
        'total_departments': total_departments,
        'total_locations': total_locations,
        'open_complaints': open_complaints,
        'resolved_complaints': resolved_complaints,
        'vouchers_issued': vouchers_issued,
        'vouchers_redeemed': vouchers_redeemed,
        'vouchers_expired': vouchers_expired,
        'average_review_rating': round(average_review_rating, 1) if average_review_rating else 0,
        'complaint_trends': json.dumps(complaint_trends),
        'requests_data': json.dumps(requests_data),
        'feedback_data': json.dumps(feedback_data),
        'occupancy_data': json.dumps(occupancy_data),
    }

    # The project now uses the new dashboard2 design as the primary dashboard.
    # Reuse the existing dashboard2_view to render the latest dashboard template
    # and context so we keep a single source of truth for the dashboard output.
    try:
        # Call dashboard2_view directly and return its response. It prepares
        # a design-oriented context. If dashboard2_view raises, fall back to
        # rendering the legacy dashboard template.
        return dashboard2_view(request)
    except Exception:
        return render(request, 'dashboard/dashboard.html', context)


@login_required
def dashboard2_view(request):
    """Render the new dashboard2 with the provided design using dynamic data."""
    from django.db.models import Count, Avg, Q
    from django.utils import timezone
    from django.contrib.auth import get_user_model
    import json
    import datetime
    
    User = get_user_model()
    today = timezone.localdate()

    # Live counts (defensive)
    try:
        total_users = User.objects.count()
    except Exception:
        total_users = 0

    try:
        total_departments = Department.objects.count()
    except Exception:
        total_departments = 0

    try:
        total_locations = Location.objects.count()
    except Exception:
        total_locations = 0

    try:
        open_complaints = Complaint.objects.filter(status__in=["pending", "in_progress"]).count()
    except Exception:
        try:
            open_complaints = Complaint.objects.count()
        except Exception:
            open_complaints = 0

    try:
        resolved_complaints = Complaint.objects.filter(status="resolved").count()
    except Exception:
        resolved_complaints = 0

    # Vouchers
    try:
        vouchers_issued = Voucher.objects.count()
        vouchers_redeemed = Voucher.objects.filter(status="redeemed").count()
        vouchers_expired = Voucher.objects.filter(status="expired").count()
    except Exception:
        vouchers_issued = vouchers_redeemed = vouchers_expired = 0

    # Reviews
    try:
        average_review_rating = Review.objects.aggregate(avg=Avg("rating"))["avg"] or 0
    except Exception:
        average_review_rating = 0

    # Complaint trends for charting
    try:
        complaint_trends = list(Complaint.objects.values("status").annotate(count=Count("id")))
    except Exception:
        complaint_trends = []

    # Requests chart data (try to derive from RequestType + ServiceRequest if available)
    try:
        request_types = list(RequestType.objects.all())
        requests_labels = [rt.name for rt in request_types]
        try:
            requests_values = [ServiceRequest.objects.filter(request_type=rt).count() for rt in request_types]
        except Exception:
            requests_values = [1 for _ in requests_labels]
    except Exception:
        requests_labels = ['Housekeeping', 'Maintenance', 'Concierge', 'F&B', 'IT Support', 'Other']
        requests_values = [95, 75, 45, 35, 25, 15]

    requests_data = {
        'labels': requests_labels,
        'values': requests_values,
    }

    # Feedback chart data (7-day buckets using Review if possible)
    try:
        labels = []
        positive = []
        neutral = []
        negative = []
        for i in range(6, -1, -1):
            day = today - datetime.timedelta(days=i)
            labels.append(day.strftime('%a'))
            reviews_on_day = Review.objects.filter(created_at__date=day)
            positive.append(reviews_on_day.filter(rating__gte=4).count())
            neutral.append(reviews_on_day.filter(rating=3).count())
            negative.append(reviews_on_day.filter(rating__lte=2).count())
        feedback_data = {
            'labels': labels,
            'positive': positive,
            'neutral': neutral,
            'negative': negative,
        }
    except Exception:
        feedback_data = {
            'labels': ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun'],
            'positive': [70, 80, 75, 90, 85, 95, 100],
            'neutral': [20, 25, 30, 25, 35, 30, 25],
            'negative': [15, 10, 20, 15, 12, 10, 8],
        }

    # Occupancy
    try:
        # Prefer datetime fields if set, otherwise use date fields
        occupancy_qs = Guest.objects.filter(
            Q(checkin_date__lte=today, checkout_date__gte=today) |
            Q(checkin_datetime__date__lte=today, checkout_datetime__date__gte=today)
        )
        occupancy_today = occupancy_qs.count()
        occupancy_rate = float(occupancy_today) / max(1, total_locations) * 100 if total_locations else 0
    except Exception:
        occupancy_today = 0
        occupancy_rate = 0

    occupancy_data = {'occupied': occupancy_today, 'rate': round(occupancy_rate, 1)}

    # Fetch actual critical tickets (high priority service requests)
    try:
        critical_tickets = ServiceRequest.objects.filter(
            priority__in=['high', 'critical']
        ).select_related('requester_user', 'department', 'location').order_by('-created_at')[:4]
        
        # Process tickets for display
        critical_tickets_data = []
        for ticket in critical_tickets:
            # Calculate time left based on SLA
            time_left = "Unknown"
            progress = 0
            if ticket.due_at and ticket.created_at:
                total_time = (ticket.due_at - ticket.created_at).total_seconds()
                elapsed_time = (timezone.now() - ticket.created_at).total_seconds()
                if total_time > 0:
                    progress = min(100, max(0, int((elapsed_time / total_time) * 100)))
                    remaining_seconds = total_time - elapsed_time
                    if remaining_seconds > 0:
                        hours = int(remaining_seconds // 3600)
                        if hours > 0:
                            time_left = f"{hours}h left"
                        else:
                            minutes = int(remaining_seconds // 60)
                            time_left = f"{minutes}m left"
                    else:
                        time_left = "Overdue"
                else:
                    time_left = "Completed"
                    progress = 100
            
            critical_tickets_data.append({
                'id': ticket.id,
                'title': ticket.request_type.name if ticket.request_type else 'Unknown Request',
                'location': str(ticket.location) if ticket.location else '',
                'department': str(ticket.department) if ticket.department else 'Unknown Department',
                'requester_user': ticket.requester_user,
                'reported': ticket.created_at.strftime('%Y-%m-%d %H:%M') if ticket.created_at else 'Unknown',
                'priority': ticket.priority.upper() if ticket.priority else 'NORM',
                'time_left': time_left,
                'progress': progress,
                'created_at': ticket.created_at,
                'completed_at': ticket.completed_at,
                'status': ticket.status,
            })
    except Exception as e:
        # Fallback to dummy data if there's an error
        critical_tickets_data = [
            {
                'id': 2847,
                'title': 'Room AC not working',
                'room': '304',
                'department': 'Maintenance',
                'guest': 'John Smith',
                'reported': '2 hours ago',
                'priority': 'HIGH',
                'time_left': '2h left',
                'color': 'red-500',
                'progress': 80
            },
            {
                'id': 2846,
                'title': 'Extra towels needed',
                'room': '218',
                'department': 'Housekeeping',
                'guest': 'Sarah Johnson',
                'reported': '45 minutes ago',
                'priority': 'MED',
                'time_left': '6h left',
                'color': 'yellow-400',
                'progress': 25
            },
            {
                'id': 2845,
                'title': 'Restaurant reservation',
                'room': '',
                'department': 'Guest Services',
                'guest': 'Michael Brown',
                'reported': '20 minutes ago',
                'priority': 'NORM',
                'time_left': '18h left',
                'color': 'sky-600',
                'progress': 10
            },
            {
                'id': 2844,
                'title': 'WiFi password reset',
                'room': '156',
                'department': 'IT Support',
                'guest': 'Emily Davis',
                'reported': 'Resolved: 15 minutes ago',
                'priority': 'RESOLVED',
                'time_left': 'Completed',
                'color': 'green-500',
                'progress': 100
            }
        ]

    # Fetch actual guest feedback (recent reviews)
    try:
        recent_feedback = Review.objects.select_related('guest').order_by('-created_at')[:3]
        
        # Process feedback for display
        feedback_data_list = []
        for feedback in recent_feedback:
            # Determine sentiment based on rating
            if feedback.rating >= 4:
                sentiment = 'POSITIVE'
            elif feedback.rating == 3:
                sentiment = 'NEUTRAL'
            else:
                sentiment = 'NEGATIVE'
                
            feedback_data_list.append({
                'id': feedback.id,
                'rating': feedback.rating,
                'location': getattr(feedback.guest, 'room_number', '') if feedback.guest else '',
                'created_at': feedback.created_at,
                'comment': feedback.comment or 'No comment provided',
                'guest': feedback.guest,
                'sentiment': sentiment,
            })
    except Exception as e:
        # Fallback to dummy data if there's an error
        feedback_data_list = [
            {
                'rating': 5,
                'location': 'Room 405',
                'time': '2 min ago',
                'comment': 'Amazing service and spotless room! The staff went above and beyond.',
                'guest': 'Amanda Wilson',
                'sentiment': 'Positive',
                'color': 'green-500'
            },
            {
                'rating': 3,
                'location': 'Restaurant',
                'time': '15 min ago',
                'comment': 'Food quality was good but service was slower than expected.',
                'guest': 'Robert Chen',
                'sentiment': 'Neutral',
                'color': 'yellow-400'
            },
            {
                'rating': 2,
                'location': 'Room 201',
                'time': '1 hour ago',
                'comment': 'Room was not properly cleaned upon arrival. Bathroom had issues.',
                'guest': 'Lisa Martinez',
                'sentiment': 'Negative',
                'color': 'red-500'
            }
        ]

    # Map the data to dashboard2 template variables
    context = {
        'user_name': request.user.get_full_name() or request.user.username,
        # Stats data
        'active_tickets': open_complaints,
        'avg_review_rating': round(average_review_rating, 1) if average_review_rating else 0,
        'sla_breaches': 0,  # This would need to be calculated from actual SLA breaches
        'vouchers_redeemed': vouchers_redeemed,
        'guest_satisfaction': round(average_review_rating * 20) if average_review_rating else 0,  # Convert 5-star to 100%
        'avg_response_time': '12m',  # This would need to be calculated from actual response times
        'staff_efficiency': 87,  # This would need to be calculated from actual metrics
        'active_gym_members': 389,  # This would need to be fetched from actual data
        'active_guests': occupancy_today,
        # Chart data - simplified for dashboard2
        'tickets_data': requests_values[:7] if len(requests_values) >= 7 else requests_values + [0] * (7 - len(requests_values)),
        'feedback_data': feedback_data['positive'][:7] if len(feedback_data['positive']) >= 7 else feedback_data['positive'] + [0] * (7 - len(feedback_data['positive'])),
        'peak_day_tickets': max(requests_values) if requests_values else 0,
        'peak_day_feedback': max(feedback_data['positive']) if feedback_data['positive'] else 0,
        'weekly_growth': 18,  # This would need to be calculated from actual data
        # Sentiment data
        'positive_reviews': sum(feedback_data['positive']),
        'neutral_reviews': sum(feedback_data['neutral']),
        'negative_reviews': sum(feedback_data['negative']),
        'positive_count': sum(feedback_data['positive']),
        'neutral_count': sum(feedback_data['neutral']),
        'negative_count': sum(feedback_data['negative']),
        # Department data - using real data where possible
        'departments': [
            {'name': 'Housekeeping', 'tickets': requests_values[0] if len(requests_values) > 0 else 15, 'color': 'sky-600'},
            {'name': 'Maintenance', 'tickets': requests_values[1] if len(requests_values) > 1 else 12, 'color': 'yellow-400'},
            {'name': 'Guest Services', 'tickets': requests_values[2] if len(requests_values) > 2 else 8, 'color': 'teal-500'},
            {'name': 'Restaurant', 'tickets': requests_values[3] if len(requests_values) > 3 else 6, 'color': 'green-500'},
            {'name': 'Front Desk', 'tickets': requests_values[4] if len(requests_values) > 4 else 4, 'color': 'fuchsia-700'},
            {'name': 'Concierge', 'tickets': requests_values[5] if len(requests_values) > 5 else 2, 'color': 'red-500'},
        ],
        # Critical tickets - now using actual data
        'critical_tickets': critical_tickets_data,
        # Guest feedback - now using actual data
        'guest_feedback': feedback_data_list
    }
    
    return render(request, 'dashboard/dashboard.html', context)


@login_required
@require_role(['admin', 'staff'])
def manage_users(request):
    """Render the Manage Users / User Groups screen on the right panel.

    Provides lightweight metrics and a groups list when available. Falls back
    to sensible dummy values if some models/fields are missing.
    """
    User = None
    try:
        from django.contrib.auth import get_user_model
        User = get_user_model()
    except Exception:
        User = None

    # Basic safe metrics
    total_groups = 0
    total_group_members = 0
    recent_additions = 0
    active_groups = 0
    groups = None

    try:
        # Try optional models that may exist in this project
        from hotel_app.models import UserGroup, UserGroupMembership
        # Annotate member counts if possible
        try:
            groups_qs = UserGroup.objects.all()
            # Attempt to annotate members_count if a related_name exists
            try:
                groups = groups_qs.annotate(members_count=Count('members'))
            except Exception:
                groups = groups_qs

            total_groups = groups_qs.count()
            # Sum members_count if available
            try:
                total_group_members = sum(getattr(g, 'members_count', 0) for g in groups)
            except Exception:
                # fallback to membership count
                try:
                    total_group_members = UserGroupMembership.objects.count()
                except Exception:
                    total_group_members = 0

            # recent additions if created_at exists
            try:
                recent_additions = groups_qs.filter(created_at__gte=timezone.now()-datetime.timedelta(days=7)).count()
            except Exception:
                recent_additions = 0

            # active groups (if boolean field exists)
            try:
                active_groups = UserGroup.objects.filter(active=True).count()
            except Exception:
                active_groups = 0

        except Exception:
            total_groups = 0
            total_group_members = 0
            groups = None
    except Exception:
        # If models are absent, keep defaults and let template show fallback content
        total_groups = 0
        total_group_members = 0
        recent_additions = 0
        active_groups = 0
        groups = None

    # total users (best-effort)
    total_users = 0
    try:
        if User is not None:
            total_users = User.objects.count()
    except Exception:
        total_users = 0

    context = {
        'total_users': total_users,
        'total_groups': total_groups,
        'total_group_members': total_group_members,
        'recent_additions': recent_additions,
        'active_groups': active_groups,
        'groups': groups,
    }

    # Previously this view returned only the component fragment which lacks the
    # full page head and CSS. Redirect to the full Manage Users page which
    # renders `dashboard/users.html` (see `manage_users_all`). This ensures the
    # head block (fonts, Tailwind, scripts) is included when opening
    # `/dashboard/manage-users/`.
    return redirect('dashboard:manage_users_all')


@require_permission([ADMINS_GROUP, STAFF_GROUP])
def messaging_setup(request):
    """Messaging Setup main screen. Provides templates, stats and connection info.

    This view is defensive: it falls back to sensible mock data when models or
    services are not available so templates can render during front-end work.
    """
    # Templates list (mock/fallback)
    templates = []
    try:
        # If a Template model exists, prefer real data
        from hotel_app.models import MessageTemplate
        templates_qs = MessageTemplate.objects.all().order_by('-updated_at')[:20]
        templates = [
            {
                'id': t.id,
                'name': getattr(t, 'name', 'Template'),
                'preview': getattr(t, 'preview', '') or getattr(t, 'body', '')[:120],
                'updated_at': getattr(t, 'updated_at', None),
            }
            for t in templates_qs
        ]
    except Exception:
        # Provide sample templates for UI
        templates = [
            {'id': 1, 'name': 'Welcome Message', 'preview': 'Hi {{guest_name}}, welcome to our hotel!', 'updated_at': None},
            {'id': 2, 'name': 'Checkout Reminder', 'preview': 'Reminder: Your checkout is at 12:00 PM today.', 'updated_at': None},
            {'id': 3, 'name': 'Post-Stay Review', 'preview': 'Thanks for staying with us — please share feedback.', 'updated_at': None},
        ]

    # Stats (mock/fallback)
    stats = {
        'connected': False,
        'messages_sent_7d': 0,
        'open_templates': len(templates),
    }
    try:
        # If WhatsAppService provides connection info, use it
        service = WhatsAppService()
        stats['connected'] = service.is_connected()
        stats['messages_sent_7d'] = service.messages_sent(days=7)
    except Exception:
        # keep fallback values
        stats.setdefault('connected', False)

    context = {
        'templates': templates,
        'stats': stats,
    }
    return render(request, 'dashboard/messaging_setup.html', context)


# Camera Settings and Data & Exports pages removed per request. If you want them restored,
# re-add their view functions and URL patterns and recreate the templates under
# templates/dashboard/camera_settings.html and templates/dashboard/data_exports.html

# ---- User Management ----
@require_permission([ADMINS_GROUP])
def dashboard_users(request):
    users = User.objects.all().select_related("userprofile__department")
    departments = Department.objects.all()
    groups = Group.objects.all()
    context = {
        "users": users,
        "departments": departments,
        "groups": groups,
    }
    return render(request, "dashboard/users.html", context)

@login_required
@require_role(['admin', 'staff'])
def manage_users_all(request):
    # Provide users queryset and related data to the template so the users table
    # can render real data server-side and be used by the client-side poller.
    users_qs = User.objects.all().select_related('userprofile').prefetch_related('groups')
    total_users = users_qs.count()
    # Build role counts for the UI (best-effort)
    role_names = ['Staff', 'Manager', 'Concierge', 'Maintenance', 'Housekeeping', 'Super Admin']
    role_counts = {}
    try:
        for rn in role_names:
            role_counts[rn.replace(' ', '_')] = User.objects.filter(groups__name__iexact=rn).distinct().count()
    except Exception:
        # Fallback: zero counts
        for rn in role_names:
            role_counts[rn.replace(' ', '_')] = 0

    # Departments list for dropdowns/modals
    try:
        departments = list(Department.objects.all().values('id', 'name'))
    except Exception:
        departments = []

    ctx = dict(active_tab="all",
               breadcrumb_title="Users",
               page_title="Manage Users",
               page_subtitle="Manage user accounts, roles, and permissions across your property.",
               search_placeholder="Search users...",
               primary_label="create User",
               users=users_qs,
               total_users=total_users,
               role_counts=role_counts,
               departments=departments)
    return render(request, 'dashboard/users.html', ctx)




        

@login_required
def manage_users_api_users(request, user_id=None):
    """Return a JSON list of users for the Manage Users frontend poller.

    Each user contains: id, username, first_name, last_name, full_name, email,
    avatar_url, roles (list), departments (single or null), is_active, last_login (iso),
    last_active_human (e.g. '2 hours').
    """
    # Handle DELETE request for user deletion
    if request.method == 'DELETE':
        if not user_id:
            return JsonResponse({'success': False, 'error': 'User ID is required'}, status=400)
        try:
            user = get_object_or_404(User, pk=user_id)
            user.delete()
            return JsonResponse({'success': True, 'message': 'User deleted successfully'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
    try:
        from django.utils.timesince import timesince
        from django.utils import timezone
    except Exception:
        timesince = None
    # Support filters: q (search), role, department, status (active/inactive), enabled (true/false)
    qs = User.objects.all().select_related('userprofile').prefetch_related('groups')
    q = request.GET.get('q')
    role = request.GET.get('role')
    department = request.GET.get('department')
    status = request.GET.get('status')
    enabled = request.GET.get('enabled')
    if q:
        qs = qs.filter(Q(username__icontains=q) | Q(first_name__icontains=q) | Q(last_name__icontains=q) | Q(email__icontains=q))
    if role:
        qs = qs.filter(groups__name__iexact=role)
    if department:
        qs = qs.filter(userprofile__department__name__iexact=department)
    if status:
        if status == 'active':
            qs = qs.filter(userprofile__enabled=True)
        elif status == 'inactive':
            qs = qs.filter(userprofile__enabled=False)
    if enabled is not None:
        if enabled.lower() in ('1', 'true', 'yes'):
            qs = qs.filter(userprofile__enabled=True)
        elif enabled.lower() in ('0', 'false', 'no'):
            qs = qs.filter(userprofile__enabled=False)

    # Pagination
    try:
        from django.core.paginator import Paginator, EmptyPage
    except Exception:
        Paginator = None
    page = int(request.GET.get('page', 1))
    page_size = int(request.GET.get('page_size', 10))
    total_count = qs.count()
    total_pages = 1
    page_obj_list = qs
    if Paginator:
        paginator = Paginator(qs, page_size)
        total_pages = paginator.num_pages
        try:
            page_obj = paginator.page(page)
            page_obj_list = page_obj.object_list
        except EmptyPage:
            page_obj_list = []
    users = []
    for u in page_obj_list[:1000]:
        profile = getattr(u, 'userprofile', None)
        avatar = getattr(profile, 'avatar_url', None) or ''
        dept = None
        if profile and profile.department:
            try:
                dept = profile.department.name
            except Exception:
                dept = None
        roles = [g.name for g in u.groups.all()]
        last_login = u.last_login
        if last_login and timesince:
            try:
                human = timesince(last_login, timezone.now()) + ' ago'
            except Exception:
                human = ''
        else:
            human = ''
        users.append({
            'id': u.pk,
            'username': u.username,
            'first_name': u.first_name,
            'last_name': u.last_name,
            'full_name': (u.get_full_name() or u.username),
            'email': u.email,
            'avatar_url': avatar,
            'roles': roles,
            'department': dept,
            'is_active': bool(getattr(profile, 'enabled', True)),  # Use UserProfile.enabled instead of User.is_active
            'enabled': bool(getattr(profile, 'enabled', True)),
            'last_login_iso': last_login.isoformat() if last_login else None,
            'last_active_human': human,
        })

    return JsonResponse({'users': users, 'total': total_count, 'page': page, 'page_size': page_size, 'total_pages': total_pages})


@login_required
def manage_users_api_filters(request):
    """Return available roles and departments for the Manage Users filters.

    Expected JSON shape:
    { roles: ["Admins","Staff",...], departments: ["Housekeeping", ...] }
    """
    roles = []
    departments = []
    try:
        roles = list(Group.objects.values_list('name', flat=True).distinct())
    except Exception:
        roles = []
    try:
        departments = list(Department.objects.values_list('name', flat=True).distinct())
    except Exception:
        departments = []
    return JsonResponse({'roles': roles, 'departments': departments})


@require_http_methods(['POST'])
@login_required
@user_passes_test(lambda u: u.is_superuser or u.is_staff)
def manage_users_api_bulk_action(request):
    """Bulk action endpoint. Expects JSON body with 'action' and 'user_ids' list.
    Supported actions: enable, disable
    """
    try:
        body = json.loads(request.body.decode('utf-8'))
    except Exception:
        return HttpResponseBadRequest('invalid json')
    action = body.get('action')
    ids = body.get('user_ids') or []
    if action not in ('enable', 'disable'):
        return HttpResponseBadRequest('unsupported action')
    if not isinstance(ids, list):
        return HttpResponseBadRequest('user_ids must be list')
    users = User.objects.filter(id__in=ids).select_related('userprofile')
    changed = []
    for u in users:
        profile = getattr(u, 'userprofile', None)
        if not profile:
            continue
        new_val = True if action == 'enable' else False
        if profile.enabled != new_val:
            profile.enabled = new_val
            profile.save(update_fields=['enabled'])
            changed.append(u.id)
    return JsonResponse({'changed': changed, 'action': action})

@login_required
def manage_users_groups(request):
    q = (request.GET.get('q') or '').strip()

    departments = []
    total_groups = 0
    total_group_members = 0
    recent_additions = 0
    active_groups = 0

    try:
        from hotel_app.models import Department, UserGroup, UserGroupMembership, UserProfile
        from django.db.models import Count, Max
        from django.utils.timesince import timesince
        from django.utils import timezone
        from datetime import timedelta
        from django.utils.text import slugify

        depts_qs = Department.objects.all().order_by('name')
        if q:
            depts_qs = depts_qs.filter(Q(name__icontains=q) | Q(description__icontains=q))

        total_groups = UserGroup.objects.count()
        memberships_qs = UserGroupMembership.objects.all()
        total_group_members = memberships_qs.count()
        recent_additions = memberships_qs.filter(joined_at__gte=timezone.now() - timedelta(hours=24)).count()
        active_groups = UserGroup.objects.annotate(mem_count=Count('usergroupmembership')).filter(mem_count__gt=0).count()

        color_map = {
            'Housekeeping': {'icon_bg': 'bg-green-500/10', 'tag_bg': 'bg-green-500/10', 'icon_color': 'green-500', 'dot_bg': 'bg-green-500'},
            'Front Office': {'icon_bg': 'bg-sky-600/10', 'tag_bg': 'bg-sky-600/10', 'icon_color': 'sky-600', 'dot_bg': 'bg-sky-600'},
            'Food & Beverage': {'icon_bg': 'bg-yellow-400/10', 'tag_bg': 'bg-yellow-400/10', 'icon_color': 'yellow-400', 'dot_bg': 'bg-yellow-400'},
            'Maintenance': {'icon_bg': 'bg-teal-500/10', 'tag_bg': 'bg-teal-500/10', 'icon_color': 'teal-500', 'dot_bg': 'bg-teal-500'},
            'Security': {'icon_bg': 'bg-red-500/10', 'tag_bg': 'bg-red-500/10', 'icon_color': 'red-500', 'dot_bg': 'bg-red-500'},
        }

        for dept_index, dept in enumerate(depts_qs):
            profiles_qs = UserProfile.objects.filter(department=dept)
            members_count = profiles_qs.count()
            supervisors_count = profiles_qs.filter(title__iregex=r'(supervisor|manager)').count()
            staff_count = max(0, members_count - supervisors_count)
            last_updated = profiles_qs.aggregate(max_up=Max('updated_at'))['max_up']
            human_updated = timesince(last_updated) + ' ago' if last_updated else 'N/A'

            featured = {
                'id': dept.pk,
                'name': dept.name,
                'description': dept.description or 'Department description',
                'members_count': members_count,
                'supervisors_count': supervisors_count,
                'staff_count': staff_count,
                'updated_at': human_updated,
                'image': f'images/manage_users/{slugify(dept.name)}.svg',
                'position_top': dept_index * 270,  # For CSS positioning
            }

            colors = color_map.get(dept.name, {'icon_bg': 'bg-gray-500/10', 'tag_bg': 'bg-gray-500/10', 'icon_color': 'gray-500', 'dot_bg': 'bg-gray-500'})
            featured.update(colors)

            # Get groups associated with this department
            groups_data = []
            groups_qs = dept.user_groups.all().order_by('name')
            for group_index, g in enumerate(groups_qs):
                mem_qs = g.usergroupmembership_set.all()
                mem_count = mem_qs.count()
                last_mem = mem_qs.order_by('-joined_at').first()
                updated_at = getattr(last_mem, 'joined_at', None)
                human_updated = timesince(updated_at) + ' ago' if updated_at else 'N/A'
                groups_data.append({
                    'pk': g.pk,  # Use pk instead of id for consistency
                    'name': g.name,
                    'members_count': mem_count,
                    'description': g.description or '',
                    'updated_at': human_updated,
                    'dot_bg': 'bg-green-500' if mem_count > 0 else 'bg-gray-300',
                    'position_top': group_index * 52,  # For CSS positioning
                })

            departments.append({'featured_group': featured, 'groups': groups_data})

    except Exception:
        # Fallback static data to match the provided HTML
        departments = [
            {
                'featured_group': {
                    'name': 'Housekeeping',
                    'description': 'Room cleaning, maintenance, and guest services',
                    'members_count': 42,
                    'supervisors_count': 6,
                    'staff_count': 36,
                    'updated_at': '2h ago',
                    'image': 'images/manage_users/house_keeping.svg',
                    'icon_bg': 'bg-green-500/10',
                    'tag_bg': 'bg-green-500/10',
                    'icon_color': 'green-500',
                    'dot_bg': 'bg-green-500',
                    'position_top': 0,
                },
                'groups': [
                    {'name': 'Floor Supervisors', 'members_count': 6, 'dot_bg': 'bg-green-500', 'position_top': 0},
                    {'name': 'Room Attendants', 'members_count': 28, 'dot_bg': 'bg-green-500', 'position_top': 52},
                    {'name': 'Laundry Team', 'members_count': 8, 'dot_bg': 'bg-green-500', 'position_top': 104},
                ]
            },
            {
                'featured_group': {
                    'name': 'Front Office',
                    'description': 'Reception, concierge, and guest relations',
                    'members_count': 18,
                    'supervisors_count': 3,
                    'staff_count': 15,
                    'updated_at': '1h ago',
                    'image': 'images/manage_users/front_office.svg',
                    'icon_bg': 'bg-sky-600/10',
                    'tag_bg': 'bg-sky-600/10',
                    'icon_color': 'sky-600',
                    'dot_bg': 'bg-sky-600',
                    'position_top': 270,
                },
                'groups': []
            },
            {
                'featured_group': {
                    'name': 'Food & Beverage',
                    'description': 'Kitchen, restaurant, bar, and room service',
                    'members_count': 31,
                    'supervisors_count': 8,
                    'staff_count': 23,
                    'updated_at': '30m ago',
                    'image': 'images/manage_users/food_beverage.svg',
                    'icon_bg': 'bg-yellow-400/10',
                    'tag_bg': 'bg-yellow-400/10',
                    'icon_color': 'yellow-400',
                    'dot_bg': 'bg-yellow-400',
                    'position_top': 540,
                },
                'groups': []
            },
            {
                'featured_group': {
                    'name': 'Maintenance',
                    'description': 'Technical support, repairs, and facility management',
                    'members_count': 12,
                    'supervisors_count': 3,
                    'staff_count': 9,
                    'updated_at': '4h ago',
                    'image': 'images/manage_users/maintainence.svg',
                    'icon_bg': 'bg-teal-500/10',
                    'tag_bg': 'bg-teal-500/10',
                    'icon_color': 'teal-500',
                    'dot_bg': 'bg-teal-500',
                    'position_top': 810,
                },
                'groups': []
            },
            {
                'featured_group': {
                    'name': 'Security',
                    'description': 'Property security, surveillance, and emergency response',
                    'members_count': 8,
                    'supervisors_count': 2,
                    'staff_count': 6,
                    'updated_at': '1h ago',
                    'image': 'images/manage_users/security.svg',
                    'icon_bg': 'bg-red-500/10',
                    'tag_bg': 'bg-red-500/10',
                    'icon_color': 'red-500',
                    'dot_bg': 'bg-red-500',
                    'position_top': 1080,
                },
                'groups': []
            },
        ]
        total_groups = 8
        total_group_members = 247
        recent_additions = 18
        active_groups = 5

    ctx = dict(
        active_tab="groups",
        breadcrumb_title="User Groups",
        page_title="User Groups",
        page_subtitle="Organize staff members by department, role, or location for targeted communication and management.",
        search_placeholder="Search groups...",
        primary_label="Create Group",
        departments=departments,
        total_groups=total_groups,
        total_group_members=total_group_members,
        recent_additions=recent_additions,
        active_groups=active_groups,
        q=q,
    )
    return render(request, 'dashboard/groups.html', ctx)


@login_required
@require_permission([ADMINS_GROUP, STAFF_GROUP])
def api_notify_all_groups(request):
    """POST endpoint to notify all groups (bulk notify).

    Expects JSON body: { "message": "..." } or will use a default message.
    Uses WhatsAppService.send_text as a best-effort mock integration.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    try:
        body = json.loads(request.body.decode('utf-8'))
    except Exception:
        body = {}
    message = body.get('message') or 'This is a bulk notification from Hotel Admin.'

    # Attempt to collect phone numbers from UserProfile and send messages
    sent = 0
    failed = 0
    try:
        from hotel_app.models import UserProfile
        profiles = UserProfile.objects.filter(enabled=True).exclude(phone__isnull=True).exclude(phone__exact='')
        phones = [p.phone for p in profiles]
    except Exception:
        phones = []

    service = WhatsAppService()
    for phone in phones:
        ok = service.send_text(phone, message)
        if ok:
            sent += 1
        else:
            failed += 1

    return JsonResponse({'sent': sent, 'failed': failed, 'attempted': len(phones)})


@login_required
@require_permission([ADMINS_GROUP, STAFF_GROUP])
def api_notify_department(request, dept_id):
    """POST endpoint to notify all members of a department.

    URL: /dashboard/api/departments/<dept_id>/notify/
    Body: { "message": "..." }
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    try:
        body = json.loads(request.body.decode('utf-8'))
    except Exception:
        body = {}
    message = body.get('message') or f'Notification for department {dept_id}.'

    sent = 0
    failed = 0
    try:
        from hotel_app.models import UserProfile, Department
        dept = get_object_or_404(Department, pk=dept_id)
        profiles = UserProfile.objects.filter(department=dept).exclude(phone__isnull=True).exclude(phone__exact='')
        phones = [p.phone for p in profiles]
    except Exception:
        phones = []

    service = WhatsAppService()
    for phone in phones:
        ok = service.send_text(phone, message)
        if ok:
            sent += 1
        else:
            failed += 1

    return JsonResponse({'sent': sent, 'failed': failed, 'attempted': len(phones)})


@login_required
@require_permission([ADMINS_GROUP, STAFF_GROUP])
def api_group_permissions(request, group_id):
    """Return JSON list of permissions for a user group."""
    try:
        from django.contrib.auth.models import Group
        group = get_object_or_404(Group, pk=group_id)
        
        # Define permissions based on group name (this is a simplified approach)
        # In a real application, you would have a more sophisticated permission system
        permissions = []
        
        # Default permissions for all groups
        base_permissions = [
            "view_profile",
            "update_profile",
            "view_requests",
            "update_requests"
        ]
        
        # Add specific permissions based on group name
        if group.name == "Admins":
            permissions = base_permissions + [
                "manage_users",
                "manage_groups",
                "system_config",
                "view_reports",
                "manage_departments",
                "full_access"
            ]
        elif group.name == "Staff":
            permissions = base_permissions + [
                "view_team_reports",
                "manage_team",
                "assign_requests",
                "view_dept_data"
            ]
        elif group.name == "Users":
            permissions = base_permissions
        else:
            permissions = base_permissions
            
    except Exception as e:
        permissions = []
    
    return JsonResponse({'permissions': permissions})


@login_required
@require_permission([ADMINS_GROUP])
@csrf_protect
def api_group_permissions_update(request, group_id):
    """Update permissions for a user group."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    
    try:
        from django.contrib.auth.models import Group, Permission
        group = get_object_or_404(Group, pk=group_id)
        
        # Parse JSON data
        try:
            data = json.loads(request.body.decode('utf-8'))
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
        
        # Get permissions from request
        permissions = data.get('permissions', [])
        
        # Get permission objects from codenames
        permission_objects = []
        for codename in permissions:
            try:
                perm = Permission.objects.get(codename=codename)
                permission_objects.append(perm)
            except Permission.DoesNotExist:
                # Skip permissions that don't exist
                continue
        
        # Update group permissions
        group.permissions.set(permission_objects)
        
        return JsonResponse({'success': True, 'message': 'Permissions updated successfully'})
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_permission([ADMINS_GROUP])
@csrf_protect
def api_bulk_permissions_update(request):
    """Update permissions for multiple user groups."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    
    try:
        from django.contrib.auth.models import Group, Permission
        
        # Parse JSON data
        try:
            data = json.loads(request.body.decode('utf-8'))
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
        
        # Get group IDs and permissions from request
        group_ids = data.get('group_ids', [])
        permissions = data.get('permissions', [])
        
        # Validate group IDs
        if not group_ids:
            return JsonResponse({'error': 'No groups specified'}, status=400)
        
        # Get permission objects from codenames
        permission_objects = []
        for codename in permissions:
            try:
                perm = Permission.objects.get(codename=codename)
                permission_objects.append(perm)
            except Permission.DoesNotExist:
                # Skip permissions that don't exist
                continue
        
        # Update permissions for each group
        updated_groups = []
        for group_id in group_ids:
            try:
                group = Group.objects.get(pk=group_id)
                group.permissions.set(permission_objects)
                updated_groups.append(group.name)
            except Group.DoesNotExist:
                continue  # Skip non-existent groups
        
        return JsonResponse({
            'success': True, 
            'message': f'Permissions updated for {len(updated_groups)} groups',
            'updated_groups': updated_groups
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

import json

from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse

from hotel_app.models import User
from .forms import GymMemberForm
from .models import GymMember

# Constants
ADMINS_GROUP = 'Admins'
STAFF_GROUP = 'Staff'


@login_required
@require_permission([ADMINS_GROUP])
def api_reset_user_password(request, user_id):
    """API endpoint to reset a user's password."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    
    try:
        user = get_object_or_404(User, pk=user_id)
        
        # Get the new password from the request
        try:
            data = json.loads(request.body.decode('utf-8'))
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
        
        new_password = data.get('new_password')
        confirm_password = data.get('confirm_password')
        
        # Validate passwords
        if not new_password:
            return JsonResponse({'error': 'New password is required'}, status=400)
        
        if new_password != confirm_password:
            return JsonResponse({'error': 'Passwords do not match'}, status=400)
        
        if len(new_password) < 8:
            return JsonResponse({'error': 'Password must be at least 8 characters long'}, status=400)
        
        # Set the new password
        user.set_password(new_password)
        user.save()
        
        # Log the password reset for debugging
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"Password reset for user {user.username} (ID: {user.id})")
        
        return JsonResponse({'success': True, 'message': 'Password reset successfully'})
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error resetting password for user ID {user_id}: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_permission([ADMINS_GROUP])
def export_user_data(request):
    """Export all user-related data (departments, users, groups, profiles)"""
    try:
        format = request.GET.get('format', 'json').lower()
        response = create_export_file(format)
        return response
    except Exception as e:
        logger = logging.getLogger(__name__)
        logger.error(f"Error exporting user data: {str(e)}")
        return JsonResponse({'error': 'Failed to export data'}, status=500)


@login_required
@require_permission([ADMINS_GROUP])
@csrf_exempt
def import_user_data(request):
    """Import user-related data from a JSON or Excel file"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST method required'}, status=405)
    
    try:
        # Get the uploaded file
        if 'file' not in request.FILES:
            return JsonResponse({'error': 'No file uploaded'}, status=400)
        
        uploaded_file = request.FILES['file']
        
        # Check file extension and process accordingly
        if uploaded_file.name.endswith('.json'):
            # Handle JSON file
            try:
                file_content = uploaded_file.read().decode('utf-8')
                data = json.loads(file_content)
            except json.JSONDecodeError as e:
                return JsonResponse({'error': f'Invalid JSON format: {str(e)}'}, status=400)
        elif uploaded_file.name.endswith('.xlsx'):
            # Handle Excel file
            try:
                from openpyxl import load_workbook
                from io import BytesIO
                
                # Load the workbook from the uploaded file
                file_content = uploaded_file.read()
                workbook = load_workbook(BytesIO(file_content), read_only=True)
                
                # Convert Excel data to the expected JSON structure
                data = {
                    'departments': [],
                    'user_groups': [],
                    'users': [],
                    'user_profiles': [],
                    'user_group_memberships': []
                }
                
                # Process departments sheet
                if 'Departments' in workbook.sheetnames:
                    ws = workbook['Departments']
                    rows = list(ws.iter_rows(values_only=True))
                    if len(rows) > 1:  # Header + data rows
                        headers = rows[0]
                        for row in rows[1:]:
                            if any(cell is not None for cell in row):  # Skip empty rows
                                dept_data = {}
                                for i, header in enumerate(headers):
                                    if header and i < len(row):
                                        dept_data[header] = row[i]
                                data['departments'].append(dept_data)
                
                # Process user groups sheet
                if 'User Groups' in workbook.sheetnames:
                    ws = workbook['User Groups']
                    rows = list(ws.iter_rows(values_only=True))
                    if len(rows) > 1:  # Header + data rows
                        headers = rows[0]
                        for row in rows[1:]:
                            if any(cell is not None for cell in row):  # Skip empty rows
                                group_data = {}
                                for i, header in enumerate(headers):
                                    if header and i < len(row):
                                        group_data[header] = row[i]
                                data['user_groups'].append(group_data)
                
                # Process users sheet
                if 'Users' in workbook.sheetnames:
                    ws = workbook['Users']
                    rows = list(ws.iter_rows(values_only=True))
                    if len(rows) > 1:  # Header + data rows
                        headers = rows[0]
                        for row in rows[1:]:
                            if any(cell is not None for cell in row):  # Skip empty rows
                                user_data = {}
                                for i, header in enumerate(headers):
                                    if header and i < len(row):
                                        user_data[header] = row[i]
                                data['users'].append(user_data)
                
                # Process user profiles sheet
                if 'User Profiles' in workbook.sheetnames:
                    ws = workbook['User Profiles']
                    rows = list(ws.iter_rows(values_only=True))
                    if len(rows) > 1:  # Header + data rows
                        headers = rows[0]
                        for row in rows[1:]:
                            if any(cell is not None for cell in row):  # Skip empty rows
                                profile_data = {}
                                for i, header in enumerate(headers):
                                    if header and i < len(row):
                                        profile_data[header] = row[i]
                                data['user_profiles'].append(profile_data)
                
                # Process user group memberships sheet
                if 'User Group Memberships' in workbook.sheetnames:
                    ws = workbook['User Group Memberships']
                    rows = list(ws.iter_rows(values_only=True))
                    if len(rows) > 1:  # Header + data rows
                        headers = rows[0]
                        for row in rows[1:]:
                            if any(cell is not None for cell in row):  # Skip empty rows
                                membership_data = {}
                                for i, header in enumerate(headers):
                                    if header and i < len(row):
                                        membership_data[header] = row[i]
                                data['user_group_memberships'].append(membership_data)
                
            except Exception as e:
                return JsonResponse({'error': f'Invalid Excel format: {str(e)}'}, status=400)
        else:
            return JsonResponse({'error': 'Only JSON (.json) or Excel (.xlsx) files are supported'}, status=400)
        
        # Import the data
        result = import_all_data(data)
        
        return JsonResponse({
            'success': True,
            'message': 'Data imported successfully',
            'result': result
        })
        
    except Exception as e:
        logger = logging.getLogger(__name__)
        logger.error(f"Error importing user data: {str(e)}")
        return JsonResponse({'error': f'Failed to import data: {str(e)}'}, status=500)
        
@login_required       
@require_role(['admin', 'staff'])
def tickets(request):
    """Render the Tickets Management page."""
    from hotel_app.models import ServiceRequest, RequestType, Department, User
    from django.db.models import Count, Q
    from datetime import timedelta
    from django.utils import timezone
    from django.core.paginator import Paginator
    
    # Get filter parameters from request
    department_filter = request.GET.get('department', '')
    priority_filter = request.GET.get('priority', '')
    status_filter = request.GET.get('status', '')
    search_query = request.GET.get('search', '')
    
    # Get departments with active ticket counts
    departments_data = []
    departments = Department.objects.all()
    for dept in departments:
        # Count active tickets for this department
        active_tickets_count = ServiceRequest.objects.filter(
            request_type__name__icontains=dept.name
        ).exclude(
            status__in=['completed', 'closed']
        ).count()
      # Calculate SLA compliance (simplified calculation)
        total_tickets = ServiceRequest.objects.filter(
            request_type__name__icontains=dept.name
        ).count()        
        completed_tickets = ServiceRequest.objects.filter(
            request_type__name__icontains=dept.name,
            status='completed'
        ).count()
        sla_compliance = 0
        if total_tickets > 0:
            sla_compliance = int((completed_tickets / total_tickets) * 100)
        color_mapping = {
            'Housekeeping': {'color': 'sky-600', 'icon_color': 'sky-600'},
            'Maintenance': {'color': 'yellow-400', 'icon_color': 'sky-600'},
            'Guest Services': {'color': 'green-500', 'icon_color': 'sky-600'},
        }
        
        dept_colors = color_mapping.get(dept.name, {'color': 'blue-500', 'icon_color': 'sky-600'})
        
        # Get logo URL if available using the new method
        logo_url = dept.get_logo_url() if dept.get_logo_url() else f'/static/images/manage_users/{dept.name.lower().replace(" ", "_")}.svg'
        
        departments_data.append({
            'id': dept.id,
            'name': dept.name,
            'active_tickets': active_tickets_count,
            'sla_compliance': sla_compliance,
            'sla_color': dept_colors['color'],
            'icon_url': logo_url,
        })
    
    # Get all service requests with filters applied
    tickets_queryset = ServiceRequest.objects.select_related(
        'request_type', 'location', 'requester_user', 'assignee_user'
    ).all().order_by('-id')
    
    # Apply filters
    if department_filter and department_filter != 'All Departments':
        tickets_queryset = tickets_queryset.filter(
            request_type__name__icontains=department_filter
        )
    
    if priority_filter and priority_filter != 'All Priorities':
        # Map display values to model values
        priority_mapping = {
            'High': 'high',
            'Medium': 'normal',
            'Low': 'low'
        }
        model_priority = priority_mapping.get(priority_filter)
        if model_priority:
            tickets_queryset = tickets_queryset.filter(priority=model_priority)
    
    if status_filter and status_filter != 'All Statuses':
        # Map display values to model values
        status_mapping = {
            'Pending': 'pending',
            'Accepted': 'accepted',
            'In Progress': 'in_progress',
            'Completed': 'completed',
            'Closed': 'closed',
            'Escalated': 'escalated',
            'Rejected': 'rejected'
        }
        model_status = status_mapping.get(status_filter)
        if model_status:
            tickets_queryset = tickets_queryset.filter(status=model_status)
    
    if search_query:
        tickets_queryset = tickets_queryset.filter(
            Q(request_type__name__icontains=search_query) |
            Q(location__name__icontains=search_query) |
            Q(notes__icontains=search_query)
        )
    
    # Process tickets to add color attributes
    processed_tickets = []
    for ticket in tickets_queryset:
        # Map priority to display values
        priority_mapping = {
            'high': {'label': 'High', 'color': 'red'},
            'normal': {'label': 'Medium', 'color': 'sky'},
            'low': {'label': 'Low', 'color': 'gray'},
        }
        
        priority_data = priority_mapping.get(ticket.priority, {'label': 'Medium', 'color': 'sky'})
        
        # Map status to display values
        status_mapping = {
            'pending': {'label': 'Pending', 'color': 'yellow'},
            'assigned': {'label': 'Assigned', 'color': 'yellow'},
            'accepted': {'label': 'Accepted', 'color': 'blue'},
            'in_progress': {'label': 'In Progress', 'color': 'sky'},
            'completed': {'label': 'Completed', 'color': 'green'},
            'closed': {'label': 'Closed', 'color': 'green'},
            'escalated': {'label': 'Escalated', 'color': 'red'},
            'rejected': {'label': 'Rejected', 'color': 'red'},
        }
        
        status_data = status_mapping.get(ticket.status, {'label': 'Pending', 'color': 'yellow'})
        
        # Calculate SLA percentage
        sla_percentage = 0
        sla_color = 'green-500'
        if ticket.created_at and ticket.due_at:
            # Calculate time taken so far or total time if completed
            if ticket.completed_at:
                time_taken = ticket.completed_at - ticket.created_at
            else:
                time_taken = timezone.now() - ticket.created_at
            
            # Calculate SLA percentage (time taken / total allowed time)
            total_allowed_time = ticket.due_at - ticket.created_at
            if total_allowed_time.total_seconds() > 0:
                sla_percentage = min(100, int((time_taken.total_seconds() / total_allowed_time.total_seconds()) * 100))
            
            # Determine color based on SLA
            if sla_percentage > 90:
                sla_color = 'red-500'
            elif sla_percentage > 70:
                sla_color = 'yellow-400'
            else:
                sla_color = 'green-500'
        
        # Add attributes to the ticket object
        ticket.priority_label = priority_data['label']
        ticket.priority_color = priority_data['color']
        ticket.status_label = status_data['label']
        ticket.status_color = status_data['color']
        ticket.sla_percentage = sla_percentage
        ticket.sla_color = sla_color
        ticket.owner_avatar = 'https://placehold.co/24x24'
        
        processed_tickets.append(ticket)
    
    # --- Pagination Logic ---
    paginator = Paginator(processed_tickets, 10)  # Show 10 tickets per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'departments': departments_data,
        'tickets': page_obj,  # Pass the page_obj to the template
        'page_obj': page_obj,  # Pass it again as page_obj for clarity
        'total_tickets': tickets_queryset.count(),
        # Pass filter values back to template
        'department_filter': department_filter,
        'priority_filter': priority_filter,
        'status_filter': status_filter,
        'search_query': search_query,
    }
    return render(request, 'dashboard/tickets.html', context)


@login_required
def gym(request):
    """Render the Gym Management page."""
    # Handle form submission
    if request.method == 'POST':
        form = GymMemberForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Gym member created successfully!')
            return redirect('dashboard:gym')
        else:
            messages.error(request, 'Please correct the errors below.')
            # Print form errors for debugging
            print("Form errors:", form.errors)
    else:
        form = GymMemberForm()
    
    # Get gym members from database with pagination
    gym_members_list = GymMember.objects.all().order_by('-id')
    
    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(gym_members_list, 10)  # Show 10 members per page
    
    try:
        gym_members = paginator.page(page)
    except PageNotAnInteger:
        gym_members = paginator.page(1)
    except EmptyPage:
        gym_members = paginator.page(paginator.num_pages)
    
    # Convert to the format expected by the template
    gym_members_data = []
    for member in gym_members:
        gym_members_data.append({
            'id': member.id,
            'name': member.full_name,
            'city': member.city or '',
            'phone': member.phone or '',
            'email': member.email or '',
            'start_date': member.start_date or '',
            'end_date': member.end_date or '',
            'qr_code': 'https://placehold.co/30x32'  # Placeholder QR code
        })
    
    context = {
        'gym_members': gym_members_data,
        'total_members': gym_members_list.count(),
        'page_size': 10,
        'current_page': gym_members.number,
        'paginator': gym_members,
        'form': form
    }
    return render(request, 'dashboard/gym.html', context)


def ticket_detail(request, ticket_id):
    """Render the Ticket Detail page."""
    from hotel_app.models import ServiceRequest, User, AuditLog
    from django.utils import timezone
    from django.db.models import Q
    
    # Get the service request
    service_request = get_object_or_404(ServiceRequest, id=ticket_id)
    
    # Check SLA breaches to ensure status is up to date
    service_request.check_sla_breaches()
    
    # Calculate SLA progress percentage
    sla_progress_percent = 0
    if service_request.created_at and service_request.sla_hours > 0:
        # For completed/closed tickets, show 100% progress
        if service_request.status in ['completed', 'closed']:
            sla_progress_percent = 100
        else:
            # Calculate time taken so far or total time if completed
            if service_request.completed_at:
                time_taken = service_request.completed_at - service_request.created_at
            else:
                time_taken = timezone.now() - service_request.created_at
            
            # Calculate SLA percentage (time taken / total allowed time)
            total_allowed_time = service_request.sla_hours * 3600  # Convert hours to seconds
            if total_allowed_time > 0:
                sla_progress_percent = min(100, int((time_taken.total_seconds() / total_allowed_time) * 100))
    
    # Map priority to display values
    priority_mapping = {
        'critical': {'label': 'Critical', 'color': 'red-500'},
        'high': {'label': 'High', 'color': 'orange-500'},
        'normal': {'label': 'Normal', 'color': 'sky-600'},
        'low': {'label': 'Low', 'color': 'gray-100'},
    }
    
    priority_data = priority_mapping.get(service_request.priority, {'label': 'Normal', 'color': 'sky-600'})
    
    # Map status to display values
    status_mapping = {
        'pending': {'label': 'Pending', 'color': 'yellow-400'},
        'assigned': {'label': 'Assigned', 'color': 'yellow-400'},
        'accepted': {'label': 'Accepted', 'color': 'blue-500'},
        'in_progress': {'label': 'In Progress', 'color': 'sky-600'},
        'completed': {'label': 'Completed', 'color': 'green-500'},
        'closed': {'label': 'Closed', 'color': 'green-500'},
        'escalated': {'label': 'Escalated', 'color': 'red-500'},
        'rejected': {'label': 'Rejected', 'color': 'red-500'},
    }
    
    status_data = status_mapping.get(service_request.status, {'label': 'Pending', 'color': 'yellow-400'})
    
    # Get requester name
    requester_name = 'Unknown'
    if service_request.requester_user:
        requester_name = service_request.requester_user.get_full_name() or service_request.requester_user.username
    
    # Get assignee name
    assignee_name = 'Unassigned'
    if service_request.assignee_user:
        assignee_name = service_request.assignee_user.get_full_name() or service_request.assignee_user.username
    
    # Check if current user is the assignee
    is_assignee = (service_request.assignee_user == request.user)
    
    # Get available users for assignment
    available_users = User.objects.filter(is_active=True).exclude(id=request.user.id)
    
    # Get request type name
    request_type_name = 'Unknown Request'
    if service_request.request_type:
        request_type_name = service_request.request_type.name
    
    # Get location info
    location_name = 'Unknown Location'
    room_number = 'N/A'
    floor = 'N/A'
    building = 'N/A'
    room_type = 'N/A'
    
    if service_request.location:
        location_name = service_request.location.name
        room_number = getattr(service_request.location, 'room_no', 'N/A') or 'N/A'
        if service_request.location.floor:
            floor = f"{service_request.location.floor.floor_number} Floor"
            if service_request.location.floor.building:
                building = service_request.location.floor.building.name
        if service_request.location.type:
            room_type = service_request.location.type.name
    
    # Get department info - Use the actual department from the service request
    department_name = 'Unknown Department'
    if service_request.department:
        department_name = service_request.department.name
    elif service_request.request_type:
        # Fallback to work_family or request_family if no department is assigned
        if hasattr(service_request.request_type, 'work_family') and service_request.request_type.work_family:
            department_name = service_request.request_type.work_family.name
        elif hasattr(service_request.request_type, 'request_family') and service_request.request_type.request_family:
            department_name = service_request.request_type.request_family.name
    
    # Format created time
    created_time = service_request.created_at.strftime("%b %d, %H:%M") if service_request.created_at else "Unknown"
    
    # Get notification count (for now, we'll simulate this with a static value)
    # In a real implementation, this would come from a notification model
    notification_count = 3  # This will be replaced with dynamic count
    
    # Get activity log for this ticket
    activity_log = []
    
    # Get audit logs for this ticket
    ticket_audit_logs = AuditLog.objects.filter(
        model_name='ServiceRequest',
        object_pk=str(service_request.pk)
    ).order_by('-created_at')
    
    # Convert audit logs to activity log format
    for log in ticket_audit_logs:
        # Format the timestamp
        time_ago = ""
        if log.created_at:
            # Calculate time difference
            diff = timezone.now() - log.created_at
            if diff.days > 0:
                time_ago = f"{diff.days} day{'s' if diff.days != 1 else ''} ago"
            elif diff.seconds > 3600:
                hours = diff.seconds // 3600
                time_ago = f"{hours} hour{'s' if hours != 1 else ''} ago"
            elif diff.seconds > 60:
                minutes = diff.seconds // 60
                time_ago = f"{minutes} minute{'s' if minutes != 1 else ''} ago"
            else:
                time_ago = "Just now"
        
        # Determine action description
        action_desc = ""
        actor_name = "System"
        if log.actor:
            actor_name = log.actor.get_full_name() or log.actor.username
        
        if log.action == 'create':
            action_desc = "Ticket created"
        elif log.action == 'update':
            # Check what was updated
            if log.changes:
                if 'status' in log.changes:
                    old_status = log.changes['status'][0] if isinstance(log.changes['status'], list) else log.changes['status']
                    new_status = log.changes['status'][1] if isinstance(log.changes['status'], list) else log.changes['status']
                    # Map status codes to display names
                    old_label = status_mapping.get(old_status, {'label': old_status})['label']
                    new_label = status_mapping.get(new_status, {'label': new_status})['label']
                    action_desc = f"Status changed from {old_label} to {new_label}"
                elif 'priority' in log.changes:
                    old_priority = log.changes['priority'][0] if isinstance(log.changes['priority'], list) else log.changes['priority']
                    new_priority = log.changes['priority'][1] if isinstance(log.changes['priority'], list) else log.changes['priority']
                    # Map priority codes to display names
                    old_label = priority_mapping.get(old_priority, {'label': old_priority})['label']
                    new_label = priority_mapping.get(new_priority, {'label': new_priority})['label']
                    action_desc = f"Priority changed from {old_label} to {new_label}"
                elif 'assignee_user' in log.changes:
                    old_assignee = log.changes['assignee_user'][0] if isinstance(log.changes['assignee_user'], list) else log.changes['assignee_user']
                    new_assignee = log.changes['assignee_user'][1] if isinstance(log.changes['assignee_user'], list) else log.changes['assignee_user']
                    if old_assignee and new_assignee:
                        action_desc = "Ticket reassigned"
                    elif new_assignee:
                        action_desc = "Ticket assigned"
                    else:
                        action_desc = "Ticket unassigned"
                elif 'notes' in log.changes:
                    action_desc = "Internal comment added"
                else:
                    action_desc = "Ticket updated"
            else:
                action_desc = "Ticket updated"
        elif log.action == 'delete':
            action_desc = "Ticket deleted"
        else:
            action_desc = f"{log.action.capitalize()} action performed"
        
        activity_log.append({
            'description': action_desc,
            'actor': actor_name,
            'time_ago': time_ago,
            'timestamp': log.created_at
        })
    
    # Add the ticket creation event if not already in logs
    if service_request.created_at and not any(log.get('timestamp') == service_request.created_at for log in activity_log):
        # Format the creation time
        time_ago = ""
        diff = timezone.now() - service_request.created_at
        if diff.days > 0:
            time_ago = f"{diff.days} day{'s' if diff.days != 1 else ''} ago"
        elif diff.seconds > 3600:
            hours = diff.seconds // 3600
            time_ago = f"{hours} hour{'s' if hours != 1 else ''} ago"
        elif diff.seconds > 60:
            minutes = diff.seconds // 60
            time_ago = f"{minutes} minute{'s' if minutes != 1 else ''} ago"
        else:
            time_ago = "Just now"
            
        activity_log.append({
            'description': 'Ticket created',
            'actor': requester_name,
            'time_ago': time_ago,
            'timestamp': service_request.created_at
        })
    
    # Sort activity log by timestamp (newest first)
    activity_log.sort(key=lambda x: x['timestamp'] if x['timestamp'] else timezone.datetime.min.replace(tzinfo=timezone.utc), reverse=True)
    
    # Limit to last 10 activities
    activity_log = activity_log[:10]
    
    context = {
        'ticket': service_request,
        'ticket_priority_label': priority_data['label'],
        'ticket_priority_color': priority_data['color'],
        'ticket_status_label': status_data['label'],
        'ticket_status_color': status_data['color'],
        'requester_name': requester_name,
        'assignee_name': assignee_name,
        'is_assignee': is_assignee,
        'available_users': available_users,
        'request_type_name': request_type_name,
        'location_name': location_name,
        'room_number': room_number,
        'floor': floor,
        'building': building,
        'room_type': room_type,
        'department_name': department_name,
        'created_time': created_time,
        'notification_count': notification_count,
        'activity_log': activity_log,
        'sla_progress_percent': sla_progress_percent,
        'resolution_notes': service_request.resolution_notes  # Pass resolution notes to template
    }
    
    return render(request, 'dashboard/ticket_detail.html', context)


@login_required
@require_role(['admin', 'staff', 'user'])
def my_tickets(request):
    """Render the My Tickets page with dynamic status cards."""
    from django.db.models import Q, Count
    from .models import ServiceRequest
    
    # Get the current user's tickets
    user_tickets = ServiceRequest.objects.filter(
        Q(assignee_user=request.user) | Q(requester_user=request.user)
    ).order_by('-created_at')
    
    # Calculate status counts for the status cards
    status_counts = user_tickets.aggregate(
        pending=Count('id', filter=Q(status='pending')),
        accepted=Count('id', filter=Q(status='accepted')),
        in_progress=Count('id', filter=Q(status='in_progress')),
        completed=Count('id', filter=Q(status='completed')),
        closed=Count('id', filter=Q(status='closed')),
        escalated=Count('id', filter=Q(status='escalated')),
        rejected=Count('id', filter=Q(status='rejected'))
    )
    
    # Calculate overdue count
    from django.utils import timezone
    overdue_count = user_tickets.filter(
        due_at__lt=timezone.now(),
        status__in=['pending', 'accepted', 'in_progress']
    ).count()
    
    # Handle filtering
    priority_filter = request.GET.get('priority', '')
    status_filter = request.GET.get('status', '')
    search_query = request.GET.get('search', '')
    
    # Convert display status values to database status values
    status_mapping = {
        'Pending': 'pending',
        'Accepted': 'accepted',
        'In Progress': 'in_progress',
        'Completed': 'completed',
        'Closed': 'closed',
        'Escalated': 'escalated',
        'Rejected': 'rejected'
    }
    
    if priority_filter:
        user_tickets = user_tickets.filter(priority=priority_filter.lower())
    
    if status_filter:
        # Convert display status to database status
        db_status = status_mapping.get(status_filter, status_filter.lower())
        user_tickets = user_tickets.filter(status=db_status)
    
    if search_query:
        user_tickets = user_tickets.filter(
            Q(notes__icontains=search_query) |
            Q(request_type__name__icontains=search_query) |
            Q(location__name__icontains=search_query)
        )
    
    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(user_tickets, 10)  # Show 10 tickets per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'tickets': page_obj,
        'page_obj': page_obj,
        'status_counts': status_counts,
        'overdue_count': overdue_count,
        'priority_filter': priority_filter,
        'status_filter': status_filter,
        'search_query': search_query,
    }
    
    return render(request, 'dashboard/my_tickets.html', context)


@login_required
@require_permission([ADMINS_GROUP, STAFF_GROUP])
def gym_report(request):
    """Render the Gym Report page."""
    from hotel_app.models import GymVisit
    
    # Fetch gym visits for the current month
    from django.utils import timezone
    current_month = timezone.now().month
    current_year = timezone.now().year
    gym_visits = GymVisit.objects.filter(
        visit_date__month=current_month,
        visit_date__year=current_year
    ).order_by('-visit_date')
    
    context = {
        'gym_visits': gym_visits,
        'total_visits': 24,  # Total number of gym visits
        'page_size': 10,     # Number of visits per page
        'current_page': 1    # Current page number
    }
    return render(request, 'dashboard/gym_report.html', context)


@login_required
@require_role(['admin', 'staff', 'user'])
def my_tickets(request):
    """Render the My Tickets page with dynamic status cards."""
    from django.db.models import Q, Count
    from .models import ServiceRequest
    from django.utils import timezone
    from django.core.paginator import Paginator
    
    # Get the current user's department
    user_department = None
    if hasattr(request.user, 'userprofile') and request.user.userprofile.department:
        user_department = request.user.userprofile.department
    
    # Get service requests assigned to the current user (either as assignee or requester)
    # Also include pending tickets in the user's department that are not yet assigned
    user_tickets = ServiceRequest.objects.filter(
        Q(assignee_user=request.user) | 
        Q(requester_user=request.user) |
        (Q(department=user_department) & Q(status='pending') & Q(assignee_user=None))
    ).select_related(
        'request_type', 'location', 'requester_user', 'assignee_user', 'department'
    ).order_by('-created_at')
    
    # Calculate status counts for the status cards
    status_counts = user_tickets.aggregate(
        pending=Count('id', filter=Q(status='pending')),
        accepted=Count('id', filter=Q(status='accepted')),
        in_progress=Count('id', filter=Q(status='in_progress')),
        completed=Count('id', filter=Q(status='completed')),
        closed=Count('id', filter=Q(status='closed')),
        escalated=Count('id', filter=Q(status='escalated')),
        rejected=Count('id', filter=Q(status='rejected'))
    )
    
    # Calculate overdue count
    overdue_count = user_tickets.filter(
        due_at__lt=timezone.now(),
        status__in=['pending', 'accepted', 'in_progress']
    ).count()
    
    # Handle filtering
    priority_filter = request.GET.get('priority', '')
    status_filter = request.GET.get('status', '')
    search_query = request.GET.get('search', '')
    
    # Convert display status values to database status values
    status_mapping = {
        'Pending': 'pending',
        'Accepted': 'accepted',
        'In Progress': 'in_progress',
        'Completed': 'completed',
        'Closed': 'closed',
        'Escalated': 'escalated',
        'Rejected': 'rejected'
    }
    
    if priority_filter:
        user_tickets = user_tickets.filter(priority=priority_filter.lower())
    
    if status_filter:
        # Convert display status to database status
        db_status = status_mapping.get(status_filter, status_filter.lower())
        user_tickets = user_tickets.filter(status=db_status)
    
    if search_query:
        user_tickets = user_tickets.filter(
            Q(notes__icontains=search_query) |
            Q(request_type__name__icontains=search_query) |
            Q(location__name__icontains=search_query)
        )
    
    # Process tickets to add color attributes and workflow permissions
    processed_tickets = []
    for ticket in user_tickets:
        # Map priority to display values
        priority_mapping = {
            'high': {'label': 'High', 'color': 'red'},
            'normal': {'label': 'Medium', 'color': 'sky'},
            'low': {'label': 'Low', 'color': 'gray'},
        }
        
        priority_data = priority_mapping.get(ticket.priority, {'label': 'Medium', 'color': 'sky'})
        
        # Map status to display values
        status_mapping = {
            'pending': {'label': 'Pending', 'color': 'yellow'},
            'assigned': {'label': 'Assigned', 'color': 'yellow'},
            'accepted': {'label': 'Accepted', 'color': 'blue'},
            'in_progress': {'label': 'In Progress', 'color': 'sky'},
            'completed': {'label': 'Completed', 'color': 'green'},
            'closed': {'label': 'Closed', 'color': 'green'},
            'escalated': {'label': 'Escalated', 'color': 'red'},
            'rejected': {'label': 'Rejected', 'color': 'red'},
        }
        
        status_data = status_mapping.get(ticket.status, {'label': 'Pending', 'color': 'yellow'})
        
        # Check SLA breaches
        ticket.check_sla_breaches()
        
        # Calculate SLA progress percentage
        sla_progress_percent = 0
        if ticket.created_at and ticket.due_at:
            # Calculate time taken so far or total time if completed
            if ticket.completed_at:
                time_taken = ticket.completed_at - ticket.created_at
            else:
                time_taken = timezone.now() - ticket.created_at
            
            # Calculate SLA percentage (time taken / total allowed time)
            total_allowed_time = ticket.due_at - ticket.created_at
            if total_allowed_time.total_seconds() > 0:
                sla_progress_percent = min(100, int((time_taken.total_seconds() / total_allowed_time.total_seconds()) * 100))
        
        # Add attributes to the ticket object
        ticket.priority_label = priority_data['label']
        ticket.priority_color = priority_data['color']
        ticket.status_label = status_data['label']
        ticket.status_color = status_data['color']
        ticket.owner_avatar = 'https://placehold.co/24x24'
        ticket.sla_progress_percent = sla_progress_percent
        
        # Add user-specific workflow permissions
        ticket.can_accept = False
        ticket.can_start = False
        ticket.can_complete = False
        ticket.can_close = False
        
        # Determine what actions the user can take based on workflow
        # For pending tickets, any user can accept (which will assign to them)
        if ticket.status == 'pending' and ticket.assignee_user is None:
            # Unassigned ticket - user can accept if it's in their department or they're the requester
            if ticket.department == user_department or ticket.requester_user == request.user:
                ticket.can_accept = True
        elif ticket.status == 'accepted' and ticket.assignee_user == request.user:
            # Accepted by current user - can start work
            ticket.can_start = True
        elif ticket.status == 'in_progress' and ticket.assignee_user == request.user:
            # In progress by current user - can complete
            ticket.can_complete = True
        elif ticket.status in ['completed', 'in_progress']:
            # Check if user can close (requester, front desk, or superuser)
            is_requester = (ticket.requester_user == request.user)
            is_front_desk = user_in_group(request.user, 'Front Desk') or user_in_group(request.user, 'Front Desk Team')
            is_superuser = request.user.is_superuser
            if is_requester or is_front_desk or is_superuser:
                ticket.can_close = True
        
        processed_tickets.append(ticket)
    
    # --- Pagination Logic ---
    paginator = Paginator(processed_tickets, 10)  # Show 10 tickets per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'tickets': page_obj,
        'page_obj': page_obj,
        'status_counts': status_counts,
        'overdue_count': overdue_count,
        'priority_filter': priority_filter,
        'status_filter': status_filter,
        'search_query': search_query,
        'user_department': user_department,
    }
    
    return render(request, 'dashboard/my_tickets.html', context)


# Removed claim_ticket_api as we're removing the claim functionality
# Tickets are now directly assigned when accepted
    


# Removed claim_ticket_api as we're removing the claim functionality
# Tickets are now directly assigned when accepted

    # Apply filters
    if search_query:
        tickets_queryset = tickets_queryset.filter(
            Q(request_type__name__icontains=search_query) |
            Q(location__name__icontains=search_query) |
            Q(notes__icontains=search_query)
        )
    
    # Process tickets to add color attributes
    processed_tickets = []
    for ticket in tickets_queryset:
        # Map priority to display values
        priority_mapping = {
            'high': {'label': 'High', 'color': 'red'},
            'normal': {'label': 'Medium', 'color': 'sky'},
            'low': {'label': 'Low', 'color': 'gray'},
        }
        
        priority_data = priority_mapping.get(ticket.priority, {'label': 'Medium', 'color': 'sky'})
        
        # Map status to display values
        status_mapping = {
            'pending': {'label': 'Pending', 'color': 'yellow'},
            'assigned': {'label': 'Assigned', 'color': 'yellow'},
            'accepted': {'label': 'Accepted', 'color': 'blue'},
            'in_progress': {'label': 'In Progress', 'color': 'sky'},
            'completed': {'label': 'Completed', 'color': 'green'},
            'closed': {'label': 'Closed', 'color': 'green'},
            'escalated': {'label': 'Escalated', 'color': 'red'},
            'rejected': {'label': 'Rejected', 'color': 'red'},
        }
        
        status_data = status_mapping.get(ticket.status, {'label': 'Pending', 'color': 'yellow'})
        
        # Calculate SLA percentage
        sla_percentage = 0
        sla_color = 'green-500'
        if ticket.created_at and ticket.due_at:
            # Calculate time taken so far or total time if completed
            if ticket.completed_at:
                time_taken = ticket.completed_at - ticket.created_at
            else:
                time_taken = timezone.now() - ticket.created_at
            
            # Calculate SLA percentage (time taken / total allowed time)
            total_allowed_time = ticket.due_at - ticket.created_at
            if total_allowed_time.total_seconds() > 0:
                sla_percentage = min(100, int((time_taken.total_seconds() / total_allowed_time.total_seconds()) * 100))
            
            # Determine color based on SLA
            if sla_percentage > 90:
                sla_color = 'red-500'
            elif sla_percentage > 70:
                sla_color = 'yellow-400'
            else:
                sla_color = 'green-500'
        
        # Add attributes to the ticket object
        ticket.priority_label = priority_data['label']
        ticket.priority_color = priority_data['color']
        ticket.status_label = status_data['label']
        ticket.status_color = status_data['color']
        ticket.sla_percentage = sla_percentage
        ticket.sla_color = sla_color
        ticket.owner_avatar = 'https://placehold.co/24x24'
        
        # Add user-specific information
        ticket.can_accept = False
        ticket.can_start = False
        ticket.can_complete = False
        ticket.can_close = False
        
        # Determine what actions the user can take
        if ticket.status == 'assigned' and ticket.assignee_user == request.user:
            ticket.can_accept = True
        elif ticket.status == 'accepted' and ticket.assignee_user == request.user:
            ticket.can_start = True
        elif ticket.status == 'in_progress' and ticket.assignee_user == request.user:
            ticket.can_complete = True
        elif ticket.status in ['completed', 'in_progress']:
            # Check if user is requester, front desk, or superuser
            is_requester = (ticket.requester_user == request.user)
            is_front_desk = user_in_group(request.user, 'Front Desk')
            is_superuser = request.user.is_superuser
            if is_requester or is_front_desk or is_superuser:
                ticket.can_close = True
        elif ticket.status == 'pending' and ticket.department == user_department and ticket.assignee_user is None:
            # Unassigned ticket in user's department - user can claim it
            ticket.can_claim = True
        
        processed_tickets.append(ticket)
    
    # --- Pagination Logic ---
    paginator = Paginator(processed_tickets, 10)  # Show 10 tickets per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'tickets': page_obj,  # Pass the page_obj to the template
        'page_obj': page_obj,  # Pass it again as page_obj for clarity
        'total_tickets': tickets_queryset.count(),
        'user_department': user_department,
        # Pass filter values back to template
        'priority_filter': priority_filter,
        'status_filter': status_filter,
        'search_query': search_query,
    }
    return render(request, 'dashboard/my_tickets.html', context)


@login_required
@require_permission([ADMINS_GROUP])
def configure_requests(request):
    """Render the Predefined / Configure Requests page.
    Uses actual department data from the database.
    """
    # Get all departments from the database
    departments = Department.objects.all().order_by('name')
    
    # Get all request types and associate them with departments
    request_types = RequestType.objects.select_related('work_family').all()
    
    # Create requests list with actual department data
    requests_list = []
    
    # First, add request types that have a work_family (department association)
    for request_type in request_types:
        department_name = 'General'
        icon = 'images/manage_users/general.svg'
        icon_bg = 'bg-gray-500/10'
        tag_bg = 'bg-gray-500/10'
        
        if request_type.work_family:
            department_name = request_type.work_family.name
            # Map department names to appropriate icons
            department_lower = department_name.lower()
            if 'housekeeping' in department_lower:
                icon = 'images/manage_users/house_keeping.svg'
                icon_bg = 'bg-green-500/10'
                tag_bg = 'bg-green-500/10'
            elif 'maintenance' in department_lower:
                icon = 'images/manage_users/maintainence.svg'
                icon_bg = 'bg-yellow-400/10'
                tag_bg = 'bg-yellow-400/10'
            elif 'concierge' in department_lower:
                icon = 'images/manage_users/concierge.svg'
                icon_bg = 'bg-fuchsia-700/10'
                tag_bg = 'bg-fuchsia-700/10'
            elif 'food' in department_lower or 'restaurant' in department_lower:
                icon = 'images/manage_users/food_beverage.svg'
                icon_bg = 'bg-red-500/10'
                tag_bg = 'bg-red-500/10'
            elif 'front' in department_lower or 'desk' in department_lower:
                icon = 'images/manage_users/front_office.svg'
                icon_bg = 'bg-sky-500/10'
                tag_bg = 'bg-sky-500/10'
            else:
                # Default icon for other departments
                icon = f'images/manage_users/{department_lower.replace(" ", "_")}.svg'
                icon_bg = 'bg-blue-500/10'
                tag_bg = 'bg-blue-500/10'
        
        requests_list.append({
            'title': request_type.name,
            'department': department_name,
            'description': request_type.description or f'Request type for {department_name}',
            'fields': 4,  # This would need to be calculated based on actual fields
            'exposed': request_type.active,
            'icon': icon,
            'icon_bg': icon_bg,
            'tag_bg': tag_bg,
        })
    
    # Add departments that don't have request types yet as placeholders
    for department in departments:
        # Check if we already have requests for this department
        has_requests = any(req['department'] == department.name for req in requests_list)
        
        if not has_requests:
            # Add a placeholder request for the department
            department_lower = department.name.lower()
            icon = 'images/manage_users/general.svg'
            icon_bg = 'bg-gray-500/10'
            tag_bg = 'bg-gray-500/10'
            
            # Map department names to appropriate icons
            if 'housekeeping' in department_lower:
                icon = 'images/manage_users/house_keeping.svg'
                icon_bg = 'bg-green-500/10'
                tag_bg = 'bg-green-500/10'
            elif 'maintenance' in department_lower:
                icon = 'images/manage_users/maintainence.svg'
                icon_bg = 'bg-yellow-400/10'
                tag_bg = 'bg-yellow-400/10'
            elif 'concierge' in department_lower:
                icon = 'images/manage_users/concierge.svg'
                icon_bg = 'bg-fuchsia-700/10'
                tag_bg = 'bg-fuchsia-700/10'
            elif 'food' in department_lower or 'restaurant' in department_lower:
                icon = 'images/manage_users/food_beverage.svg'
                icon_bg = 'bg-red-500/10'
                tag_bg = 'bg-red-500/10'
            elif 'front' in department_lower or 'desk' in department_lower:
                icon = 'images/manage_users/front_office.svg'
                icon_bg = 'bg-sky-500/10'
                tag_bg = 'bg-sky-500/10'
            else:
                # Default icon for other departments
                icon = f'images/manage_users/{department_lower.replace(" ", "_")}.svg'
                icon_bg = 'bg-blue-500/10'
                tag_bg = 'bg-blue-500/10'
            
            requests_list.append({
                'title': f'{department.name} Request',
                'department': department.name,
                'description': f'Request services from the {department.name} department',
                'fields': 3,
                'exposed': True,
                'icon': icon,
                'icon_bg': icon_bg,
                'tag_bg': tag_bg,
            })
    
    counts = {
        'all': len(requests_list),
        'portal': len([r for r in requests_list if r['exposed']]),
        'internal': len([r for r in requests_list if not r['exposed']]),
    }

    context = {
        'requests': requests_list,
        'counts': counts,
        'active_tab': 'all',
    }
    return render(request, 'dashboard/predefined_requests.html', context)
import json

from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.shortcuts import get_object_or_404

# Fix the imports - use the local require_permission function and ADMINS_GROUP constant
# from hotel_app.decorators import require_permission
# from hotel_app.groups import ADMINS_GROUP


@login_required
@require_permission([ADMINS_GROUP])
def configure_requests_api(request):
    """API endpoint to manage requests.

    Supports CRUD operations for requests.
    """
    if request.method == 'GET':
        # Return list of requests
        try:
            from hotel_app.models import RequestType
            requests = RequestType.objects.all().order_by('name')
            requests_list = []
            for r in requests:
                # Since RequestType doesn't have a department field, we'll use work_family or request_family as fallback
                department_name = 'Unknown Department'
                if hasattr(r, 'work_family') and r.work_family:
                    department_name = r.work_family.name
                elif hasattr(r, 'request_family') and r.request_family:
                    department_name = r.request_family.name
                    
                requests_list.append({
                    'id': r.id,
                    'title': r.name,
                    'department': department_name,
                    'description': r.description,
                    'active': r.active,
                })
            return JsonResponse({'requests': requests_list})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    elif request.method == 'POST':
        # Create a new request
        try:
            data = json.loads(request.body.decode('utf-8'))
            title = data.get('title')
            # Note: department_id is not used since RequestType doesn't have a department field
            description = data.get('description')
            active = data.get('active', True)

            if not title:
                return JsonResponse({'error': 'Title is required'}, status=400)

            from hotel_app.models import RequestType
            request_type = RequestType.objects.create(
                name=title,
                description=description,
                active=active,
            )
            return JsonResponse({'id': request_type.id, 'message': 'Request created successfully'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    elif request.method == 'PUT':
        # Update an existing request
        try:
            data = json.loads(request.body.decode('utf-8'))
            request_id = data.get('id')
            title = data.get('title')
            # Note: department_id is not used since RequestType doesn't have a department field
            description = data.get('description')
            active = data.get('active', True)

            if not request_id or not title:
                return JsonResponse({'error': 'ID and title are required'}, status=400)

            from hotel_app.models import RequestType
            request_type = get_object_or_404(RequestType, pk=request_id)
            request_type.name = title
            request_type.description = description
            request_type.active = active
            request_type.save()
            return JsonResponse({'id': request_type.id, 'message': 'Request updated successfully'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    elif request.method == 'DELETE':
        # Delete a request
        try:
            data = json.loads(request.body.decode('utf-8'))
            request_id = data.get('id')

            if not request_id:
                return JsonResponse({'error': 'ID is required'}, status=400)

            from hotel_app.models import RequestType
            request_type = get_object_or_404(RequestType, pk=request_id)
            request_type.delete()
            return JsonResponse({'message': 'Request deleted successfully'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    else:
        return JsonResponse({'error': 'Unsupported method'}, status=405)


def create_sample_service_requests():
    """Create sample service requests for testing purposes."""
    from hotel_app.models import ServiceRequest, RequestType, Department, Location, User
    from django.utils import timezone
    from datetime import timedelta
    import random
    
    # Create sample departments if they don't exist
    departments_data = [
        {'name': 'Housekeeping', 'description': 'Housekeeping services'},
        {'name': 'Maintenance', 'description': 'Maintenance services'},
        {'name': 'Guest Services', 'description': 'Guest services'},
    ]
    
    departments = []
    for dept_data in departments_data:
        dept, created = Department.objects.get_or_create(
            name=dept_data['name'],
            defaults={'description': dept_data['description']}
        )
        departments.append(dept)
    
    # Create sample request types if they don't exist
    request_types_data = [
        {'name': 'Room Cleaning'},
        {'name': 'AC Repair'},
        {'name': 'TV Issue'},
        {'name': 'Extra Towels'},
        {'name': 'Restaurant Reservation'},
    ]
    
    request_types = []
    for rt_data in request_types_data:
        rt, created = RequestType.objects.get_or_create(
            name=rt_data['name'],
            defaults={}
        )
        request_types.append(rt)
    
    # Create sample locations if they don't exist
    locations_data = [
        {'name': 'Room 101', 'room_no': '101'},
        {'name': 'Room 205', 'room_no': '205'},
        {'name': 'Room 304', 'room_no': '304'},
        {'name': 'Room 412', 'room_no': '412'},
        {'name': 'Lobby', 'room_no': 'Lobby'},
    ]
    
    locations = []
    for loc_data in locations_data:
        loc, created = Location.objects.get_or_create(
            name=loc_data['name'],
            defaults={'room_no': loc_data['room_no']}
        )
        locations.append(loc)
    
    # Get some users (use existing ones or create new ones)
    users = list(User.objects.all())
    if len(users) < 5:
        # Create some sample users if needed
        for i in range(5 - len(users)):
            user = User.objects.create_user(
                username=f'user{i}',
                email=f'user{i}@example.com',
                password='password123'
            )
            users.append(user)
    
    # Create sample service requests
    priorities = ['low', 'normal', 'high']
    statuses = ['pending', 'assigned', 'accepted', 'in_progress', 'completed', 'closed']
    
    sample_requests = [
        {
            'request_type': request_types[0],
            'location': locations[0],
            'requester_user': users[0],
            'priority': 'high',
            'status': 'in_progress',
        },
        {
            'request_type': request_types[1],
            'location': locations[2],
            'requester_user': users[1],
            'priority': 'high',
            'status': 'in_progress',
        },
        {
            'request_type': request_types[3],
            'location': locations[1],
            'requester_user': users[2],
            'priority': 'normal',
            'status': 'pending',
        },
        {
            'request_type': request_types[4],
            'location': locations[4],
            'requester_user': users[3],
            'priority': 'low',
            'status': 'completed',
        },
        {
            'request_type': request_types[2],
            'location': locations[3],
            'requester_user': users[4],
            'priority': 'high',
            'status': 'in_progress',
        },
    ]
    
    for req_data in sample_requests:
        # Randomly assign an assignee user (or leave unassigned)
        assignee = random.choice(users) if random.choice([True, False]) else None
        
        # Create the service request
        sr = ServiceRequest.objects.create(
            request_type=req_data['request_type'],
            location=req_data['location'],
            requester_user=req_data['requester_user'],
            assignee_user=assignee,
            priority=req_data['priority'],
            status=req_data['status'],
            notes='Sample request for testing',
        )
        
        # If status is completed, set a closed_at time
        if req_data['status'] == 'completed':
            sr.closed_at = timezone.now()
            sr.save()


@login_required
@require_permission([ADMINS_GROUP])
def configure_requests_api_fields(request, request_id):
    """API endpoint to manage request fields.

    Supports CRUD operations for request fields.
    """
    if request.method == 'GET':
        # Return list of fields for a request
        try:
            from hotel_app.models import RequestType, RequestTypeField
            request_type = get_object_or_404(RequestType, pk=request_id)
            fields = RequestTypeField.objects.filter(request_type=request_type).order_by('order')
            fields_list = [
                {
                    'id': f.id,
                    'label': f.label,
                    'type': f.type,
                    'required': f.required,
                    'order': f.order,
                }
                for f in fields
            ]
            return JsonResponse({'fields': fields_list})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    elif request.method == 'POST':
        # Create a new field for a request
        try:
            data = json.loads(request.body.decode('utf-8'))
            label = data.get('label')
            type = data.get('type')
            required = data.get('required')
            order = data.get('order')

            if not label or not type:
                return JsonResponse({'error': 'Label and type are required'}, status=400)

            from hotel_app.models import RequestType, RequestTypeField
            request_type = get_object_or_404(RequestType, pk=request_id)
            field = RequestTypeField.objects.create(
                request_type=request_type,
                label=label,
                type=type,
                required=required,
                order=order,
            )
            return JsonResponse({'id': field.id, 'message': 'Field created successfully'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    elif request.method == 'PUT':
        # Update an existing field for a request
        try:
            data = json.loads(request.body.decode('utf-8'))
            field_id = data.get('id')
            label = data.get('label')
            type = data.get('type')
            required = data.get('required')
            order = data.get('order')

            if not field_id or not label or not type:
                return JsonResponse({'error': 'ID, label, and type are required'}, status=400)

            from hotel_app.models import RequestTypeField
            field = get_object_or_404(RequestTypeField, pk=field_id)
            field.label = label
            field.type = type
            field.required = required
            field.order = order
            field.save()
            return JsonResponse({'id': field.id, 'message': 'Field updated successfully'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    elif request.method == 'DELETE':
        # Delete a field for a request
        try:
            data = json.loads(request.body.decode('utf-8'))
            field_id = data.get('id')

            if not field_id:
                return JsonResponse({'error': 'ID is required'}, status=400)

            from hotel_app.models import RequestTypeField
            field = get_object_or_404(RequestTypeField, pk=field_id)
            field.delete()
            return JsonResponse({'message': 'Field deleted successfully'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    else:
        return JsonResponse({'error': 'Unsupported method'}, status=405)


def create_sample_service_requests():
    """Create sample service requests for testing purposes."""
    from hotel_app.models import ServiceRequest, RequestType, Location, User
    from django.utils import timezone
    import random
    
    # Create sample request types if they don't exist
    request_types_data = [
        {'name': 'Room Cleaning'},
        {'name': 'AC Repair'},
        {'name': 'TV Issue'},
        {'name': 'Extra Towels'},
        {'name': 'Restaurant Reservation'},
    ]
    
    request_types = []
    for rt_data in request_types_data:
        rt, created = RequestType.objects.get_or_create(
            name=rt_data['name'],
            defaults={}
        )
        request_types.append(rt)
    
    # Create sample locations if they don't exist
    locations_data = [
        {'name': 'Room 101', 'room_no': '101'},
        {'name': 'Room 205', 'room_no': '205'},
        {'name': 'Room 304', 'room_no': '304'},
        {'name': 'Room 412', 'room_no': '412'},
        {'name': 'Lobby', 'room_no': 'Lobby'},
    ]
    
    locations = []
    for loc_data in locations_data:
        loc, created = Location.objects.get_or_create(
            name=loc_data['name'],
            defaults={'room_no': loc_data['room_no']}
        )
        locations.append(loc)
    
    # Get some users (use existing ones or create new ones)
    users = list(User.objects.all())
    if len(users) < 5:
        # Create some sample users if needed
        for i in range(5 - len(users)):
            user = User.objects.create_user(
                username=f'user{i}',
                email=f'user{i}@example.com',
                password='password123'
            )
            users.append(user)
    
    # Create sample service requests
    priorities = ['low', 'normal', 'high']
    statuses = ['pending', 'assigned', 'accepted', 'in_progress', 'completed', 'closed']
    
    sample_requests = [
        {
            'request_type': request_types[0],
            'location': locations[0],
            'requester_user': users[0],
            'priority': 'high',
            'status': 'in_progress',
        },
        {
            'request_type': request_types[1],
            'location': locations[2],
            'requester_user': users[1],
            'priority': 'high',
            'status': 'in_progress',
        },
        {
            'request_type': request_types[3],
            'location': locations[1],
            'requester_user': users[2],
            'priority': 'normal',
            'status': 'pending',
        },
        {
            'request_type': request_types[4],
            'location': locations[4],
            'requester_user': users[3],
            'priority': 'low',
            'status': 'completed',
        },
        {
            'request_type': request_types[2],
            'location': locations[3],
            'requester_user': users[4],
            'priority': 'high',
            'status': 'in_progress',
        },
    ]
    
    for req_data in sample_requests:
        # Randomly assign an assignee user (or leave unassigned)
        assignee = random.choice(users) if random.choice([True, False]) else None
        
        # Create the service request
        sr = ServiceRequest.objects.create(
            request_type=req_data['request_type'],
            location=req_data['location'],
            requester_user=req_data['requester_user'],
            assignee_user=assignee,
            priority=req_data['priority'],
            status=req_data['status'],
            notes='Sample request for testing',
        )
        
        # If status is completed, set a closed_at time
        if req_data['status'] == 'completed':
            sr.closed_at = timezone.now()
            sr.save()


@login_required
@require_permission([ADMINS_GROUP])
def configure_requests_api_bulk_action(request):
    """Bulk action endpoint for requests.

    Expects JSON body with 'action' and 'request_ids' list.
    Supported actions: enable, disable
    """
    try:
        body = json.loads(request.body.decode('utf-8'))
    except Exception:
        return HttpResponseBadRequest('invalid json')
    action = body.get('action')
    ids = body.get('request_ids') or []
    if action not in ('enable', 'disable'):
        return HttpResponseBadRequest('unsupported action')
    if not isinstance(ids, list):
        return HttpResponseBadRequest('request_ids must be list')
    requests = RequestType.objects.filter(id__in=ids)
    changed = []
    for r in requests:
        new_val = True if action == 'enable' else False
        if r.exposed != new_val:
            r.exposed = new_val
            r.save(update_fields=['exposed'])
            changed.append(r.id)
    return JsonResponse({'changed': changed, 'action': action})


@login_required
@require_permission([ADMINS_GROUP])
@csrf_protect
def api_bulk_permissions_update(request):
    """Update permissions for multiple user groups."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    
    try:
        from django.contrib.auth.models import Group, Permission
        
        # Parse JSON data
        try:
            data = json.loads(request.body.decode('utf-8'))
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
        
        # Get group IDs and permissions from request
        group_ids = data.get('group_ids', [])
        permissions = data.get('permissions', [])
        
        # Validate group IDs
        if not group_ids:
            return JsonResponse({'error': 'No groups specified'}, status=400)
        
        # Get permission objects from codenames
        permission_objects = []
        for codename in permissions:
            try:
                perm = Permission.objects.get(codename=codename)
                permission_objects.append(perm)
            except Permission.DoesNotExist:
                # Skip permissions that don't exist
                continue
        
        # Update permissions for each group
        updated_groups = []
        for group_id in group_ids:
            try:
                group = Group.objects.get(pk=group_id)
                group.permissions.set(permission_objects)
                updated_groups.append(group.name)
            except Group.DoesNotExist:
                continue  # Skip non-existent groups
        
        return JsonResponse({
            'success': True, 
            'message': f'Permissions updated for {len(updated_groups)} groups',
            'updated_groups': updated_groups
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@require_permission([ADMINS_GROUP, STAFF_GROUP])
def api_department_members(request, dept_id):
    """Return JSON list of members for a department (id, full_name, phone, email) with pagination support."""
    try:
        from hotel_app.models import UserProfile, Department
        dept = get_object_or_404(Department, pk=dept_id)
        
        # Get page and page size from query parameters
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 10))
        
        # Get all profiles for this department
        profiles = UserProfile.objects.filter(department=dept)
        
        # Calculate pagination
        total_profiles = profiles.count()
        total_pages = (total_profiles + page_size - 1) // page_size  # Ceiling division
        page = max(1, min(page, total_pages))  # Clamp page to valid range
        start_idx = (page - 1) * page_size
        end_idx = start_idx + page_size
        
        # Get profiles for current page
        page_profiles = profiles[start_idx:end_idx]
        
        members = []
        for p in page_profiles:
            members.append({
                'id': getattr(p, 'user_id', None), 
                'full_name': p.full_name, 
                'phone': p.phone, 
                'email': getattr(p, 'user', None).email if getattr(p, 'user', None) else None
            })
    except Exception as e:
        members = []
        total_profiles = 0
        total_pages = 1
        page = 1
    
    return JsonResponse({
        'members': members,
        'total': total_profiles,
        'page': page,
        'total_pages': total_pages
    })


@login_required
@require_permission([ADMINS_GROUP, STAFF_GROUP])
def api_group_members(request, group_id):
    """Return JSON list of members for a user group (id, username, email).

    Uses UserGroupMembership and User model where available.
    """
    try:
        from hotel_app.models import UserGroup, UserGroupMembership
        group = get_object_or_404(UserGroup, pk=group_id)
        memberships = UserGroupMembership.objects.filter(group=group).select_related('user')
        members = []
        for m in memberships:
            u = getattr(m, 'user', None)
            members.append({'id': getattr(u, 'pk', None), 'username': getattr(u, 'username', ''), 'email': getattr(u, 'email', '')})
    except Exception:
        members = []
    return JsonResponse({'members': members})

@login_required
def manage_users_roles(request):
    ctx = dict(active_tab="roles",
               breadcrumb_title="Roles & Permissions",
               page_title="Roles & Permissions",
               page_subtitle="Create and manage permission templates for your teams.",
               search_placeholder="Search roles...",
               primary_label="New Role")
    return render(request, 'dashboard/roles.html', ctx)

@login_required
def manage_users_profiles(request):
    # Get the standard Django groups (Admins, Staff, Users)
    from django.contrib.auth.models import Group
    groups = Group.objects.filter(name__in=['Admins', 'Staff', 'Users']).order_by('name')
    
    # Get user count for each group
    group_user_counts = {}
    for group in groups:
        count = group.user_set.count()
        group_user_counts[group.name] = count
    
    # Check user permissions to determine what they can see
    user_permissions = []
    if request.user.is_superuser:
        # Superuser has all permissions
        user_permissions = ['manage_users', 'manage_groups', 'system_config', 'view_reports', 'manage_departments', 'full_access']
    elif request.user.groups.filter(name='Admins').exists():
        # Admin permissions
        user_permissions = ['manage_users', 'manage_groups', 'view_reports', 'manage_departments']
    elif request.user.groups.filter(name='Staff').exists():
        # Staff permissions
        user_permissions = ['view_team_reports', 'manage_team', 'assign_requests', 'view_dept_data']
    else:
        # Basic user permissions
        user_permissions = ['view_profile', 'update_profile']
    
    ctx = dict(
        active_tab="profiles",
        breadcrumb_title="User Profiles",
        page_title="User Profiles",
        page_subtitle="Manage role templates and permissions for staff members",
        search_placeholder="Search profiles...",
        primary_label="Create Profile",
        groups=groups,
        group_user_counts=group_user_counts,
        user_permissions=user_permissions  # Pass user permissions to template
    )
    return render(request, 'dashboard/user_profiles.html', ctx)


@login_required

@require_http_methods(['POST'])
@csrf_protect
def user_create(request):
    """
    Create a user + profile.
    - role: mapped to permission flags (not saved in DB).
    - department: linked to Department by name (optional).
    - also associates the user to an existing UserGroup if its name matches the role or department (but does NOT create new groups to keep 'role not in DB').
    """
    data = request.POST

    username = (data.get("username") or "").strip()
    email = (data.get("email") or "").strip()
    full_name = (data.get("full_name") or "").strip()
    phone = (data.get("phone") or "").strip()
    department_name = (data.get("department") or "").strip()
    role = (data.get("role") or "").strip()
    password = (data.get("password") or "").strip()
    is_active = data.get("is_active") in ("1", "true", "True", "yes")

    errors = {}
    if not username:
        errors["username"] = ["Username is required."]
    if not email:
        errors["email"] = ["Email is required."]
    if not full_name:
        errors["full_name"] = ["Full name is required."]

    if errors:
        return JsonResponse({"success": False, "errors": errors}, status=400)

    # Ensure username uniqueness; if exists, make it unique
    base_username = username
    counter = 1
    while User.objects.filter(username=username).exists():
        username = f"{base_username}-{counter}"
        counter += 1

    dept_obj = None
    if department_name:
        dept_obj = Department.objects.filter(name__iexact=department_name).first()
        if not dept_obj:
            errors["department"] = [f"Department '{department_name}' not found."]
            return JsonResponse({"success": False, "errors": errors}, status=400)

    is_staff, is_superuser = _role_to_flags(role)

    try:
        with transaction.atomic():
            # Set password or generate a random one if blank
            if password:
                user_password = password
            else:
                user_password = User.objects.make_random_password()
            
            user = User.objects.create_user(
                username=username,
                email=email,
                password=user_password,
            )
            user.is_active = bool(is_active)
            user.is_staff = is_staff
            user.is_superuser = is_superuser
            user.save()

            # Create/attach user profile
            profile, created = UserProfile.objects.get_or_create(
                user=user,
                defaults={
                    'full_name': full_name,
                    'phone': phone or None,
                    'department': dept_obj,
                    'enabled': True,
                }
            )
            # If profile already existed, update its fields
            if not created:
                profile.full_name = full_name
                profile.phone = phone or None
                profile.department = dept_obj
                profile.enabled = True
                profile.save()

            # Assign user to the appropriate Django group based on role
            # First, remove user from all groups
            user.groups.clear()
            
            # Then add to the appropriate group based on role
            role_mapping = {
                'admin': 'Admins',
                'admins': 'Admins',
                'administrator': 'Admins',
                'superuser': 'Admins',
                'staff': 'Staff',
                'front desk': 'Staff',
                'front desk team': 'Staff',
                'user': 'Users',
                'users': 'Users'
            }
            
            group_name = role_mapping.get(role.lower(), 'Users')
            try:
                group = Group.objects.get(name=group_name)
                user.groups.add(group)
            except Group.DoesNotExist:
                # If the group doesn't exist, create it
                group = Group.objects.create(name=group_name)
                user.groups.add(group)

            # OPTIONAL: Attach to an existing group if one matches role or department (do not create new groups)
            candidate_group_names = []
            if role:
                candidate_group_names.append(role)
            if department_name:
                candidate_group_names.append(department_name)

            attached_group = None
            if candidate_group_names:
                attached_group = UserGroup.objects.filter(name__in=candidate_group_names).first()
                if attached_group:
                    UserGroupMembership.objects.get_or_create(user=user, group=attached_group)

            # Handle profile picture upload if provided (from FormData)
            profile_picture = request.FILES.get('profile_picture') if hasattr(request, 'FILES') else None
            if profile_picture:
                try:
                    # Create user directory if it doesn't exist
                    user_dir = os.path.join(settings.MEDIA_ROOT, 'users', str(user.pk))
                    os.makedirs(user_dir, exist_ok=True)

                    filename = f"profile_picture{os.path.splitext(profile_picture.name)[1]}"
                    file_path = os.path.join(user_dir, filename)

                    with open(file_path, 'wb+') as destination:
                        for chunk in profile_picture.chunks():
                            destination.write(chunk)

                    media_url = settings.MEDIA_URL or '/media/'
                    if not media_url.endswith('/'):
                        media_url = media_url + '/'
                    profile.avatar_url = f"{media_url}users/{user.pk}/{filename}"
                    profile.save(update_fields=['avatar_url'])
                except Exception:
                    # Non-fatal: continue but log if available
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.exception('Failed to save uploaded profile picture for user %s', user.pk)

    except Exception as e:
        # Surface clear error back to client
        return JsonResponse({"success": False, "errors": {"non_field_errors": [str(e)]}}, status=500)

    return JsonResponse({
        "success": True,
        "user": {
            "id": user.id,
            "username": user.username,
            "email": user.email,
            "is_active": user.is_active,
            "is_staff": user.is_staff,
            "is_superuser": user.is_superuser,
            "profile": {
                "full_name": profile.full_name,
                "phone": profile.phone,
                "department": dept_obj.name if dept_obj else None
            }
        }
    })


@require_http_methods(['POST'])
@require_http_methods(['POST'])
@csrf_protect
@require_permission([ADMINS_GROUP])
def department_create(request):
    """Create a department with optional logo."""
    form = DepartmentForm(request.POST, request.FILES)
    if form.is_valid():
        try:
            dept = form.save()
            
            # Handle logo upload if provided (from FormData)
            logo = request.FILES.get('logo') if hasattr(request, 'FILES') else None
            if logo:
                try:
                    # Create department directory if it doesn't exist
                    dept_dir = os.path.join(settings.MEDIA_ROOT, 'departments', str(dept.pk))
                    os.makedirs(dept_dir, exist_ok=True)

                    filename = f"logo{os.path.splitext(logo.name)[1]}"
                    file_path = os.path.join(dept_dir, filename)

                    with open(file_path, 'wb+') as destination:
                        for chunk in logo.chunks():
                            destination.write(chunk)

                    media_url = settings.MEDIA_URL or '/media/'
                    if not media_url.endswith('/'):
                        media_url = media_url + '/'
                    dept.logo = f"{media_url}departments/{dept.pk}/{filename}"
                    dept.save(update_fields=['logo'])
                except Exception:
                    # Non-fatal: continue but log if available
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.exception('Failed to save uploaded logo for department %s', dept.pk)

            return JsonResponse({
                "success": True, 
                "department": {
                    "id": dept.id, 
                    "name": dept.name,
                    "logo_url": dept.logo.url if dept.logo else None
                }
            })
        except Exception as e:
            return JsonResponse({"success": False, "errors": {"non_field_errors": [str(e)]}}, status=500)
    else:
        return JsonResponse({"success": False, "errors": form.errors}, status=400)


@require_permission([ADMINS_GROUP])
def department_update(request, dept_id):
    department = get_object_or_404(Department, pk=dept_id)
    if request.method == "POST":
        form = DepartmentForm(request.POST, request.FILES, instance=department)
        if form.is_valid():
            dept = form.save()
            
            # Handle logo upload if provided (from FormData)
            logo = request.FILES.get('logo') if hasattr(request, 'FILES') else None
            if logo:
                try:
                    # Create department directory if it doesn't exist
                    dept_dir = os.path.join(settings.MEDIA_ROOT, 'departments', str(dept.pk))
                    os.makedirs(dept_dir, exist_ok=True)

                    filename = f"logo{os.path.splitext(logo.name)[1]}"
                    file_path = os.path.join(dept_dir, filename)

                    with open(file_path, 'wb+') as destination:
                        for chunk in logo.chunks():
                            destination.write(chunk)

                    media_url = settings.MEDIA_URL or '/media/'
                    if not media_url.endswith('/'):
                        media_url = media_url + '/'
                    dept.logo = f"{media_url}departments/{dept.pk}/{filename}"
                    dept.save(update_fields=['logo'])
                except Exception:
                    # Non-fatal: continue but log if available
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.exception('Failed to save uploaded logo for department %s', dept.pk)
            
            messages.success(request, "Department updated successfully.")
        else:
            messages.error(request, "Error updating department. Please check the form.")
    return redirect("dashboard:departments")


@require_permission([ADMINS_GROUP, STAFF_GROUP])
def user_update(request, user_id):
    user = get_object_or_404(User, pk=user_id)
    if request.method == "POST":
        # Support both normal form posts and AJAX/Fetch FormData (which may include files)
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        full_name = request.POST.get('full_name', '').strip()
        phone = request.POST.get('phone', '').strip()
        title = request.POST.get('title', '').strip()
        department_id = request.POST.get('department', '').strip()
        is_active = request.POST.get('is_active', '0') == '1'
        role = request.POST.get('role', '').strip()

        # Update user fields
        if username:
            user.username = username
        if email:
            user.email = email
        user.is_active = is_active
        
        # Update staff and superuser flags based on role
        is_staff, is_superuser = _role_to_flags(role)
        user.is_staff = is_staff
        user.is_superuser = is_superuser
        
        user.save()

        # Update or create user profile
        profile, created = UserProfile.objects.get_or_create(user=user)
        profile.full_name = full_name
        profile.phone = phone
        profile.title = title

        # Handle department if provided
        if department_id:
            try:
                department = Department.objects.get(id=department_id)
                profile.department = department
            except Department.DoesNotExist:
                pass
        else:
            profile.department = None

        # Handle role assignment to Django groups
        if role:
            # First, remove user from all groups
            user.groups.clear()
            
            # Then add to the appropriate group based on role
            role_mapping = {
                'admin': 'Admins',
                'admins': 'Admins',
                'administrator': 'Admins',
                'superuser': 'Admins',
                'staff': 'Staff',
                'front desk': 'Staff',
                'front desk team': 'Staff',
                'user': 'Users',
                'users': 'Users'
            }
            
            group_name = role_mapping.get(role.lower(), 'Users')
            try:
                group = Group.objects.get(name=group_name)
                user.groups.add(group)
            except Group.DoesNotExist:
                # If the group doesn't exist, create it
                group = Group.objects.create(name=group_name)
                user.groups.add(group)

        # Handle profile picture upload
        profile_picture = request.FILES.get('profile_picture') if hasattr(request, 'FILES') else None
        if profile_picture:
            try:
                user_dir = os.path.join(settings.MEDIA_ROOT, 'users', str(user.pk))
                os.makedirs(user_dir, exist_ok=True)

                filename = f"profile_picture{os.path.splitext(profile_picture.name)[1]}"
                file_path = os.path.join(user_dir, filename)

                with open(file_path, 'wb+') as destination:
                    for chunk in profile_picture.chunks():
                        destination.write(chunk)

                media_url = settings.MEDIA_URL or '/media/'
                if not media_url.endswith('/'):
                    media_url = media_url + '/'
                profile.avatar_url = f"{media_url}users/{user.pk}/{filename}"
            except Exception:
                import logging
                logger = logging.getLogger(__name__)
                logger.exception('Failed to save uploaded profile picture for user %s', user.pk)

        profile.save()

        # Return JSON response for AJAX requests
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.META.get('CONTENT_TYPE', '').startswith('multipart/form-data'):
            return JsonResponse({
                'success': True,
                'message': 'User updated successfully!'
            })
        
        # For non-AJAX requests, add a message and redirect
        messages.success(request, "User updated successfully.")
        return redirect("dashboard:manage_user_detail", user_id=user_id)
    
    # For GET requests, redirect to user detail page
    return redirect("dashboard:manage_user_detail", user_id=user_id)


@require_permission([ADMINS_GROUP, STAFF_GROUP])
def manage_user_detail(request, user_id):
    """Render a dynamic full-page user detail / edit view for a single user.

    Context to template:
    - user: User instance
    - profile: UserProfile or None
    - groups: Queryset of Group objects
    - departments: Queryset of Department objects
    - avatar_url: str or None
    - requests_handled, messages_sent, avg_rating, response_rate
    """
    user = get_object_or_404(User, pk=user_id)
    profile = getattr(user, 'userprofile', None)
    groups = user.groups.all()
    try:
        departments = Department.objects.all()
    except Exception:
        departments = []

    # Basic stats (defensive)
    try:
        requests_handled = ServiceRequest.objects.filter(assignee_user=user).count()
    except Exception:
        requests_handled = 0

    # placeholder for messages_sent (if you have a messaging model, replace this)
    messages_sent = 0

    try:
        avg_rating = Review.objects.aggregate(Avg("rating"))["rating__avg"] or 0
    except Exception:
        avg_rating = 0

    try:
        closed_count = ServiceRequest.objects.filter(assignee_user=user, status__in=['closed', 'resolved', 'completed']).count()
        total_assigned = ServiceRequest.objects.filter(assignee_user=user).count()
        response_rate = (closed_count / total_assigned) if total_assigned else 0.98
    except Exception:
        response_rate = 0.98

    avatar_url = None
    if profile and getattr(profile, 'avatar_url', None):
        avatar_url = profile.avatar_url

    # Determine user role based on groups
    user_role = "User"
    if user.is_superuser:
        user_role = "Admin"
    elif user.is_staff and groups.filter(name="Staff").exists():
        user_role = "Staff"
    elif groups.filter(name="Users").exists():
        user_role = "User"

    context = {
        'user': user,
        'profile': profile,
        'groups': groups,
        'departments': departments,
        'avatar_url': avatar_url,
        'requests_handled': requests_handled,
        'messages_sent': messages_sent,
        'avg_rating': round(avg_rating, 1) if avg_rating else 0,
        'response_rate': int(response_rate * 100) if isinstance(response_rate, float) else response_rate,
        'user_role': user_role,
    }
    # Build a simple mapping of group name -> permission names to render in template
    try:
        group_permissions = {}
        for g in groups:
            perms = g.permissions.all()
            group_permissions[g.name] = [p.name for p in perms]
        # Attach JSON string for template consumption
        context['group_permissions_json'] = json.dumps(group_permissions)
    except Exception:
        context['group_permissions_json'] = json.dumps({})

    return render(request, 'dashboard/manage_user_detail.html', context)


@require_permission([ADMINS_GROUP])
def user_delete(request, user_id):
    user = get_object_or_404(User, pk=user_id)
    if request.method == "POST":
        user.delete()
        messages.success(request, "User deleted successfully.")
    return redirect("dashboard:users")


@require_permission([ADMINS_GROUP])
def manage_users_toggle_enabled(request, user_id):
    """Toggle the 'enabled' flag on a user's UserProfile. Expects POST."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    user = get_object_or_404(User, pk=user_id)
    profile = getattr(user, 'userprofile', None)
    if not profile:
        return JsonResponse({'error': 'UserProfile missing'}, status=400)
    profile.enabled = not bool(profile.enabled)
    profile.save(update_fields=['enabled'])
    return JsonResponse({'id': user.pk, 'enabled': profile.enabled})


# ---- Department Management ----
@require_permission([ADMINS_GROUP, STAFF_GROUP])
def dashboard_departments(request):
    # Keep existing department queryset for list rendering and metrics
    depts_qs = Department.objects.all().annotate(user_count=Count("userprofile")).order_by('name')
    # Server-side status filter (optional): active / paused / archived
    status = request.GET.get('status', '').lower()
    if status:
        if status == 'active':
            depts_qs = depts_qs.filter(user_count__gt=0)
        elif status == 'archived':
            depts_qs = depts_qs.filter(user_count__lte=0)
        elif status == 'paused':
            depts_qs = depts_qs.filter(user_count__gt=0, user_count__lte=2)

    # Simple pagination
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    page = request.GET.get('page', 1)
    paginator = Paginator(depts_qs, 10)  # 10 departments per page
    try:
        depts_page = paginator.page(page)
    except PageNotAnInteger:
        depts_page = paginator.page(1)
    except EmptyPage:
        depts_page = paginator.page(paginator.num_pages)
    form = DepartmentForm()

    # Build a serializable list for the template with featured_group (matching the template expectations)
    departments = []
    try:
        from hotel_app.models import UserProfile
        for index, d in enumerate(depts_page):
            profiles = UserProfile.objects.filter(department=d)
            members = []
            lead = None
            for p in profiles:
                members.append({'user_id': getattr(p, 'user_id', None), 'full_name': p.full_name, 'email': getattr(p, 'user', None).email if getattr(p, 'user', None) else None, 'avatar_url': getattr(p, 'avatar_url', None)})
                if p.title and 'department head' in p.title.lower():
                    lead = {'user_id': getattr(p, 'user_id', None), 'full_name': p.full_name, 'email': getattr(p, 'user', None).email if getattr(p, 'user', None) else None, 'avatar_url': getattr(p, 'avatar_url', None)}

            # Get logo URL if available using the new method
            logo_url = d.get_logo_url()
            if logo_url:
                image = logo_url
            else:
                # Provide proper fallback icons based on department name
                dept_name_slug = d.name.lower().replace(" ", "_").replace("-", "_").replace("&", "_").replace("___", "_")
                # Map common department names to their icons
                icon_mapping = {
                    'front_office': 'front_office.svg',
                    'housekeeping': 'housekeeping.svg',
                    'food_beverage': 'food_beverage.svg',
                    'food&beverage': 'food_beverage.svg',
                    'food_&_beverage': 'food_beverage.svg',
                    'security': 'security.svg',
                    'maintenance': 'maintainence.svg',
                    'it': 'name.svg',
                    'hr': 'name.svg',
                    'finance': 'name.svg',
                    'marketing': 'name.svg',
                    'sales': 'name.svg',
                }
                # Try to find a matching icon or use default
                icon_file = icon_mapping.get(dept_name_slug, 'name.svg')
                image = f'images/manage_users/{icon_file}'

            icon_bg = 'bg-gray-500/10'
            tag_bg = 'bg-gray-500/10'
            icon_color = 'gray-500'
            dot_bg = 'bg-gray-500'

            # Dummy metrics
            members_count = profiles.count()
            open_tickets = 0
            performance_pct = f"{min(100, 50 + members_count)}%"
            performance_color = 'green-500' if members_count > 5 else 'yellow-400'
            performance_width = '8' if members_count > 5 else '4'

            sla_label = 'Good' if members_count > 5 else 'Monitor'
            sla_tag_bg = 'bg-green-100' if sla_label == 'Good' else 'bg-yellow-100'
            sla_color = 'green-700' if sla_label == 'Good' else 'yellow-700'

            status_label = 'Active' if members_count > 0 else 'Inactive'
            status_bg = 'bg-green-500/10' if members_count > 0 else 'bg-gray-200'
            status_color = 'green-500' if members_count > 0 else 'gray-500'

            featured_group = {
                'id': d.pk,
                'name': d.name,
                'description': d.description or 'Department description',
                'members_count': members_count,
                'image': image,
                'icon_bg': icon_bg,
                'tag_bg': tag_bg,
                'icon_color': icon_color,
                'dot_bg': dot_bg,
                'position_top': index * 270,
            }

            # Attach groups info for groups template (best-effort)
            try:
                groups_qs = d.user_groups.all()
                groups_list = []
                for g in groups_qs:
                    groups_list.append({'pk': g.pk, 'name': g.name, 'members_count': getattr(g, 'members_count', 0), 'dot_bg': 'bg-green-500'})
                featured_group['groups'] = groups_list
            except Exception:
                featured_group['groups'] = []

            departments.append({
                'featured_group': featured_group,
                'members': members,
                'lead': lead,
                'open_tickets': open_tickets,
                'sla_label': sla_label,
                'sla_tag_bg': sla_tag_bg,
                'sla_color': sla_color,
                'performance_pct': performance_pct,
                'performance_color': performance_color,
                'performance_width': performance_width,
                'status_label': status_label,
                'status_bg': status_bg,
                'status_color': status_color,
            })
    except Exception:
        # fallback to simple data matching the expected structure
        for index, d in enumerate(depts_page):
            # Get logo URL if available using the new method
            logo_url = d.get_logo_url()
            if logo_url:
                image = logo_url
            else:
                # Provide proper fallback icons based on department name
                dept_name_slug = d.name.lower().replace(" ", "_").replace("-", "_").replace("&", "_").replace("___", "_")
                # Map common department names to their icons
                icon_mapping = {
                    'front_office': 'front_office.svg',
                    'housekeeping': 'housekeeping.svg',
                    'food_beverage': 'food_beverage.svg',
                    'food&beverage': 'food_beverage.svg',
                    'food_&_beverage': 'food_beverage.svg',
                    'security': 'security.svg',
                    'maintenance': 'maintainence.svg',
                    'it': 'name.svg',
                    'hr': 'name.svg',
                    'finance': 'name.svg',
                    'marketing': 'name.svg',
                    'sales': 'name.svg',
                }
                # Try to find a matching icon or use default
                icon_file = icon_mapping.get(dept_name_slug, 'name.svg')
                image = f'images/manage_users/{icon_file}'
            featured_group = {'id': getattr(d, 'pk', ''), 'name': getattr(d, 'name', ''), 'description': getattr(d, 'description', '') or '', 'members_count': getattr(d, 'user_count', 0), 'image': image, 'icon_bg': 'bg-gray-500/10', 'tag_bg': 'bg-gray-500/10', 'icon_color': 'gray-500', 'dot_bg': 'bg-gray-500', 'position_top': index * 270}
            departments.append({'featured_group': featured_group, 'members': [], 'lead': None, 'open_tickets': 0, 'sla_label': 'N/A', 'sla_tag_bg': 'bg-gray-200', 'sla_color': 'gray-600', 'performance_pct': '0%', 'performance_color': 'gray-500', 'performance_width': '2', 'status_label': 'Unknown', 'status_bg': 'bg-gray-200', 'status_color': 'gray-500'})

    # Render the Manage Users base template when navigated via the Manage Users tabs.
    # The template expects `active_tab` to determine which header/content to show.
    context = {
        "departments": departments,
        "page_obj": depts_page,
        "paginator": paginator,
        "total_departments": depts_qs.count(),
        "form": form,
        "active_tab": 'departments',
        "crumb_section": 'Admin',
        "crumb_title": 'Departments',
        "title": 'Manage Departments',
        "subtitle": 'Manage hotel departments, heads, and staff assignments',
        "primary_label": "Add Department",
    }
    return render(request, "dashboard/manage_users_base.html", context)


@require_permission([ADMINS_GROUP])
@require_http_methods(['POST'])
def add_group_member(request, group_id):
    """Add a user to a UserGroup and ensure their UserProfile.department is set to the group's department."""
    try:
        from hotel_app.models import UserGroup, UserGroupMembership, UserProfile
        group = get_object_or_404(UserGroup, pk=group_id)
    except Exception:
        return JsonResponse({'error': 'group not found'}, status=404)

    user_id = request.POST.get('user_id') or request.POST.get('id')
    if not user_id:
        return JsonResponse({'error': 'user_id required'}, status=400)

    try:
        user = User.objects.get(pk=int(user_id))
    except Exception:
        return JsonResponse({'error': 'user not found'}, status=404)

    # create membership if not exists
    membership, created = UserGroupMembership.objects.get_or_create(user=user, group=group)

    # ensure user's profile department set to group's department
    profile = getattr(user, 'userprofile', None)
    if not profile:
        from hotel_app.models import UserProfile as UP
        profile = UP.objects.create(user=user, full_name=(user.get_full_name() or user.username))

    if group.department and profile.department_id != group.department_id:
        profile.department = group.department
        profile.save(update_fields=['department'])

    return JsonResponse({'success': True, 'created': created, 'user_id': user.pk, 'group_id': group.pk})


@require_permission([ADMINS_GROUP])
@require_http_methods(['POST'])
def remove_group_member(request, group_id):
    """Remove a user from a UserGroup. Does not change department membership automatically."""
    try:
        from hotel_app.models import UserGroup, UserGroupMembership
        group = get_object_or_404(UserGroup, pk=group_id)
    except Exception:
        return JsonResponse({'error': 'group not found'}, status=404)

    user_id = request.POST.get('user_id') or request.POST.get('id')
    if not user_id:
        return JsonResponse({'error': 'user_id required'}, status=400)

    try:
        user = User.objects.get(pk=int(user_id))
    except Exception:
        return JsonResponse({'error': 'user not found'}, status=404)

    try:
        membership = UserGroupMembership.objects.get(user=user, group=group)
        membership.delete()
    except UserGroupMembership.DoesNotExist:
        return JsonResponse({'error': 'membership not found'}, status=404)

    return JsonResponse({'success': True, 'user_id': user.pk, 'group_id': group.pk})

@require_http_methods(['POST'])
@csrf_protect
@require_permission([ADMINS_GROUP])
def department_create(request):
    """Create a department with optional logo, head, and assigned staff."""
    form = DepartmentForm(request.POST, request.FILES)
    if form.is_valid():
        try:
            dept = form.save()
            
            # Handle additional fields
            head_id = request.POST.get('head')
            email = request.POST.get('email')
            assigned_staff_ids = request.POST.getlist('assigned_staff')
            
            # Set department head if provided
            if head_id:
                try:
                    head_user = User.objects.get(id=head_id)
                    # Create or update user profile to set department
                    profile, created = UserProfile.objects.get_or_create(user=head_user)
                    profile.department = dept
                    profile.title = profile.title or 'Department Head'
                    profile.save()
                    
                    # Set the department head (if your Department model supports this)
                    # dept.head = head_user
                    # dept.save(update_fields=['head'])
                except User.DoesNotExist:
                    pass
            
            # Assign staff to department
            if assigned_staff_ids:
                for user_id in assigned_staff_ids:
                    try:
                        user = User.objects.get(id=user_id)
                        profile, created = UserProfile.objects.get_or_create(user=user)
                        profile.department = dept
                        # Only set title if not already set
                        if not profile.title:
                            profile.title = 'Staff'
                        profile.save()
                    except User.DoesNotExist:
                        continue
            
            # Handle logo upload if provided (from FormData)
            logo = request.FILES.get('logo') if hasattr(request, 'FILES') else None
            if logo:
                try:
                    # Create department directory if it doesn't exist
                    dept_dir = os.path.join(settings.MEDIA_ROOT, 'departments', str(dept.pk))
                    os.makedirs(dept_dir, exist_ok=True)

                    filename = f"logo{os.path.splitext(logo.name)[1]}"
                    file_path = os.path.join(dept_dir, filename)

                    with open(file_path, 'wb+') as destination:
                        for chunk in logo.chunks():
                            destination.write(chunk)

                    media_url = settings.MEDIA_URL or '/media/'
                    if not media_url.endswith('/'):
                        media_url = media_url + '/'
                    dept.logo = f"{media_url}departments/{dept.pk}/{filename}"
                    dept.save(update_fields=['logo'])
                except Exception:
                    # Non-fatal: continue but log if available
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.exception('Failed to save uploaded logo for department %s', dept.pk)

            return JsonResponse({
                "success": True, 
                "department": {
                    "id": dept.id, 
                    "name": dept.name,
                    "logo_url": dept.logo.url if dept.logo else None
                }
            })
        except Exception as e:
            return JsonResponse({"success": False, "errors": {"non_field_errors": [str(e)]}}, status=500)
    else:
        return JsonResponse({"success": False, "errors": form.errors}, status=400)

@require_permission([ADMINS_GROUP])
def department_update(request, dept_id):
    department = get_object_or_404(Department, pk=dept_id)
    if request.method == "POST":
        form = DepartmentForm(request.POST, request.FILES, instance=department)
        if form.is_valid():
            dept = form.save()
            
            # Handle logo upload if provided (from FormData)
            logo = request.FILES.get('logo') if hasattr(request, 'FILES') else None
            if logo:
                try:
                    # Create department directory if it doesn't exist
                    dept_dir = os.path.join(settings.MEDIA_ROOT, 'departments', str(dept.pk))
                    os.makedirs(dept_dir, exist_ok=True)

                    filename = f"logo{os.path.splitext(logo.name)[1]}"
                    file_path = os.path.join(dept_dir, filename)

                    with open(file_path, 'wb+') as destination:
                        for chunk in logo.chunks():
                            destination.write(chunk)

                    media_url = settings.MEDIA_URL or '/media/'
                    if not media_url.endswith('/'):
                        media_url = media_url + '/'
                    dept.logo = f"{media_url}departments/{dept.pk}/{filename}"
                    dept.save(update_fields=['logo'])
                except Exception:
                    # Non-fatal: continue but log if available
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.exception('Failed to save uploaded logo for department %s', dept.pk)
            
            messages.success(request, "Department updated successfully.")
        else:
            messages.error(request, "Error updating department. Please check the form.")
    return redirect("dashboard:departments")

@require_permission([ADMINS_GROUP])
def department_delete(request, dept_id):
    department = get_object_or_404(Department, pk=dept_id)
    if request.method == "POST":
        department.delete()
        messages.success(request, "Department deleted successfully.")
    return redirect("dashboard:departments")
@require_http_methods(['POST'])
def assign_department_lead(request, dept_id):
    """Assign a department lead.

    Expects POST body form data: user_id (int)
    Sets the chosen user's UserProfile.department to dept and title to 'Department Head'.
    Clears the title on any other profile in the same department previously marked as Department Head.
    Returns JSON with success and lead info.
    """
    try:
        dept = get_object_or_404(Department, pk=dept_id)
    except Exception:
        return JsonResponse({'error': 'department not found'}, status=404)

    user_id = request.POST.get('user_id') or request.POST.get('lead_user_id')
    if not user_id:
        return JsonResponse({'error': 'user_id required'}, status=400)

    try:
        user = User.objects.get(pk=int(user_id))
    except Exception:
        return JsonResponse({'error': 'user not found'}, status=404)

    profile = getattr(user, 'userprofile', None)
    if not profile:
        # If no profile exists, create a minimal one
        from hotel_app.models import UserProfile
        profile = UserProfile.objects.create(user=user, full_name=(user.get_full_name() or user.username))

    # Clear existing leads in this department (best-effort)
    try:
        from hotel_app.models import UserProfile as UP
        previous_leads = UP.objects.filter(department=dept, title__icontains='Department Head').exclude(user=user)
        for pl in previous_leads:
            pl.title = ''
            pl.save(update_fields=['title'])
    except Exception:
        # ignore if model shape differs
        pass

    # Assign selected user as department lead
    profile.department = dept
    profile.title = 'Department Head'
    profile.save(update_fields=['department', 'title'])

    return JsonResponse({'success': True, 'lead': {'user_id': user.pk, 'full_name': profile.full_name, 'department': dept.name}})


# ---- Group Management ----
@require_permission([ADMINS_GROUP])
def dashboard_groups(request):
    groups = Group.objects.all().annotate(user_count=Count("user"))
    form = GroupForm()
    context = {
        "groups": groups,
        "form": form,
    }
    return render(request, "dashboard/groups.html", context)

@require_http_methods(['POST'])
@csrf_protect
def group_create(request):
    """
    Create a user group. Department is matched by name if provided.
    If department is provided, the group name will be formatted as "Department - Group Name".
    """
    name = (request.POST.get("name") or "").strip()
    description = (request.POST.get("description") or "").strip()
    department_name = (request.POST.get("department") or "").strip()

    errors = {}
    if not name:
        errors["name"] = ["Group name is required."]
    
    dept_obj = None
    final_group_name = name
    if department_name:
        dept_obj = Department.objects.filter(name__iexact=department_name).first()
        if not dept_obj:
            errors["department"] = [f"Department '{department_name}' not found."]
        else:
            # Format group name as "Department - Group Name"
            final_group_name = f"{dept_obj.name} - {name}"
    
    # Check if a group with the final name already exists
    if UserGroup.objects.filter(name__iexact=final_group_name).exists():
        errors["name"] = [f"Group '{final_group_name}' already exists."]

    if errors:
        return JsonResponse({"success": False, "errors": errors}, status=400)

    try:
        grp = UserGroup.objects.create(
            name=final_group_name,
            description=description or None,
            department=dept_obj
        )
    except Exception as e:
        return JsonResponse({"success": False, "errors": {"non_field_errors": [str(e)]}}, status=500)

    # Return additional information about the group including department
    response_data = {
        "success": True, 
        "group": {
            "id": grp.id, 
            "name": grp.name,
            "description": grp.description,
            "department_id": grp.department.id if grp.department else None,
            "department_name": grp.department.name if grp.department else None
        }
    }
    return JsonResponse(response_data)

@require_permission([ADMINS_GROUP])
def group_update(request, group_id):
    group = get_object_or_404(Group, pk=group_id)
    if request.method == "POST":
        form = GroupForm(request.POST, instance=group)
        if form.is_valid():
            form.save()
            messages.success(request, "Group updated successfully.")
    return redirect("dashboard:groups")

@require_permission([ADMINS_GROUP])
def group_delete(request, group_id):
    group = get_object_or_404(Group, pk=group_id)
    if request.method == "POST":
        group.delete()
        messages.success(request, "Group deleted successfully.")
    return redirect("dashboard:groups")


# ---- Location Management ----
@require_permission([ADMINS_GROUP, STAFF_GROUP])
def dashboard_locations(request):
    locations = Location.objects.all().select_related("building", "floor", "type")
    form = LocationForm()
    context = {
        "locations": locations,
        "form": form,
    }
    return render(request, "dashboard/locations.html", context)

@require_permission([ADMINS_GROUP])
def location_create(request):
    if request.method == "POST":
        form = LocationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Location created successfully.")
    return redirect("dashboard:locations")

@require_permission([ADMINS_GROUP])
def location_update(request, loc_id):
    location = get_object_or_404(Location, pk=loc_id)
    if request.method == "POST":
        form = LocationForm(request.POST, instance=location)
        if form.is_valid():
            form.save()
            messages.success(request, "Location updated successfully.")
    return redirect("dashboard:locations")

@require_permission([ADMINS_GROUP])
def location_delete(request, loc_id):
    location = get_object_or_404(Location, pk=loc_id)
    if request.method == "POST":
        location.delete()
        messages.success(request, "Location deleted successfully.")
    return redirect("dashboard:locations")



@require_permission([ADMINS_GROUP])
def request_type_create(request):
    if request.method == "POST":
        form = RequestTypeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Request Type created successfully.")
    return redirect("dashboard:request_types")

@require_permission([ADMINS_GROUP])
def request_type_update(request, rt_id):
    request_type = get_object_or_404(RequestType, pk=rt_id)
    if request.method == "POST":
        form = RequestTypeForm(request.POST, instance=request_type)
        if form.is_valid():
            form.save()
            messages.success(request, "Request Type updated successfully.")
    return redirect("dashboard:request_types")

@require_permission([ADMINS_GROUP])
def request_type_delete(request, rt_id):
    request_type = get_object_or_404(RequestType, pk=rt_id)
    if request.method == "POST":
        request_type.delete()
        messages.success(request, "Request Type deleted successfully.")
    return redirect("dashboard:request_types")


# ---- Checklist Management ----


@require_permission([ADMINS_GROUP])
def checklist_create(request):
    if request.method == "POST":
        form = ChecklistForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Checklist created successfully.")
    return redirect("dashboard:checklists")

@require_permission([ADMINS_GROUP])
def checklist_update(request, cl_id):
    checklist = get_object_or_404(Checklist, pk=cl_id)
    if request.method == "POST":
        form = ChecklistForm(request.POST, instance=checklist)
        if form.is_valid():
            form.save()
            messages.success(request, "Checklist updated successfully.")
    return redirect("dashboard:checklists")

@require_permission([ADMINS_GROUP])
def checklist_delete(request, cl_id):
    checklist = get_object_or_404(Checklist, pk=cl_id)
    if request.method == "POST":
        checklist.delete()
        messages.success(request, "Checklist deleted successfully.")
    return redirect("dashboard:checklists")



@require_permission([ADMINS_GROUP])
def complaint_create(request):
    if request.method == "POST":
        form = ComplaintForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Complaint logged successfully.")
    return redirect("dashboard:complaints")

@require_permission([ADMINS_GROUP])
def complaint_update(request, complaint_id):
    complaint = get_object_or_404(Complaint, pk=complaint_id)
    if request.method == "POST":
        form = ComplaintForm(request.POST, instance=complaint)
        if form.is_valid():
            form.save()
            messages.success(request, "Complaint updated successfully.")
    return redirect("dashboard:complaints")

@require_permission([ADMINS_GROUP])
def complaint_delete(request, complaint_id):
    complaint = get_object_or_404(Complaint, pk=complaint_id)
    if request.method == "POST":
        complaint.delete()
        messages.success(request, "Complaint deleted successfully.")
    return redirect("dashboard:complaints")


# ---- Review Management ----

@require_permission([ADMINS_GROUP])
def review_create(request):
    if request.method == "POST":
        form = ReviewForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Review submitted successfully.")
    return redirect("dashboard:reviews")

@require_permission([ADMINS_GROUP])
def review_update(request, review_id):
    review = get_object_or_404(Review, pk=review_id)
    if request.method == "POST":
        form = ReviewForm(request.POST, instance=review)
        if form.is_valid():
            form.save()
            messages.success(request, "Review updated successfully.")
    return redirect("dashboard:reviews")

@require_permission([ADMINS_GROUP])
def review_delete(request, review_id):
    review = get_object_or_404(Review, pk=review_id)
    if request.method == "POST":
        review.delete()
        messages.success(request, "Review deleted successfully.")
    return redirect("dashboard:reviews")


# ---- New Voucher System: Analytics ----
@require_permission([ADMINS_GROUP, STAFF_GROUP])
def sla_escalations(request):
    """SLA & Escalations dashboard."""
    context = {
        'active_tab': 'sla_escalations',
        'title': 'SLA & Escalations',
        'subtitle': 'Define service level agreements and escalation workflows for guest requests',
    }
    return render(request, 'dashboard/sla_escalations.html', context)


@require_permission([ADMINS_GROUP, STAFF_GROUP])
def voucher_analytics(request):
    """Voucher analytics dashboard with actual data."""
    today = timezone.now().date()
    
    total_vouchers = Voucher.objects.count()
    active_vouchers = Voucher.objects.filter(status='active').count()
    redeemed_vouchers = Voucher.objects.filter(status='redeemed').count()
    expired_vouchers = Voucher.objects.filter(status='expired').count()
    redeemed_today = VoucherScan.objects.filter(scanned_at__date=today, redemption_successful=True).count()
    
    vouchers_by_type = dict(
        Voucher.objects.values('voucher_type').annotate(count=Count('id')).values_list('voucher_type', 'count')
    )
    
    recent_vouchers = Voucher.objects.select_related('guest').order_by('-created_at')[:20]
    recent_scans = VoucherScan.objects.select_related('voucher', 'voucher__guest', 'scanned_by').order_by('-scanned_at')[:10]
    
    # Peak redemption hours using Django ORM's TruncHour
    peak_hours_data = list(
        VoucherScan.objects.filter(redemption_successful=True)
        .annotate(hour_truncated=TruncHour('scanned_at'))
        .values('hour_truncated')
        .annotate(count=Count('id'))
        .order_by('hour_truncated')
    )
    # Reformat for chart
    peak_hours = [{'hour': item['hour_truncated'].hour, 'count': item['count']} for item in peak_hours_data if item['hour_truncated']]

    analytics_data = {
        'total_vouchers': total_vouchers,
        'active_vouchers': active_vouchers,
        'redeemed_vouchers': redeemed_vouchers,
        'expired_vouchers': expired_vouchers,
        'redeemed_today': redeemed_today,
        'vouchers_by_type': vouchers_by_type,
        'peak_hours': peak_hours,
    }
    
    context = {
        'analytics': analytics_data,
        'analytics_json': json.dumps(analytics_data),
        'recent_vouchers': recent_vouchers,
        'recent_scans': recent_scans,
    }
    return render(request, "dashboard/voucher_analytics.html", context)


@login_required
@require_role(['admin', 'staff'])
def analytics_dashboard(request):
    from django.db.models import Avg, Count
    from datetime import datetime, timedelta
    import json
    
    # Date range for analytics (last 30 days)
    today = datetime.now().date()
    thirty_days_ago = today - timedelta(days=30)
    
    # Ticket volume trends (last 30 days, grouped by day)
    ticket_trends = []
    ticket_dates = []
    
    for i in range(30):
        date = thirty_days_ago + timedelta(days=i)
        count = ServiceRequest.objects.filter(created_at__date=date).count()
        ticket_trends.append(count)
        ticket_dates.append(date.strftime('%b %d'))
    
    # Feedback volume trends (last 30 days, grouped by day)
    feedback_trends = []
    feedback_dates = []
    
    for i in range(30):
        date = thirty_days_ago + timedelta(days=i)
        count = Review.objects.filter(created_at__date=date).count()
        feedback_trends.append(count)
        feedback_dates.append(date.strftime('%b %d'))
    
    # Guest satisfaction score over time (last 4 weeks)
    satisfaction_scores = []
    satisfaction_weeks = []
    
    for i in range(4):
        week_start = today - timedelta(days=today.weekday()) - timedelta(weeks=3-i)
        week_end = week_start + timedelta(days=7)
        reviews = Review.objects.filter(created_at__date__gte=week_start, created_at__date__lt=week_end)
        avg_rating = reviews.aggregate(Avg('rating'))['rating__avg'] or 0
        satisfaction_weeks.append(f'Week {i+1}')
        satisfaction_scores.append(round(avg_rating, 1))
    
    # Department performance data
    departments = Department.objects.all()
    dept_performance = []
    
    for dept in departments:
        dept_requests = ServiceRequest.objects.filter(department=dept)
        avg_resolution_time = 0
        avg_satisfaction = 0
        
        if dept_requests.exists():
            # Calculate average resolution time
            resolved_requests = dept_requests.filter(status='completed')
            if resolved_requests.exists():
                total_resolution_time = timedelta()
                for req in resolved_requests:
                    if req.completed_at and req.created_at:
                        total_resolution_time += (req.completed_at - req.created_at)
                avg_resolution_time = total_resolution_time.total_seconds() / 3600 / resolved_requests.count()  # in hours
            
            # Calculate average satisfaction
            # Since there's no direct relationship between ServiceRequest and Review,
            # we'll use all reviews for now. In a real implementation, you would need
            # to establish a proper relationship between requests and reviews.
            dept_reviews = Review.objects.all()
            avg_satisfaction = dept_reviews.aggregate(Avg('rating'))['rating__avg'] or 0
        
        dept_performance.append({
            'name': dept.name,
            'resolution_time': round(avg_resolution_time, 1),
            'satisfaction': round(avg_satisfaction, 1)
        })
    
    # Room type feedback distribution
    room_types = ['Standard', 'Deluxe', 'Suite', 'Executive']
    room_feedback = []
    
    for room_type in room_types:
        # This is sample data - in a real implementation, you would join with actual room data
        reviews = Review.objects.filter(created_at__date__gte=thirty_days_ago)
        avg_rating = reviews.aggregate(Avg('rating'))['rating__avg'] or 0
        room_feedback.append({
            'type': room_type,
            'satisfaction': round(avg_rating, 1)
        })
    
    # Busiest hours heatmap data (sample data for demonstration)
    busiest_hours_data = []
    days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
    hours = list(range(24))
    
    for day in days:
        for hour in hours:
            # Generate sample data - in a real implementation, you would query actual data
            value = (hour * 2 + days.index(day)) % 20  # Sample calculation
            busiest_hours_data.append({
                'day': day,
                'hour': hour,
                'value': value
            })
    
    # Overall statistics
    total_tickets = ServiceRequest.objects.count()
    total_reviews = Review.objects.count()
    avg_rating = Review.objects.aggregate(Avg('rating'))['rating__avg'] or 0
    completed_tickets = ServiceRequest.objects.filter(status='completed').count()
    completion_rate = (completed_tickets / total_tickets * 100) if total_tickets > 0 else 0
    
    # Top performing departments
    top_departments = sorted(dept_performance, key=lambda x: x['satisfaction'], reverse=True)[:3]
    
    # Recent activity
    recent_tickets = ServiceRequest.objects.select_related('request_type', 'department').order_by('-created_at')[:5]
    recent_reviews = Review.objects.select_related('guest').order_by('-created_at')[:5]
    
    # Scheduled reports data
    scheduled_reports = [
        {
            'name': 'Weekly Performance Summary',
            'schedule': 'Every Monday at 9:00 AM',
            'next_run': 'Dec 18, 2023',
            'status': 'Active'
        },
        {
            'name': 'Guest Satisfaction Report',
            'schedule': 'Monthly • 1st of each month',
            'next_run': 'Jan 1, 2024',
            'status': 'Active'
        },
        {
            'name': 'SLA Breach Alert',
            'schedule': 'Real-time • When SLA is breached',
            'next_run': '',
            'status': 'Paused'
        }
    ]
    
    # Quick templates data
    quick_templates = [
        {
            'name': 'Daily Operations',
            'description': 'Tickets, feedback, SLA status'
        },
        {
            'name': 'Guest Experience',
            'description': 'Satisfaction trends, reviews'
        },
        {
            'name': 'Staff Performance',
            'description': 'Resolution times, workload'
        },
        {
            'name': 'Executive Summary',
            'description': 'High-level KPIs, trends'
        }
    ]
    
    context = {
        'ticket_trends': json.dumps(ticket_trends),
        'ticket_dates': json.dumps(ticket_dates),
        'feedback_trends': json.dumps(feedback_trends),
        'feedback_dates': json.dumps(feedback_dates),
        'satisfaction_scores': json.dumps(satisfaction_scores),
        'satisfaction_weeks': json.dumps(satisfaction_weeks),
        'dept_performance': json.dumps(dept_performance),
        'room_feedback': json.dumps(room_feedback),
        'busiest_hours_data': json.dumps(busiest_hours_data),
        'total_tickets': total_tickets,
        'total_reviews': total_reviews,
        'avg_rating': round(avg_rating, 1),
        'completion_rate': round(completion_rate, 1),
        'top_departments': top_departments,
        'recent_tickets': recent_tickets,
        'recent_reviews': recent_reviews,
        'scheduled_reports': scheduled_reports,
        'quick_templates': quick_templates
    }
    
    return render(request, 'dashboard/analytics_dashboard.html', context)


@login_required
@require_permission([ADMINS_GROUP, STAFF_GROUP])
def create_ticket_api(request):
    """API endpoint to create a new ticket with department routing."""
    if request.method == 'POST':
        try:
            from hotel_app.models import ServiceRequest, RequestType, Location, Department, User
            import json
            
            data = json.loads(request.body.decode('utf-8'))
            
            # Extract data from request
            guest_name = data.get('guest_name')
            room_number = data.get('room_number')
            department_name = data.get('department')
            category = data.get('category')
            priority = data.get('priority')
            description = data.get('description')
            
            # Validate required fields
            if not guest_name or not room_number or not department_name or not category or not priority:
                return JsonResponse({'error': 'Missing required fields'}, status=400)
            
            # Get or create location
            location, _ = Location.objects.get_or_create(
                room_no=room_number,
                defaults={'name': f'Room {room_number}'}
            )
            
            # Get or create request type
            request_type, _ = RequestType.objects.get_or_create(
                name=category,
                defaults={}
            )
            
            # Get department
            try:
                department = Department.objects.get(name=department_name)
            except Department.DoesNotExist:
                return JsonResponse({'error': 'Department not found'}, status=400)
            
            # Map priority to model values
            priority_mapping = {
                'High': 'high',
                'Medium': 'normal',
                'Normal': 'normal',
                'Low': 'low',
            }
            model_priority = priority_mapping.get(priority, 'normal')
            
            # Create service request
            service_request = ServiceRequest.objects.create(
                request_type=request_type,
                location=location,
                requester_user=request.user,
                department=department,
                priority=model_priority,
                status='pending',
                notes=description
            )
            
            # Notify department staff
            service_request.notify_department_staff()
            
            return JsonResponse({
                'success': True,
                'message': 'Ticket created successfully',
                'ticket_id': service_request.id
            })
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)


@login_required
@require_permission([ADMINS_GROUP, STAFF_GROUP])
def assign_ticket_api(request, ticket_id):
    """API endpoint to assign a ticket to a user."""
    if request.method == 'POST':
        try:
            from hotel_app.models import ServiceRequest, User
            import json
            
            data = json.loads(request.body.decode('utf-8'))
            assignee_id = data.get('assignee_id')
            
            if not assignee_id:
                return JsonResponse({'error': 'Assignee ID is required'}, status=400)
            
            # Get the service request and assignee user
            service_request = get_object_or_404(ServiceRequest, id=ticket_id)
            assignee = get_object_or_404(User, id=assignee_id)
            
            # Assign the ticket to the user
            service_request.assign_to_user(assignee)
            
            return JsonResponse({
                'success': True,
                'message': f'Ticket assigned to {assignee.get_full_name() or assignee.username}',
                'ticket_id': service_request.id
            })
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)


@login_required
@require_permission([ADMINS_GROUP, STAFF_GROUP])
def accept_ticket_api(request, ticket_id):
    """API endpoint for a user to accept a ticket."""
    if request.method == 'POST':
        try:
            from hotel_app.models import ServiceRequest
            
            # Get the service request
            service_request = get_object_or_404(ServiceRequest, id=ticket_id)
            
            # Check if the ticket is pending and in the user's department
            # Users can accept pending tickets in their department
            user_department = None
            if hasattr(request.user, 'userprofile') and request.user.userprofile.department:
                user_department = request.user.userprofile.department
            
            if service_request.status != 'pending':
                return JsonResponse({'error': 'Ticket is not in pending status'}, status=400)
            
            if service_request.department != user_department:
                return JsonResponse({'error': 'You are not in the department for this ticket'}, status=403)
            
            # Assign the ticket to the current user if not already assigned
            if not service_request.assignee_user:
                service_request.assignee_user = request.user
                service_request.save()
            
            # Accept the ticket (change status to accepted)
            service_request.accept_task()
            
            return JsonResponse({
                'success': True,
                'message': 'Ticket accepted successfully',
                'ticket_id': service_request.id
            })
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)


@login_required
@require_role(['admin', 'staff', 'user'])
def start_ticket_api(request, ticket_id):
    """API endpoint to start working on a ticket."""
    if request.method == 'POST':
        try:
            from hotel_app.models import ServiceRequest
            
            # Get the service request
            service_request = get_object_or_404(ServiceRequest, id=ticket_id)
            
            # Check if the current user is the assignee
            if service_request.assignee_user != request.user:
                return JsonResponse({'error': 'You are not assigned to this ticket'}, status=403)
            
            # Start working on the ticket (change status to in_progress)
            service_request.start_work()
            
            return JsonResponse({
                'success': True,
                'message': 'Work started on ticket',
                'ticket_id': service_request.id
            })
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)


@login_required
@require_role(['admin', 'staff', 'user'])
def complete_ticket_api(request, ticket_id):
    """API endpoint to mark a ticket as completed."""
    if request.method == 'POST':
        try:
            from hotel_app.models import ServiceRequest
            import json
            
            data = json.loads(request.body.decode('utf-8'))
            resolution_notes = data.get('resolution_notes', '')
            
            # Get the service request
            service_request = get_object_or_404(ServiceRequest, id=ticket_id)
            
            # Check if the current user is the assignee
            if service_request.assignee_user != request.user:
                return JsonResponse({'error': 'You are not assigned to this ticket'}, status=403)
            
            # Complete the ticket (change status to completed)
            service_request.complete_task(resolution_notes)
            
            return JsonResponse({
                'success': True,
                'message': 'Ticket marked as completed',
                'ticket_id': service_request.id
            })
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)


@login_required
@require_role(['admin', 'staff', 'user'])
def close_ticket_api(request, ticket_id):
    """API endpoint to close a ticket."""
    if request.method == 'POST':
        try:
            from hotel_app.models import ServiceRequest
            
            # Get the service request
            service_request = get_object_or_404(ServiceRequest, id=ticket_id)
            
            # Check if user can close (requester, front desk, or superuser)
            is_requester = (service_request.requester_user == request.user)
            is_front_desk = (user_in_group(request.user, 'Front Desk') or 
                           user_in_group(request.user, 'Front Desk Team'))
            is_superuser = request.user.is_superuser
            
            if not (is_requester or is_front_desk or is_superuser):
                return JsonResponse({'error': 'You do not have permission to close this ticket'}, status=403)
            
            # Close the ticket (change status to closed)
            service_request.close_task()
            
            return JsonResponse({
                'success': True,
                'message': 'Ticket closed successfully',
                'ticket_id': service_request.id
            })
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)


@login_required
@require_role(['admin', 'staff', 'user'])
def escalate_ticket_api(request, ticket_id):
    """API endpoint to escalate a ticket."""
    if request.method == 'POST':
        try:
            from hotel_app.models import ServiceRequest
            
            # Get the service request
            service_request = get_object_or_404(ServiceRequest, id=ticket_id)
            
            # Escalate the ticket
            service_request.escalate_task()
            
            return JsonResponse({
                'success': True,
                'message': 'Ticket escalated successfully',
                'ticket_id': service_request.id
            })
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)


@login_required
@require_role(['admin', 'staff', 'user'])
def reject_ticket_api(request, ticket_id):
    """API endpoint to reject a ticket."""
    if request.method == 'POST':
        try:
            from hotel_app.models import ServiceRequest
            
            # Get the service request
            service_request = get_object_or_404(ServiceRequest, id=ticket_id)
            
            # Reject the ticket
            service_request.reject_task()
            
            return JsonResponse({
                'success': True,
                'message': 'Ticket rejected successfully',
                'ticket_id': service_request.id
            })
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)


# ---- New Voucher System: Guest Management ----
@require_permission([ADMINS_GROUP, STAFF_GROUP])
def dashboard_guests(request):
    """Guest management dashboard with filters."""
    search = request.GET.get('search', '')
    breakfast_filter = request.GET.get('breakfast_filter', '')
    status_filter = request.GET.get('status_filter', '')
    qr_filter = request.GET.get('qr_filter', '')
    
    guests = Guest.objects.all().order_by('-created_at')
    
    if search:
        guests = guests.filter(
            Q(full_name__icontains=search) | Q(email__icontains=search) |
            Q(room_number__icontains=search) | Q(guest_id__icontains=search) |
            Q(phone__icontains=search)
        )
    if breakfast_filter == 'yes':
        guests = guests.filter(breakfast_included=True)
    elif breakfast_filter == 'no':
        guests = guests.filter(breakfast_included=False)
    if qr_filter == 'with_qr':
        guests = guests.exclude(details_qr_code='')
    elif qr_filter == 'without_qr':
        guests = guests.filter(details_qr_code='')
    if status_filter:
        today = timezone.now().date()
        if status_filter == 'current':
            guests = guests.filter(checkin_date__lte=today, checkout_date__gte=today)
        elif status_filter == 'past':
            guests = guests.filter(checkout_date__lt=today)
        elif status_filter == 'future':
            guests = guests.filter(checkin_date__gt=today)
    
    context = {
        "guests": guests,
        "search": search,
        "breakfast_filter": breakfast_filter,
        "status_filter": status_filter,
        "qr_filter": qr_filter,
        "title": "Guest Management"
    }
    return render(request, "dashboard/guests.html", context)

@require_permission([ADMINS_GROUP, STAFF_GROUP])
def guest_detail(request, guest_id):
    """Guest detail view with vouchers and stay information."""
    guest = get_object_or_404(Guest, pk=guest_id)
    vouchers = guest.vouchers.all().order_by('-created_at')
    
    stay_duration = "N/A"
    if guest.checkin_date and guest.checkout_date:
        duration = guest.checkout_date - guest.checkin_date
        stay_duration = f"{duration.days} days"
    
    context = {
        "guest": guest,
        "vouchers": vouchers,
        "stay_duration": stay_duration,
        "title": f"Guest: {guest.full_name}"
    }
    return render(request, "dashboard/guest_detail.html", context)


# ---- New Voucher System: Voucher Management ----
@require_permission([ADMINS_GROUP, STAFF_GROUP])
def dashboard_vouchers(request):
    """Voucher management dashboard."""
    vouchers = Voucher.objects.all().select_related('guest').order_by('-created_at')
    
    for voucher in vouchers:
        if not voucher.qr_image:
            voucher.generate_qr_code(size='xxlarge')
    
    context = {
        "vouchers": vouchers,
        "total_vouchers": vouchers.count(),
        "active_vouchers": vouchers.filter(status='active').count(),
        "redeemed_vouchers": vouchers.filter(status='redeemed').count(),
        "expired_vouchers": vouchers.filter(status='expired').count(),
        "title": "Voucher Management"
    }
    return render(request, "dashboard/vouchers.html", context)

@require_permission([ADMINS_GROUP])
def voucher_create(request):
    if request.method == "POST":
        form = VoucherForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Voucher created successfully.")
    return redirect("dashboard:vouchers")

@require_permission([ADMINS_GROUP])
def voucher_update(request, voucher_id):
    voucher = get_object_or_404(Voucher, pk=voucher_id)
    if request.method == "POST":
        form = VoucherForm(request.POST, instance=voucher)
        if form.is_valid():
            form.save()
            messages.success(request, "Voucher updated successfully.")
    return redirect("dashboard:vouchers")

@require_permission([ADMINS_GROUP])
def voucher_delete(request, voucher_id):
    voucher = get_object_or_404(Voucher, pk=voucher_id)
    if request.method == "POST":
        voucher.delete()
        messages.success(request, "Voucher deleted successfully.")
    return redirect("dashboard:vouchers")

@require_permission([ADMINS_GROUP, STAFF_GROUP])
def voucher_detail(request, voucher_id):
    """Voucher detail view with scan history."""
    voucher = get_object_or_404(Voucher, pk=voucher_id)
    scans = voucher.scans.all().order_by('-scanned_at')
    
    if not voucher.qr_image:
        if voucher.generate_qr_code(size='xxlarge'):
            messages.success(request, 'QR code generated successfully!')
        else:
            messages.error(request, 'Failed to generate QR code.')
    
    context = {
        "voucher": voucher,
        "scans": scans,
        "title": f"Voucher: {voucher.voucher_code}"
    }
    return render(request, "dashboard/voucher_detail.html", context)

@require_permission([ADMINS_GROUP, STAFF_GROUP])
def regenerate_voucher_qr(request, voucher_id):
    """Regenerate QR code for a specific voucher."""
    voucher = get_object_or_404(Voucher, pk=voucher_id)
    if request.method == 'POST':
        qr_size = request.POST.get('qr_size', 'xxlarge')
        if voucher.generate_qr_code(size=qr_size):
            messages.success(request, f'QR code regenerated with size: {qr_size}!')
        else:
            messages.error(request, 'Failed to regenerate QR code.')
    return redirect('dashboard:voucher_detail', voucher_id=voucher.id)

@require_permission([ADMINS_GROUP, STAFF_GROUP])
def share_voucher_whatsapp(request, voucher_id):
    """Share voucher via WhatsApp API."""
    voucher = get_object_or_404(Voucher, pk=voucher_id)
    if request.method == 'POST':
        if not voucher.guest or not voucher.guest.phone:
            return JsonResponse({'success': False, 'error': 'Guest phone number is not available.'})
        try:
            whatsapp_service = WhatsAppService()
            result = whatsapp_service.send_voucher_message(voucher)
            if result.get('success'):
                return JsonResponse({'success': True, 'message': 'Voucher shared via WhatsApp!'})
            else:
                return JsonResponse({'success': False, 'error': 'WhatsApp API unavailable.', 'fallback': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Service error: {str(e)}', 'fallback': True})
    return JsonResponse({'success': False, 'error': 'Invalid request method.'})


# ---- Guest QR Codes Dashboard ----
@require_permission([ADMINS_GROUP, STAFF_GROUP])
def guest_qr_codes(request):
    """Display all guest QR codes in a grid layout with filters."""


@require_permission([ADMINS_GROUP])
def sla_configuration(request):
    """SLA Configuration page with pagination and filtering."""
    # Get page and page size from query parameters
    page = int(request.GET.get('page', 1))
    page_size = int(request.GET.get('page_size', 20))
    
    # Get filter parameters
    search_query = request.GET.get('search', '')
    department_filter = request.GET.get('department', '')
    
    # Get general SLA configurations (these are always shown)
    sla_configs = SLAConfiguration.objects.all().order_by('priority')
    
    # Get all departments for the new configuration section
    departments = Department.objects.all().order_by('name')
    
    # Get distinct department/request combinations
    # We need to get one entry per department/request_type combination
    department_sla_configs = DepartmentRequestSLA.objects.select_related(
        'department', 'request_type'
    ).order_by('department__name', 'request_type__name')
    
    # Since we can't use distinct() with select_related in PostgreSQL, we'll handle this in Python
    # Get all entries and then filter to unique combinations
    all_configs = list(department_sla_configs)
    unique_configs = []
    seen_combinations = set()
    
    for config in all_configs:
        combination = (config.department_id, config.request_type_id)
        if combination not in seen_combinations:
            unique_configs.append(config)
            seen_combinations.add(combination)
    
    department_sla_configs = unique_configs
    
    # Apply filters
    if search_query:
        department_sla_configs = [config for config in department_sla_configs 
                                 if search_query.lower() in config.department.name.lower() or 
                                    search_query.lower() in config.request_type.name.lower()]
    
    if department_filter:
        department_sla_configs = [config for config in department_sla_configs 
                                 if config.department.name == department_filter]
    
    # Convert to a format that can be paginated
    # Since we're working with a list, we need to manually handle pagination
    from django.core.paginator import Paginator as ListPaginator
    
    paginator = ListPaginator(department_sla_configs, page_size)
    try:
        department_sla_page = paginator.page(page)
    except PageNotAnInteger:
        department_sla_page = paginator.page(1)
    except EmptyPage:
        department_sla_page = paginator.page(paginator.num_pages)
    
    context = {
        'active_tab': 'sla_configuration',
        'title': 'SLA Configuration',
        'subtitle': 'Configure default SLA times for different priority levels',
        'sla_configs': sla_configs,
        'departments': departments,
        'department_sla_configs': department_sla_page,  # Paginated results
        'paginator': paginator,
        'page_obj': department_sla_page,
        # Filter values for the template
        'search_query': search_query,
        'department_filter': department_filter,
    }
    return render(request, 'dashboard/sla_configuration.html', context)


@require_permission([ADMINS_GROUP])
def api_sla_configuration_update(request):
    """API endpoint to update SLA configurations."""
    if request.method == 'POST':
        try:
            import json
            data = json.loads(request.body.decode('utf-8'))
            
            # Handle clear all request
            if data.get('clear_all'):
                # Delete all department SLA configurations
                DepartmentRequestSLA.objects.all().delete()
                return JsonResponse({
                    'success': True,
                    'message': 'All department SLA configurations cleared successfully'
                })
            
            # Handle import data request
            import_data = data.get('import_data', [])
            if import_data:
                imported_count = 0
                for item in import_data:
                    department_name = item.get('department')
                    request_type_name = item.get('request_type')
                    response_time = item.get('response_time', 30)
                    resolution_time = item.get('resolution_time', 120)
                    
                    if department_name and request_type_name:
                        # Get or create the department
                        department, dept_created = Department.objects.get_or_create(
                            name=department_name,
                            defaults={'description': f'Department for {department_name}'}
                        )
                        
                        # Get or create the request type
                        request_type, req_created = RequestType.objects.get_or_create(
                            name=request_type_name,
                            defaults={'description': f'Request type for {request_type_name}'}
                        )
                        
                        # For each priority level, create or update the SLA configuration
                        for priority in ['critical', 'high', 'normal', 'low']:
                            DepartmentRequestSLA.objects.update_or_create(
                                department=department,
                                request_type=request_type,
                                priority=priority,
                                defaults={
                                    'response_time_minutes': response_time,
                                    'resolution_time_minutes': resolution_time
                                }
                            )
                        imported_count += 1
                
                return JsonResponse({
                    'success': True,
                    'message': f'Successfully imported {imported_count} SLA configurations'
                })
            
            # Handle add department config request
            add_config = data.get('add_department_config')
            if add_config:
                department_id = add_config.get('department_id')
                request_type_name = add_config.get('request_type')
                response_time = add_config.get('response_time_minutes')
                resolution_time = add_config.get('resolution_time_minutes')
                
                if department_id and request_type_name and response_time and resolution_time:
                    # Get the department
                    try:
                        department = Department.objects.get(id=department_id)
                    except Department.DoesNotExist:
                        return JsonResponse({'error': 'Department not found'}, status=400)
                    
                    # Get or create the request type
                    request_type, created = RequestType.objects.get_or_create(
                        name=request_type_name,
                        defaults={'description': f'Request type for {request_type_name}'}
                    )
                    
                    # For each priority level, create or update the SLA configuration
                    for priority in ['critical', 'high', 'normal', 'low']:
                        DepartmentRequestSLA.objects.update_or_create(
                            department=department,
                            request_type=request_type,
                            priority=priority,
                            defaults={
                                'response_time_minutes': response_time,
                                'resolution_time_minutes': resolution_time
                            }
                        )
                    
                    return JsonResponse({
                        'success': True,
                        'message': 'Department SLA configuration added successfully'
                    })
            
            # Handle delete department config request
            delete_config = data.get('delete_department_config')
            if delete_config:
                department_id = delete_config.get('department_id')
                request_type_name = delete_config.get('request_type')
                
                if department_id and request_type_name:
                    # Get the department
                    try:
                        department = Department.objects.get(id=department_id)
                    except Department.DoesNotExist:
                        return JsonResponse({'error': 'Department not found'}, status=400)
                    
                    # Get the request type
                    try:
                        request_type = RequestType.objects.get(name=request_type_name)
                    except RequestType.DoesNotExist:
                        return JsonResponse({'error': 'Request type not found'}, status=400)
                    
                    # Delete all SLA configurations for this department/request type combination
                    DepartmentRequestSLA.objects.filter(
                        department=department,
                        request_type=request_type
                    ).delete()
                    
                    return JsonResponse({
                        'success': True,
                        'message': 'Department SLA configuration removed successfully'
                    })
            
            # Update general SLA configurations
            for config_data in data.get('general_configs', []):
                priority = config_data.get('priority')
                response_time = config_data.get('response_time_minutes')
                resolution_time = config_data.get('resolution_time_minutes')
                
                if priority and response_time is not None and resolution_time is not None:
                    SLAConfiguration.objects.update_or_create(
                        priority=priority,
                        defaults={
                            'response_time_minutes': response_time,
                            'resolution_time_minutes': resolution_time
                        }
                    )
            
            # Update department/request-specific SLA configurations
            for config_data in data.get('department_configs', []):
                department_id = config_data.get('department_id')
                request_type_name = config_data.get('request_type')
                response_time = config_data.get('response_time_minutes')
                resolution_time = config_data.get('resolution_time_minutes')
                
                if (department_id and request_type_name and 
                    response_time is not None and resolution_time is not None):
                    # Get or create the request type
                    request_type, created = RequestType.objects.get_or_create(
                        name=request_type_name,
                        defaults={'description': f'Request type for {request_type_name}'}
                    )
                    
                    # Get the department
                    try:
                        department = Department.objects.get(id=department_id)
                    except Department.DoesNotExist:
                        continue  # Skip if department doesn't exist
                        
                    # For each priority level, create or update the SLA configuration
                    for priority in ['critical', 'high', 'normal', 'low']:
                        DepartmentRequestSLA.objects.update_or_create(
                            department=department,
                            request_type=request_type,
                            priority=priority,
                            defaults={
                                'response_time_minutes': response_time,
                                'resolution_time_minutes': resolution_time
                            }
                        )
            
            return JsonResponse({
                'success': True,
                'message': 'SLA configurations updated successfully'
            })
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    elif request.method == 'GET':
        # Return current SLA configurations
        try:
            # General SLA configurations
            general_configs = SLAConfiguration.objects.all().order_by('priority')
            general_config_data = []
            for config in general_configs:
                general_config_data.append({
                    'priority': config.priority,
                    'response_time_minutes': config.response_time_minutes,
                    'resolution_time_minutes': config.resolution_time_minutes
                })
            
            # Department/request-specific SLA configurations (distinct combinations)
            all_department_configs = DepartmentRequestSLA.objects.select_related(
                'department', 'request_type'
            ).order_by('department_id', 'request_type_id')
            
            # Get unique combinations
            unique_department_configs = []
            seen_combinations = set()
            
            for config in all_department_configs:
                combination = (config.department_id, config.request_type_id)
                if combination not in seen_combinations:
                    unique_department_configs.append({
                        'department_id': config.department_id,
                        'request_type_id': config.request_type_id,
                        'request_type_name': config.request_type.name,
                        'response_time_minutes': config.response_time_minutes,
                        'resolution_time_minutes': config.resolution_time_minutes
                    })
                    seen_combinations.add(combination)
            
            return JsonResponse({
                'success': True,
                'general_configs': general_config_data,
                'department_configs': unique_department_configs
            })
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)

@require_permission([ADMINS_GROUP, STAFF_GROUP])
def regenerate_guest_qr(request, guest_id):
    """Regenerate QR code for a specific guest."""
    guest = get_object_or_404(Guest, pk=guest_id)
    if request.method == 'POST':
        qr_size = request.POST.get('qr_size', 'xlarge')
        if guest.generate_details_qr_code(size=qr_size):
            messages.success(request, f'Guest QR code regenerated with size: {qr_size}!')

        else:
            messages.error(request, 'Failed to regenerate guest QR code.')
    return redirect('dashboard:guest_detail', guest_id=guest.id)

@require_permission([ADMINS_GROUP, STAFF_GROUP])
def share_guest_qr_whatsapp(request, guest_id):
    """Share guest QR code via WhatsApp API."""
    guest = get_object_or_404(Guest, pk=guest_id)
    if request.method == 'POST':
        if not guest.phone:
            return JsonResponse({'success': False, 'error': 'Guest phone number not available.'})
        try:
            whatsapp_service = WhatsAppService()
            result = whatsapp_service.send_guest_qr_message(guest)
            if result.get('success'):
                return JsonResponse({'success': True, 'message': 'Guest QR code shared via WhatsApp!'})
            else:
                return JsonResponse({'success': False, 'error': 'WhatsApp API unavailable.', 'fallback': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Service error: {str(e)}', 'fallback': True})
    return JsonResponse({'success': False, 'error': 'Invalid request method.'})

@require_permission([ADMINS_GROUP, STAFF_GROUP])
def get_guest_whatsapp_message(request, guest_id):
    """Get a pre-formatted WhatsApp message template for a guest."""
    guest = get_object_or_404(Guest, pk=guest_id)
    message = (
        f"Hello {guest.full_name},\n\n"
        f"Welcome! Here is your personal QR code for accessing hotel services.\n\n"
        f"Guest ID: {guest.guest_id}\n"
        f"Room: {guest.room_number}"
    )
    return JsonResponse({
        'success': True,
        'message': message,
        'guest_name': guest.full_name,
        'guest_phone': guest.phone
    })


@login_required
@require_permission([ADMINS_GROUP, STAFF_GROUP])
def get_ticket_suggestions_api(request):
    """API endpoint to get ticket suggestions based on department and SLA configurations."""
    if request.method == 'GET':
        try:
            from hotel_app.models import DepartmentRequestSLA, RequestType, Department
            
            department_name = request.GET.get('department_name')
            search_term = request.GET.get('search_term', '').lower()
            
            # Get all department SLA configurations
            if department_name:
                try:
                    department = Department.objects.get(name=department_name)
                    department_configs = DepartmentRequestSLA.objects.select_related(
                        'department', 'request_type'
                    ).filter(department=department)
                except Department.DoesNotExist:
                    department_configs = DepartmentRequestSLA.objects.select_related(
                        'department', 'request_type'
                    ).none()
            else:
                department_configs = DepartmentRequestSLA.objects.select_related(
                    'department', 'request_type'
                ).all()
            
            # Extract unique request types and their descriptions
            suggestions = []
            seen_request_types = set()
            
            for config in department_configs:
                request_type = config.request_type
                if request_type.id not in seen_request_types:
                    # Create suggestion text based on request type and department
                    suggestion_text = f"{request_type.name} - {config.department.name}"
                    if request_type.description:
                        suggestion_text += f": {request_type.description[:100]}"
                    
                    # Only include suggestions that match the search term
                    if search_term in request_type.name.lower() or search_term in suggestion_text.lower():
                        suggestions.append({
                            'id': request_type.id,
                            'name': request_type.name,
                            'description': request_type.description or '',
                            'department': config.department.name,
                            'suggestion_text': suggestion_text
                        })
                    
                    seen_request_types.add(request_type.id)
            
            # Limit to 10 suggestions
            suggestions = suggestions[:10]
            
            return JsonResponse({
                'success': True,
                'suggestions': suggestions
            })
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)



# ---- Feedback View ----
@login_required
@user_passes_test(is_staff)
def feedback_inbox(request):
    """Feedback inbox view showing all guest feedback."""
    from .models import Review, Guest
    from .forms import FeedbackForm
    

    # Handle form submission for new feedback
    if request.method == 'POST':
        # Get form data
        guest_name = request.POST.get('guest_name', '')
        room_number = request.POST.get('room_number', '')
        overall_rating = request.POST.get('overall_rating', 0)
        cleanliness_rating = request.POST.get('cleanliness_rating', 0)
        staff_rating = request.POST.get('staff_rating', 0)
        recommend = request.POST.get('recommend', '')
        comment = request.POST.get('comment', '')
        
        # Validate required fields
        if not guest_name or not room_number or not overall_rating:
            messages.error(request, 'Please fill in all required fields.')
        else:
            try:
                # Create or get guest
                guest, created = Guest.objects.get_or_create(
                    room_number=room_number,
                    defaults={'full_name': guest_name}
                )
                
                # If guest exists but name is different, update it
                if not created and guest.full_name != guest_name:
                    guest.full_name = guest_name
                    guest.save()
                
                # Format comment with all ratings
                full_comment = comment
                if full_comment:
                    full_comment += "\n\n"
                else:
                    full_comment = ""
                
                full_comment += f"Overall Rating: {overall_rating}/5\n"
                full_comment += f"Cleanliness Rating: {cleanliness_rating}/5\n"
                full_comment += f"Staff Service Rating: {staff_rating}/5\n"
                full_comment += f"Recommendation: {recommend}"
                
                # Create review
                Review.objects.create(
                    guest=guest,
                    rating=overall_rating,
                    comment=full_comment
                )
                
                messages.success(request, 'Feedback added successfully!')
                return redirect('dashboard:feedback_inbox')
            except Exception as e:
                messages.error(request, f'Error saving feedback: {str(e)}')
    else:
        form = FeedbackForm()
    
    # Get all reviews with related guest information
    reviews = Review.objects.select_related('guest').all().order_by('-created_at')
    
    # Convert to the format expected by the template
    feedback_data = []
    for review in reviews:
        # Determine sentiment based on rating
        if review.rating >= 4:
            sentiment = 'Positive'
        elif review.rating <= 2:
            sentiment = 'Negative'
        else:
            sentiment = 'Neutral'
            
        # Extract keywords from comment (simple approach)
        keywords = []
        if review.comment:
            # Simple keyword extraction - in a real implementation, you might use NLP
            common_words = ['service', 'staff', 'room', 'food', 'clean', 'location', 'wifi', 'pool', 'spa', 'breakfast']
            comment_lower = review.comment.lower()
            keywords = [word for word in common_words if word in comment_lower]
            # Limit to 3 keywords
            keywords = keywords[:3]
        
        feedback_data.append({
            'id': review.id,
            'date': review.created_at.strftime('%b %d, %Y'),
            'guest': review.guest.full_name if review.guest else 'Anonymous',
            'room': getattr(review.guest, 'room_number', 'N/A') if review.guest else 'N/A',
            'rating': float(review.rating),
            'feedback': review.comment[:100] + '...' if review.comment and len(review.comment) > 100 else review.comment or '',
            'keywords': keywords,
            'sentiment': sentiment,
            'status': 'responded' if review.updated_at else 'needs_attention'
        })
    
    # Calculate statistics
    total_feedback = reviews.count()
    avg_rating = reviews.aggregate(Avg('rating'))['rating__avg'] or 0
    needs_attention = reviews.filter(updated_at__isnull=True).count()
    response_rate = int((needs_attention / total_feedback * 100)) if total_feedback > 0 else 0
    
    context = {
        'feedback_data': feedback_data,
        'stats': {
            'total_feedback': total_feedback,
            'avg_rating': round(avg_rating, 1),
            'needs_attention': needs_attention,
            'response_rate': 100 - response_rate
        },
        'form': form
    }
    
    return render(request, 'dashboard/feedback_inbox.html', context)


@login_required
@user_passes_test(is_staff)
def feedback_detail(request, feedback_id):
    """Feedback detail view showing detailed information about a specific feedback."""
    from .models import Review, Guest
    
    # Get the review
    try:
        review = Review.objects.select_related('guest').get(id=feedback_id)
    except Review.DoesNotExist:
        # Handle the case where the review doesn't exist
        from django.http import Http404
        raise Http404("Review not found")
    
    # Determine sentiment based on rating
    if review.rating >= 4:
        sentiment = 'Positive'
    elif review.rating <= 2:
        sentiment = 'Negative'
    else:
        sentiment = 'Neutral'
    
    # Extract keywords from comment (simple approach)
    keywords = []
    if review.comment:
        # Simple keyword extraction - in a real implementation, you might use NLP
        common_words = ['service', 'staff', 'room', 'food', 'clean', 'location', 'wifi', 'pool', 'spa', 'breakfast', 'concierge', 'reception']
        comment_lower = review.comment.lower()
        keywords = [word for word in common_words if word in comment_lower]
    
    # Create feedback data structure
    feedback = {
        'id': review.id,
        'date': review.created_at.strftime('%B %d, %Y'),
        'time': review.created_at.strftime('%I:%M %p'),
        'guest': review.guest.full_name if review.guest else 'Anonymous',
        'room': getattr(review.guest, 'room_number', 'N/A') if review.guest else 'N/A',
        'room_type': 'Standard Room',  # This would come from guest data in a real implementation
        'rating': float(review.rating),
        'sentiment': sentiment,
        'title': f'{sentiment} Review - {review.rating} Stars',
        'comment': review.comment or '',
        'keywords': keywords,
        'department_impact': [
            {'department': 'Room Service', 'sentiment': 'Negative' if 'service' in keywords else 'Positive'},
            {'department': 'Housekeeping', 'sentiment': 'Negative' if 'clean' in keywords else 'Positive'},
            {'department': 'Front Desk', 'sentiment': 'Negative' if 'reception' in keywords else 'Positive'}
        ],
        'activity_timeline': [
            {'event': 'Feedback received', 'time': review.created_at.strftime('%B %d, %I:%M %p'), 'description': 'Guest submitted feedback', 'status': 'completed'},
            {'event': 'Auto-tagged by AI', 'time': review.created_at.strftime('%B %d, %I:%M %p'), 'description': 'System identified keywords and sentiment', 'status': 'completed'},
            {'event': 'Pending action', 'time': 'Now', 'description': 'Awaiting manager response', 'status': 'pending'}
        ],
        'attachments': [],  # This would be populated from actual attachments in a real implementation
        'guest_info': {
            'name': review.guest.full_name if review.guest else 'Anonymous',
            'loyalty_member': True,  # This would come from guest data in a real implementation
            'check_in': review.guest.checkin_date.strftime('%B %d, %Y') if review.guest and review.guest.checkin_date else 'N/A',
            'check_out': review.guest.checkout_date.strftime('%B %d, %Y') if review.guest and review.guest.checkout_date else 'N/A',
            'stay_duration': '3 nights',  # This would be calculated in a real implementation
            'previous_stays': 1  # This would come from guest data in a real implementation
        },
        'response_status': {
            'status': 'Responded' if review.updated_at else 'Pending Review',
            'priority': 'High' if review.rating <= 2 else 'Normal',
            'due_date': (review.created_at + timezone.timedelta(days=1)).strftime('%B %d, %Y')
        }
    }
    
    context = {
        'feedback': feedback
    }
    
    return render(request, 'dashboard/feedback_detail.html', context)


# ---- Ticket Workflow API Endpoints ----
@login_required
@require_permission([ADMINS_GROUP, STAFF_GROUP])
def create_ticket_api(request):
    """API endpoint to create a new ticket with department routing."""
    if request.method == 'POST':
        try:
            from hotel_app.models import ServiceRequest, RequestType, Location, Department, User, SLAConfiguration
            import json
            
            data = json.loads(request.body.decode('utf-8'))
            
            # Extract data from request
            guest_name = data.get('guest_name')
            room_number = data.get('room_number')
            department_name = data.get('department')
            category = data.get('category')
            priority = data.get('priority')
            description = data.get('description')
            
            # Validate required fields
            if not guest_name or not room_number or not department_name or not category or not priority:
                return JsonResponse({'error': 'Missing required fields'}, status=400)
            
            # Get or create location
            location, _ = Location.objects.get_or_create(
                room_no=room_number,
                defaults={'name': f'Room {room_number}'}
            )
            
            # Get or create request type
            request_type, _ = RequestType.objects.get_or_create(
                name=category,
                defaults={}
            )
            
            # Get department
            try:
                department = Department.objects.get(name=department_name)
            except Department.DoesNotExist:
                return JsonResponse({'error': 'Invalid department'}, status=400)
            
            # Map priority to model values
            priority_mapping = {
                'Critical': 'critical',
                'High': 'high',
                'Medium': 'normal',
                'Normal': 'normal',
                'Low': 'low',
            }
            model_priority = priority_mapping.get(priority, 'normal')
            
            # Create service request (SLA times will be set automatically in the model's save method)
            service_request = ServiceRequest.objects.create(
                request_type=request_type,
                location=location,
                requester_user=request.user,
                department=department,
                priority=model_priority,
                status='pending',
                notes=description
            )
            
            # Notify department staff
            service_request.notify_department_staff()
            
            return JsonResponse({
                'success': True,
                'message': 'Ticket created successfully',
                'ticket_id': service_request.id
            })
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)


@login_required
@require_permission([ADMINS_GROUP, STAFF_GROUP])
def assign_ticket_api(request, ticket_id):
    """API endpoint to assign a ticket to a user or department."""
    if request.method == 'POST':
        try:
            from hotel_app.models import ServiceRequest, User, Department
            import json
            
            data = json.loads(request.body.decode('utf-8'))
            assignee_id = data.get('assignee_id')
            department_id = data.get('department_id')
            
            # Get the service request
            service_request = get_object_or_404(ServiceRequest, id=ticket_id)
            
            # Assign to user if provided
            if assignee_id:
                assignee = get_object_or_404(User, id=assignee_id)
                service_request.assign_to_user(assignee)
                return JsonResponse({
                    'success': True,
                    'message': f'Ticket assigned to {assignee.get_full_name() or assignee.username}',
                    'ticket_id': service_request.id
                })
            
            # Assign to department if provided
            elif department_id:
                department = get_object_or_404(Department, id=department_id)
                service_request.assign_to_department(department)
                return JsonResponse({
                    'success': True,
                    'message': f'Ticket assigned to {department.name} department',
                    'ticket_id': service_request.id
                })
            
            return JsonResponse({'error': 'Assignee or department ID is required'}, status=400)
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)


# Removed claim_ticket_api as we're removing the claim functionality
# Tickets are now directly assigned and accepted


@login_required
@require_role(['admin', 'staff', 'user'])
def accept_ticket_api(request, ticket_id):
    """API endpoint for a user to accept a ticket."""
    if request.method == 'POST':
        try:
            from hotel_app.models import ServiceRequest
            
            # Get the service request
            service_request = get_object_or_404(ServiceRequest, id=ticket_id)
            
            # Check if the ticket is pending and in the user's department
            # Users can accept pending tickets in their department
            user_department = None
            if hasattr(request.user, 'userprofile') and request.user.userprofile.department:
                user_department = request.user.userprofile.department
            
            if service_request.status != 'pending':
                return JsonResponse({'error': 'Ticket is not in pending status'}, status=400)
            
            # Check if user can accept the ticket (either in same department or is the requester)
            if not (service_request.department == user_department or service_request.requester_user == request.user):
                return JsonResponse({'error': 'You do not have permission to accept this ticket'}, status=403)
            
            # Assign the ticket to the current user if not already assigned
            if not service_request.assignee_user:
                service_request.assignee_user = request.user
                service_request.save()
            
            # Accept the ticket (change status to accepted)
            service_request.accept_task()
            
            return JsonResponse({
                'success': True,
                'message': 'Ticket accepted successfully',
                'ticket_id': service_request.id
            })
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)


# ---- Integrations ----
@login_required
def integrations(request):
    """View for the integrations page."""
    return render(request, "dashboard/integrations.html")


@login_required
@require_role(['admin', 'staff'])
def performance_dashboard(request):
    """Render the Performance Dashboard page with dynamic data."""
    from django.db.models import Count, Avg, Q
    from .models import ServiceRequest, Department, User, UserProfile
    import datetime
    from django.utils import timezone
    
    # Calculate date ranges
    today = timezone.now().date()
    week_ago = today - datetime.timedelta(days=7)
    
    # Overall Completion Rate
    total_requests = ServiceRequest.objects.count()
    completed_requests = ServiceRequest.objects.filter(status='completed').count()
    completion_rate = round((completed_requests / total_requests * 100), 1) if total_requests > 0 else 0
    
    # SLA Breaches
    sla_breaches = ServiceRequest.objects.filter(
        Q(response_sla_breached=True) | Q(resolution_sla_breached=True)
    ).count()
    
    # Average Response Time (in minutes)
    avg_response_time = 0
    responded_requests = ServiceRequest.objects.filter(
        started_at__isnull=False
    ).exclude(accepted_at__isnull=True)
    
    if responded_requests.exists():
        total_response_time = datetime.timedelta()
        for req in responded_requests:
            total_response_time += (req.started_at - req.accepted_at)
        avg_response_time = round(total_response_time.total_seconds() / 60 / responded_requests.count())
    
    # Active Staff
    active_staff = User.objects.filter(is_active=True).count()
    
    # Completion Rates by Department
    departments = Department.objects.all()
    department_completion_data = []
    department_labels = []
    
    for dept in departments:
        dept_requests = ServiceRequest.objects.filter(department=dept)
        total_dept_requests = dept_requests.count()
        completed_dept_requests = dept_requests.filter(status='completed').count()
        dept_completion_rate = round((completed_dept_requests / total_dept_requests * 100), 1) if total_dept_requests > 0 else 0
        
        department_labels.append(dept.name)
        department_completion_data.append(dept_completion_rate)
    
    # SLA Breach Trends (last 7 days)
    sla_breach_trends = []
    sla_breach_labels = []
    
    for i in range(7):
        date = week_ago + datetime.timedelta(days=i)
        breaches = ServiceRequest.objects.filter(
            Q(response_sla_breached=True) | Q(resolution_sla_breached=True),
            created_at__date=date
        ).count()
        
        sla_breach_labels.append(date.strftime('%a'))
        sla_breach_trends.append(breaches)
    
    # Top Performers (users with highest completion rates)
    top_performers = []
    users_with_requests = User.objects.filter(
        requests_assigned__isnull=False
    ).annotate(
        total_requests=Count('requests_assigned'),
        completed_requests=Count('requests_assigned', filter=Q(requests_assigned__status='completed'))
    ).filter(total_requests__gt=0)
    
    for user in users_with_requests:
        completion_rate_user = round((user.completed_requests / user.total_requests * 100), 1)
        top_performers.append({
            'user': user,
            'completion_rate': completion_rate_user,
            'tickets_completed': user.completed_requests,
            'department': getattr(user.userprofile, 'department', None)
        })
    
    # Sort by completion rate and take top 5
    top_performers = sorted(top_performers, key=lambda x: x['completion_rate'], reverse=True)[:5]
    
    # Department Rankings
    department_rankings = []
    for dept in departments:
        dept_requests = ServiceRequest.objects.filter(department=dept)
        total_dept_requests = dept_requests.count()
        completed_dept_requests = dept_requests.filter(status='completed').count()
        dept_completion_rate = round((completed_dept_requests / total_dept_requests * 100), 1) if total_dept_requests > 0 else 0
        
        # Count staff in department
        staff_count = UserProfile.objects.filter(department=dept).count()
        
        department_rankings.append({
            'department': dept,
            'completion_rate': dept_completion_rate,
            'tickets_handled': total_dept_requests,
            'staff_count': staff_count
        })
    
    # Sort by completion rate
    department_rankings = sorted(department_rankings, key=lambda x: x['completion_rate'], reverse=True)
    
    # Staff Performance Details
    staff_performance = []
    for user in users_with_requests:
        completion_rate_user = round((user.completed_requests / user.total_requests * 100), 1) if user.total_requests > 0 else 0
        
        # Calculate breaches for this user
        user_breaches = ServiceRequest.objects.filter(
            Q(response_sla_breached=True) | Q(resolution_sla_breached=True),
            assignee_user=user
        ).count()
        
        # Determine status based on performance
        if completion_rate_user >= 95:
            status = 'Excellent'
            status_class = 'bg-green-100 text-green-700'
        elif completion_rate_user >= 85:
            status = 'Good'
            status_class = 'bg-sky-100 text-sky-700'
        else:
            status = 'Needs Improvement'
            status_class = 'bg-yellow-100 text-yellow-800'
        
        staff_performance.append({
            'user': user,
            'department': getattr(user.userprofile, 'department', None),
            'tickets_completed': user.completed_requests,
            'completion_rate': completion_rate_user,
            'avg_response': avg_response_time,  # Simplified for now
            'breaches': user_breaches,
            'status': status,
            'status_class': status_class
        })
    
    context = {
        # Stats cards
        'completion_rate': completion_rate,
        'sla_breaches': sla_breaches,
        'avg_response_time': avg_response_time,
        'active_staff': active_staff,
        
        # Charts
        'department_labels': department_labels,
        'department_completion_data': department_completion_data,
        'sla_breach_labels': sla_breach_labels,
        'sla_breach_trends': sla_breach_trends,
        
        # Tables
        'top_performers': top_performers,
        'department_rankings': department_rankings,
        'staff_performance': staff_performance,
    }
    
    return render(request, 'dashboard/performance_dashboard.html', context)


# ---- Tailwind Test ----
@login_required
def tailwind_test(request):
    """View for testing Tailwind CSS functionality."""
    return render(request, "dashboard/tailwind_test.html")


@login_required
def gym(request):
    """
    Render the Gym Management page, handle member creation, and paginate results.
    """
    # Initialize the form for the modal. It will be empty on a GET request.
    form = GymMemberForm()

    # Handle the form submission (when you click "Create Member")
    if request.method == 'POST':
        form = GymMemberForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'New gym member has been added successfully!')
            # Redirect to the same page to prevent re-submission on refresh
            return redirect('dashboard:gym') 
        else:
            # If the form has errors, the page will re-render below,
            # and the 'form' variable with errors will be passed to the template.
            messages.error(request, 'Please correct the errors in the form.')

    # Get all members from the database for the GET request
    member_list = GymMember.objects.all().order_by('-id')
    total_members = member_list.count()

    # Set up Django's built-in Paginator
    paginator = Paginator(member_list, 10) # Show 10 members per page
    page_number = request.GET.get('page')

    try:
        page_obj = paginator.get_page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    context = {
        'page_obj': page_obj,          # The template expects an object named 'page_obj'
        'total_members': total_members,
        'form': form,                  # Pass the form (empty or with errors) to the template
    }
    return render(request, 'dashboard/gym.html', context)


@login_required
def gym_report(request):
    """Render the Gym Report page."""
    # Sample gym visit data - in a real implementation, this would come from the database
    gym_visits = [
        {
            'id': '001',
            'customer_id': 'MEM001',
            'name': 'John Smith',
            'date_time': '2024-01-15 08:30 AM',
            'admin': 'Admin A'
        },
        {
            'id': '002',
            'customer_id': 'MEM002',
            'name': 'Sarah Johnson',
            'date_time': '2024-01-15 09:15 AM',
            'admin': 'Admin B'
        },
        {
            'id': '003',
            'customer_id': 'MEM003',
            'name': 'Mike Davis',
            'date_time': '2024-01-15 10:00 AM',
            'admin': 'Admin A'
        },
        {
            'id': '004',
            'customer_id': 'MEM004',
            'name': 'Emily Wilson',
            'date_time': '2024-01-15 11:30 AM',
            'admin': 'Admin C'
        },
        {
            'id': '005',
            'customer_id': 'MEM005',
            'name': 'David Brown',
            'date_time': '2024-01-15 02:15 PM',
            'admin': 'Admin B'
        },
        {
            'id': '006',
            'customer_id': 'MEM006',
            'name': 'Lisa Anderson',
            'date_time': '2024-01-15 03:45 PM',
            'admin': 'Admin A'
        },
        {
            'id': '007',
            'customer_id': 'MEM007',
            'name': 'Robert Taylor',
            'date_time': '2024-01-15 05:20 PM',
            'admin': 'Admin C'
        },
        {
            'id': '008',
            'customer_id': 'MEM008',
            'name': 'Jennifer Lee',
            'date_time': '2024-01-15 06:00 PM',
            'admin': 'Admin B'
        }
    ]
    
    context = {
        'gym_visits': gym_visits,
        'total_visits': 24,  # Total number of gym visits
        'page_size': 10,     # Number of visits per page
        'current_page': 1    # Current page number
    }
    return render(request, 'dashboard/gym_report.html', context)


@login_required
@require_permission([ADMINS_GROUP])
def export_user_data(request):
    """Export all user-related data (departments, users, groups, profiles)"""
    try:
        format = request.GET.get('format', 'json').lower()
        response = create_export_file(format)
        return response
    except Exception as e:
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f"Error exporting user data: {str(e)}")
        logger.error(traceback.format_exc())
        return JsonResponse({'error': f'Failed to export data: {str(e)}'}, status=500)


@login_required
@require_permission([ADMINS_GROUP])
@csrf_exempt
def import_user_data(request):
    """Import user-related data from a JSON file"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST method required'}, status=405)
    
    try:
        # Get the uploaded file
        if 'file' not in request.FILES:
            return JsonResponse({'error': 'No file uploaded'}, status=400)
        
        uploaded_file = request.FILES['file']
        
        # Check file extension
        if not uploaded_file.name.endswith('.json'):
            return JsonResponse({'error': 'Only JSON files are supported'}, status=400)
        
        # Read and parse the JSON data
        try:
            file_content = uploaded_file.read().decode('utf-8')
            data = json.loads(file_content)
        except json.JSONDecodeError as e:
            return JsonResponse({'error': f'Invalid JSON format: {str(e)}'}, status=400)
        
        # Import the data
        result = import_all_data(data)
        
        return JsonResponse({
            'success': True,
            'message': 'Data imported successfully',
            'result': result
        })
        
    except Exception as e:
        logger = logging.getLogger(__name__)
        logger.error(f"Error importing user data: {str(e)}")
        return JsonResponse({'error': f'Failed to import data: {str(e)}'}, status=500)
