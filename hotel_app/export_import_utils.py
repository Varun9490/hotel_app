import json
import csv
from io import StringIO
from django.http import HttpResponse
from django.contrib.auth import get_user_model
from .models import Department, UserGroup, UserProfile, UserGroupMembership
from django.core.exceptions import ValidationError
import logging
from django.db import transaction
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)
User = get_user_model()


def export_departments():
    """Export all departments as JSON"""
    departments = Department.objects.all().values(
        'id', 'name', 'description'
    )
    return list(departments)


def export_user_groups():
    """Export all user groups as JSON"""
    groups = UserGroup.objects.all().values(
        'id', 'name', 'description', 'department_id'
    )
    return list(groups)


def export_users():
    """Export all users as JSON"""
    users = User.objects.all().values(
        'id', 'username', 'email', 'first_name', 'last_name', 'is_active', 'is_staff', 'is_superuser', 'date_joined'
    )
    return list(users)


def export_user_profiles():
    """Export all user profiles as JSON"""
    profiles = UserProfile.objects.all().values(
        'user_id', 'full_name', 'phone', 'title', 'department_id', 'avatar_url',
        'enabled', 'timezone', 'preferences', 'role', 'created_at', 'updated_at'
    )
    return list(profiles)


def export_user_group_memberships():
    """Export all user group memberships as JSON"""
    memberships = UserGroupMembership.objects.all().values(
        'user_id', 'group_id', 'joined_at'
    )
    return list(memberships)


def export_all_data():
    """Export all user-related data as a single JSON object"""
    data = {
        'departments': export_departments(),
        'user_groups': export_user_groups(),
        'users': export_users(),
        'user_profiles': export_user_profiles(),
        'user_group_memberships': export_user_group_memberships()
    }
    return data


def create_export_file(format='json'):
    """Create an export file in the specified format"""
    data = export_all_data()
    
    if format.lower() == 'csv':
        return create_csv_export(data)
    elif format.lower() == 'xlsx':
        return create_xlsx_export(data)
    else:
        return create_json_export(data)


def create_json_export(data):
    """Create a JSON export file"""
    json_data = json.dumps(data, indent=2, default=str)
    
    response = HttpResponse(json_data, content_type='application/json')
    response['Content-Disposition'] = 'attachment; filename="user_data_export.json"'
    return response


def create_csv_export(data):
    """Create a CSV export file with all data"""
    output = StringIO()
    writer = csv.writer(output)
    
    # Write departments
    writer.writerow(['DEPARTMENTS'])
    writer.writerow(['id', 'name', 'description'])
    for dept in data['departments']:
        writer.writerow([dept['id'], dept['name'], dept['description']])
    
    writer.writerow([])  # Empty row as separator
    
    # Write user groups
    writer.writerow(['USER_GROUPS'])
    writer.writerow(['id', 'name', 'description', 'department_id'])
    for group in data['user_groups']:
        writer.writerow([group['id'], group['name'], group['description'], group['department_id']])
    
    writer.writerow([])  # Empty row as separator
    
    # Write users
    writer.writerow(['USERS'])
    writer.writerow(['id', 'username', 'email', 'first_name', 'last_name', 'is_active', 'is_staff', 'is_superuser', 'date_joined'])
    for user in data['users']:
        writer.writerow([user['id'], user['username'], user['email'], user['first_name'], user['last_name'], 
                         user['is_active'], user['is_staff'], user['is_superuser'], user['date_joined']])
    
    writer.writerow([])  # Empty row as separator
    
    # Write user profiles
    writer.writerow(['USER_PROFILES'])
    writer.writerow(['user_id', 'full_name', 'phone', 'title', 'department_id', 'avatar_url',
                     'enabled', 'timezone', 'role', 'created_at', 'updated_at'])
    for profile in data['user_profiles']:
        writer.writerow([profile['user_id'], profile['full_name'], profile['phone'], profile['title'], 
                         profile['department_id'], profile['avatar_url'], profile['enabled'], 
                         profile['timezone'], profile['role'], profile['created_at'], profile['updated_at']])
    
    writer.writerow([])  # Empty row as separator
    
    # Write user group memberships
    writer.writerow(['USER_GROUP_MEMBERSHIPS'])
    writer.writerow(['user_id', 'group_id', 'joined_at'])
    for membership in data['user_group_memberships']:
        writer.writerow([membership['user_id'], membership['group_id'], membership['joined_at']])
    
    csv_data = output.getvalue()
    output.close()
    
    response = HttpResponse(csv_data, content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="user_data_export.csv"'
    return response


def create_xlsx_export(data):
    """Create an Excel (XLSX) export file with all data"""
    wb = Workbook()
    
    # Remove the default sheet
    wb.remove(wb.active)
    
    # Create departments sheet
    ws_dept = wb.create_sheet("Departments")
    ws_dept.append(['id', 'name', 'description'])
    for dept in data['departments']:
        ws_dept.append([dept['id'], dept['name'], dept['description']])
    
    # Create user groups sheet
    ws_groups = wb.create_sheet("User Groups")
    ws_groups.append(['id', 'name', 'description', 'department_id'])
    for group in data['user_groups']:
        ws_groups.append([group['id'], group['name'], group['description'], group['department_id']])
    
    # Create users sheet
    ws_users = wb.create_sheet("Users")
    ws_users.append(['id', 'username', 'email', 'first_name', 'last_name', 'is_active', 'is_staff', 'is_superuser', 'date_joined'])
    for user in data['users']:
        ws_users.append([user['id'], user['username'], user['email'], user['first_name'], user['last_name'], 
                         user['is_active'], user['is_staff'], user['is_superuser'], user['date_joined']])
    
    # Create user profiles sheet
    ws_profiles = wb.create_sheet("User Profiles")
    ws_profiles.append(['user_id', 'full_name', 'phone', 'title', 'department_id', 'avatar_url',
                        'enabled', 'timezone', 'role', 'created_at', 'updated_at'])
    for profile in data['user_profiles']:
        ws_profiles.append([profile['user_id'], profile['full_name'], profile['phone'], profile['title'], 
                            profile['department_id'], profile['avatar_url'], profile['enabled'], 
                            profile['timezone'], profile['role'], profile['created_at'], profile['updated_at']])
    
    # Create user group memberships sheet
    ws_memberships = wb.create_sheet("User Group Memberships")
    ws_memberships.append(['user_id', 'group_id', 'joined_at'])
    for membership in data['user_group_memberships']:
        ws_memberships.append([membership['user_id'], membership['group_id'], membership['joined_at']])
    
    # Auto-adjust column widths
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to response
    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="user_data_export.xlsx"'
    return response


def validate_import_data(data):
    """Validate the structure of import data"""
    required_keys = ['departments', 'user_groups', 'users', 'user_profiles', 'user_group_memberships']
    
    if not isinstance(data, dict):
        raise ValidationError("Invalid data format. Expected a dictionary.")
    
    for key in required_keys:
        if key not in data:
            raise ValidationError(f"Missing required key: {key}")
        
        # Check that each key contains a list
        if not isinstance(data[key], list):
            raise ValidationError(f"Invalid format for {key}. Expected a list.")
    
    # Validate departments
    for dept in data['departments']:
        if not isinstance(dept, dict):
            raise ValidationError("Invalid department format. Expected a dictionary.")
        if 'name' not in dept:
            raise ValidationError("Missing required field 'name' in department data.")
    
    # Validate user groups
    for group in data['user_groups']:
        if not isinstance(group, dict):
            raise ValidationError("Invalid user group format. Expected a dictionary.")
        if 'name' not in group:
            raise ValidationError("Missing required field 'name' in user group data.")
    
    # Validate users
    for user in data['users']:
        if not isinstance(user, dict):
            raise ValidationError("Invalid user format. Expected a dictionary.")
        if 'username' not in user or 'email' not in user:
            raise ValidationError("Missing required fields in user data. Username and email are required.")
    
    # Validate user profiles
    for profile in data['user_profiles']:
        if not isinstance(profile, dict):
            raise ValidationError("Invalid user profile format. Expected a dictionary.")
        if 'user_id' not in profile:
            raise ValidationError("Missing required field 'user_id' in user profile data.")
    
    # Validate user group memberships
    for membership in data['user_group_memberships']:
        if not isinstance(membership, dict):
            raise ValidationError("Invalid user group membership format. Expected a dictionary.")
        if 'user_id' not in membership or 'group_id' not in membership:
            raise ValidationError("Missing required fields in user group membership data. Both user_id and group_id are required.")
    
    return True


@transaction.atomic
def import_departments(departments_data):
    """Import departments from JSON data"""
    created_count = 0
    updated_count = 0
    
    for dept_data in departments_data:
        dept_id = dept_data.pop('id', None)
        if dept_id:
            # Try to update existing department
            try:
                dept = Department.objects.get(id=dept_id)
                for key, value in dept_data.items():
                    setattr(dept, key, value)
                dept.save()
                updated_count += 1
            except Department.DoesNotExist:
                # Create new department
                dept = Department.objects.create(id=dept_id, **dept_data)
                created_count += 1
        else:
            # Create new department without specifying ID
            Department.objects.create(**dept_data)
            created_count += 1
    
    return created_count, updated_count


@transaction.atomic
def import_user_groups(groups_data):
    """Import user groups from JSON data"""
    created_count = 0
    updated_count = 0
    
    for group_data in groups_data:
        group_id = group_data.pop('id', None)
        # Handle department reference
        department_id = group_data.pop('department_id', None)
        
        if group_id:
            # Try to update existing group
            try:
                group = UserGroup.objects.get(id=group_id)
                # Handle department reference
                if department_id:
                    try:
                        department = Department.objects.get(id=department_id)
                        group.department = department
                    except Department.DoesNotExist:
                        group.department = None
                        logger.warning(f"Department with ID {department_id} not found for group {group_id}. Setting to None.")
                
                for key, value in group_data.items():
                    setattr(group, key, value)
                group.save()
                updated_count += 1
            except UserGroup.DoesNotExist:
                # Create new group
                group_data['id'] = group_id
                # Handle department reference
                if department_id:
                    try:
                        department = Department.objects.get(id=department_id)
                        group_data['department'] = department
                    except Department.DoesNotExist:
                        group_data['department'] = None
                        logger.warning(f"Department with ID {department_id} not found for new group {group_id}. Setting to None.")
                
                group = UserGroup.objects.create(**group_data)
                created_count += 1
        else:
            # Create new group without specifying ID
            # Handle department reference
            if department_id:
                try:
                    department = Department.objects.get(id=department_id)
                    group_data['department'] = department
                except Department.DoesNotExist:
                    group_data['department'] = None
                    logger.warning(f"Department with ID {department_id} not found for new group. Setting to None.")
            
            UserGroup.objects.create(**group_data)
            created_count += 1
    
    return created_count, updated_count


@transaction.atomic
def import_users(users_data):
    """Import users from JSON data"""
    created_count = 0
    updated_count = 0
    
    for user_data in users_data:
        user_id = user_data.pop('id', None)
        # Remove sensitive fields
        user_data.pop('password', None)
        
        if user_id:
            # Try to update existing user
            try:
                user = User.objects.get(id=user_id)
                for key, value in user_data.items():
                    setattr(user, key, value)
                user.save()
                updated_count += 1
            except User.DoesNotExist:
                # Create new user
                user_data['id'] = user_id
                User.objects.create(**user_data)
                created_count += 1
        else:
            # Create new user without specifying ID
            User.objects.create(**user_data)
            created_count += 1
    
    return created_count, updated_count


@transaction.atomic
def import_user_profiles(profiles_data):
    """Import user profiles from JSON data"""
    created_count = 0
    updated_count = 0
    
    for profile_data in profiles_data:
        user_id = profile_data.pop('user_id', None)
        # Handle department reference
        department_id = profile_data.pop('department_id', None)
        
        if user_id:
            # Try to update existing profile
            try:
                profile = UserProfile.objects.get(user_id=user_id)
                # Handle department reference
                if department_id:
                    try:
                        department = Department.objects.get(id=department_id)
                        profile.department = department
                    except Department.DoesNotExist:
                        profile.department = None
                        logger.warning(f"Department with ID {department_id} not found for profile of user {user_id}. Setting to None.")
                
                for key, value in profile_data.items():
                    setattr(profile, key, value)
                profile.save()
                updated_count += 1
            except UserProfile.DoesNotExist:
                # Create new profile
                try:
                    user = User.objects.get(id=user_id)
                    profile_data['user'] = user
                    # Handle department reference
                    if department_id:
                        try:
                            department = Department.objects.get(id=department_id)
                            profile_data['department'] = department
                        except Department.DoesNotExist:
                            profile_data['department'] = None
                            logger.warning(f"Department with ID {department_id} not found for new profile of user {user_id}. Setting to None.")
                    
                    UserProfile.objects.create(**profile_data)
                    created_count += 1
                except User.DoesNotExist:
                    logger.warning(f"User with ID {user_id} not found. Skipping profile creation.")
        else:
            logger.warning("Profile data missing user_id. Skipping profile creation.")
    
    return created_count, updated_count


@transaction.atomic
def import_user_group_memberships(memberships_data):
    """Import user group memberships from JSON data"""
    created_count = 0
    skipped_count = 0
    
    for membership_data in memberships_data:
        user_id = membership_data.get('user_id')
        group_id = membership_data.get('group_id')
        
        if user_id and group_id:
            try:
                user = User.objects.get(id=user_id)
                group = UserGroup.objects.get(id=group_id)
                
                # Create membership if it doesn't exist
                membership, created = UserGroupMembership.objects.get_or_create(
                    user=user,
                    group=group,
                    defaults={'joined_at': membership_data.get('joined_at')}
                )
                
                if created:
                    created_count += 1
            except (User.DoesNotExist, UserGroup.DoesNotExist):
                skipped_count += 1
                logger.warning(f"User ({user_id}) or Group ({group_id}) not found. Skipping membership creation.")
        else:
            skipped_count += 1
            logger.warning("Membership data missing user_id or group_id. Skipping membership creation.")
    
    return created_count, skipped_count


@transaction.atomic
def import_all_data(data):
    """Import all user-related data from a JSON object"""
    validate_import_data(data)
    
    # Import in the correct order to maintain relationships
    dept_created, dept_updated = import_departments(data['departments'])
    group_created, group_updated = import_user_groups(data['user_groups'])
    user_created, user_updated = import_users(data['users'])
    profile_created, profile_updated = import_user_profiles(data['user_profiles'])
    membership_created, membership_skipped = import_user_group_memberships(data['user_group_memberships'])
    
    return {
        'departments': {'created': dept_created, 'updated': dept_updated},
        'user_groups': {'created': group_created, 'updated': group_updated},
        'users': {'created': user_created, 'updated': user_updated},
        'user_profiles': {'created': profile_created, 'updated': profile_updated},
        'user_group_memberships': {'created': membership_created, 'skipped': membership_skipped}
    }